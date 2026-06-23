from email.policy import default
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.http import JsonResponse
from django.template import TemplateDoesNotExist
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Q
import json
import secrets
from datetime import datetime, time

from .models import *
from django.core.paginator import Paginator
from django.db.models import Q

# ============== AUTHENTICATION VIEWS ==============



def manager_register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        phone = request.POST.get('phone')
        branch_location = request.POST.get('branch_location')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        if password != confirm_password:
            messages.error(request, 'Passwords do not match!')
            return redirect('manager_register')
        
        if len(password) < 8:
            messages.error(request, 'Password must be at least 8 characters long!')
            return redirect('manager_register')
        
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return redirect('manager_register')
        
        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists!')
            return redirect('manager_register')
        
        # Create user with branch_location directly
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=password,
            user_type='manager',
            phone=phone,
            is_first_login=False,
            first_name=first_name,
            last_name=last_name,
            branch_location=branch_location or 'Head Office'  # Set default if empty
        )
        
        messages.success(request, f'Manager account created successfully for {first_name} {last_name}! Please login.')
        return redirect('login')
    
    return render(request, 'auth/manager_register.html')
      
def user_login(request):
    if request.user.is_authenticated:
        if request.user.user_type == 'manager':
            return redirect('manager_dashboard')
        else:
            return redirect('employee_dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            if user.user_type == 'employee' and user.is_first_login:
                messages.info(request, 'This is your first login. Please change your password to continue.')
                return redirect('change_password')
            
            if user.user_type == 'manager':
                return redirect('manager_dashboard')
            else:
                return redirect('employee_dashboard')
        else:
            messages.error(request, 'Invalid username or password!')
    
    return render(request, 'auth/login.html')

def user_logout(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully!')
    return redirect('login')
@login_required
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        user = request.user
        
        if not old_password or not new_password or not confirm_password:
            messages.error(request, 'All fields are required!')
            return redirect('change_password')
        
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match!')
            return redirect('change_password')
        
        if old_password == new_password:
            messages.error(request, 'New password cannot be same as old password!')
            return redirect('change_password')
        
        if not user.check_password(old_password):
            messages.error(request, 'Current password is incorrect!')
            return redirect('change_password')
        
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            for error in e.messages:
                messages.error(request, error)
            return redirect('change_password')
        
        user.set_password(new_password)
        user.is_first_login = False
        user.save()
        update_session_auth_hash(request, user)  
        
        messages.success(request, 'Password changed successfully! You can now access all features.')
        
        if user.user_type == 'manager':
            return redirect('manager_dashboard')
        else:
            return redirect('employee_dashboard')
    
    is_first_login = request.user.is_authenticated and request.user.user_type == 'employee' and request.user.is_first_login
    
    context = {
        'is_first_login': is_first_login,
    }
    return render(request, 'auth/change_password.html', context)
# ============== MANAGER VIEWS ==============
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Q, Sum
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import CustomUser, Project, Hardware, HardwareAssignment, HardwareSerialEntry, HardwareType

@login_required
def manager_dashboard(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    # Get data with optimized queries
    employees = CustomUser.objects.filter(user_type='employee', manager=request.user)
    projects = Project.objects.filter(created_by=request.user)
    hardware_assignments = HardwareAssignment.objects.filter(assigned_by=request.user)
    hardware_items = Hardware.objects.filter(created_by=request.user)
    
    # Hardware status counts
    available_count = hardware_items.filter(status='available').count()
    assigned_count = hardware_items.filter(status='assigned').count()
    in_use_count = hardware_items.filter(status='in_use').count()
    maintenance_count = hardware_items.filter(status='maintenance').count()
    total_hardware = hardware_items.count()
    
    # Calculate hardware by type
    hardware_by_type = {}
    for hw in hardware_items.select_related('hardware_type'):
        type_name = hw.hardware_type.name if hw.hardware_type else 'Unknown'
        hardware_by_type[type_name] = hardware_by_type.get(type_name, 0) + 1
    
    # Calculate chart height values
    max_count = max(available_count, assigned_count, in_use_count, maintenance_count) if total_hardware > 0 else 100
    chart_max = int(max_count * 1.1) + 5
    
    scale_values = [
        chart_max,
        int(chart_max * 0.75),
        int(chart_max * 0.5),
        int(chart_max * 0.25),
        0
    ]
    
    hardware_status = [
        {
            'name': 'Available',
            'count': available_count,
            'percentage': (available_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((available_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'success'
        },
        {
            'name': 'Assigned',
            'count': assigned_count,
            'percentage': (assigned_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((assigned_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'warning'
        },
        {
            'name': 'In Use',
            'count': in_use_count,
            'percentage': (in_use_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((in_use_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'info'
        },
        {
            'name': 'Maintenance',
            'count': maintenance_count,
            'percentage': (maintenance_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((maintenance_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'danger'
        },
    ]
    
    # Recent activities aggregation
    recent_activities = []
    
    # Recent assignments (last 5)
    recent_assignments = hardware_assignments.select_related('employee', 'project').order_by('-assigned_date')[:5]
    for assignment in recent_assignments:
        recent_activities.append({
            'icon': 'clipboard-check',
            'color': 'primary',
            'title': 'New hardware assignment',
            'description': f'{assignment.employee.get_full_name() or assignment.employee.username} → {assignment.project.project_name}',
            'timestamp': assignment.assigned_date,
            'type': 'assignment',
            'badge': 'New'
        })
    
    # Recent employees (last 5)
    recent_employees = employees.select_related('manager').order_by('-date_joined')[:5]
    for employee in recent_employees:
        recent_activities.append({
            'icon': 'person-plus',
            'color': 'success',
            'title': 'New employee joined',
            'description': f'{employee.get_full_name() or employee.username} - {employee.email}',
            'timestamp': employee.date_joined,
            'type': 'employee',
            'badge': 'New'
        })
    
    # Recent hardware verifications
    try:
        recent_verifications = HardwareSerialEntry.objects.filter(
            assignment_item__assignment__assigned_by=request.user,
            verified=True
        ).select_related('assignment_item__hardware', 'assignment_item__assignment__employee').order_by('-verified_at')[:5]
        
        for verification in recent_verifications:
            recent_activities.append({
                'icon': 'shield-check',
                'color': 'info',
                'title': 'Hardware verified',
                'description': f'{verification.assignment_item.hardware.hardware_type.name} - {verification.assignment_item.assignment.employee.get_full_name()}',
                'timestamp': verification.verified_at,
                'type': 'verification',
                'badge': 'Verified'
            })
    except:
        pass
    
    # Recent projects
    recent_projects = projects.order_by('-created_at')[:5]
    for project in recent_projects:
        recent_activities.append({
            'icon': 'folder-plus',
            'color': 'warning',
            'title': 'New project created',
            'description': f'{project.project_name} ({project.project_id})',
            'timestamp': project.created_at,
            'type': 'project',
            'badge': 'New'
        })
    
    # Sort activities by timestamp (most recent first)
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:8]  # Show latest 8 activities
    
    # Calculate growth percentages (30 days comparison)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    # Employee growth
    current_employees = employees.count()
    last_month_employees = employees.filter(date_joined__lt=thirty_days_ago).count()
    if last_month_employees > 0:
        employee_growth = ((current_employees - last_month_employees) / last_month_employees * 100)
    else:
        employee_growth = 0 if current_employees == 0 else 100
    
    # Project growth
    current_projects = projects.count()
    last_month_projects = projects.filter(created_at__lt=thirty_days_ago).count()
    if last_month_projects > 0:
        project_growth = ((current_projects - last_month_projects) / last_month_projects * 100)
    else:
        project_growth = 0 if current_projects == 0 else 100
    
    # Assignment growth
    current_assignments = hardware_assignments.filter(actual_return_date__isnull=True).count()
    last_month_assignments = hardware_assignments.filter(assigned_date__lt=thirty_days_ago).count()
    if last_month_assignments > 0:
        assignment_growth = ((current_assignments - last_month_assignments) / last_month_assignments * 100)
    else:
        assignment_growth = 0 if current_assignments == 0 else 100
    
    # Hardware growth
    current_hardware = total_hardware
    last_month_hardware = hardware_items.filter(created_at__lt=thirty_days_ago).count()
    if last_month_hardware > 0:
        hardware_growth = ((current_hardware - last_month_hardware) / last_month_hardware * 100)
    else:
        hardware_growth = 0 if current_hardware == 0 else 100
    
    # Get assignments with employee and project details
    assignments_list = hardware_assignments.select_related('employee', 'project').order_by('-assigned_date')[:10]
    
    # Get employees list for recent employees table
    employees_list = employees.select_related('manager').order_by('-date_joined')[:10]
    
    # Calculate verification stats
    total_verifications = HardwareSerialEntry.objects.filter(
        assignment_item__assignment__assigned_by=request.user
    ).count()
    
    verified_count = HardwareSerialEntry.objects.filter(
        assignment_item__assignment__assigned_by=request.user,
        verified=True
    ).count()
    
    pending_verifications = total_verifications - verified_count
    
    verification_rate = int((verified_count / total_verifications * 100)) if total_verifications > 0 else 0
    
    # Get hardware types distribution for chart (FIXED: removed created_by filter)
    hardware_type_data = []
    for hw_type in HardwareType.objects.all():  # Get all hardware types
        count = hardware_items.filter(hardware_type=hw_type).count()
        if count > 0:
            hardware_type_data.append({
                'name': hw_type.name,
                'count': count,
                'percentage': int((count / total_hardware * 100)) if total_hardware > 0 else 0
            })
    
    context = {
        # Statistics
        'total_employees': current_employees,
        'total_projects': current_projects,
        'active_assignments': current_assignments,
        'hardware_count': total_hardware,
        
        # Hardware status
        'hardware_status': hardware_status,
        'available_count': available_count,
        'assigned_count': assigned_count,
        'in_use_count': in_use_count,
        'maintenance_count': maintenance_count,
        
        # Hardware by type
        'hardware_by_type': hardware_by_type,
        'hardware_type_data': hardware_type_data,
        
        # Lists
        'employees': employees_list,
        'projects': projects[:10],
        'assignments': assignments_list,
        'recent_activities': recent_activities,
        
        # Growth percentages
        'employee_growth': round(employee_growth, 1),
        'project_growth': round(project_growth, 1),
        'assignment_growth': round(assignment_growth, 1),
        'hardware_growth': round(hardware_growth, 1),
        
        # Chart settings
        'chart_max': chart_max,
        'scale_values': scale_values,
        'max_count': max_count,
        
        # Verification stats
        'verification_rate': verification_rate,
        'pending_verifications': pending_verifications,
        'verified_count': verified_count,
        
        # Date
        'today': timezone.now().date(),
        'current_time': timezone.now(),
    }
    return render(request, 'manager/dashboard.html', context)

from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.hashers import make_password

# Default password constant
DEFAULT_PASSWORD = 'Eduquity@2024'  # Change this to your desired default password


from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.hashers import make_password

# Default password constant
DEFAULT_PASSWORD = 'Eduquity@2024'  # Change this to your desired default password

@login_required
def create_employee(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    manager_branch = getattr(request.user, 'branch_location', None) or 'Head Office'
 
    
    if request.method == 'POST':
        name = request.POST.get('name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        
        if not name or len(name.strip()) < 2:
            messages.error(request, 'Please enter a valid name!')
            return redirect('create_employee')
        
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return redirect('create_employee')
        
        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists!')
            return redirect('create_employee')
        
        # Use DEFAULT_PASSWORD
        default_password = DEFAULT_PASSWORD
        
        name_parts = name.strip().split(' ', 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ''
        
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=default_password,
            user_type='employee',
            manager=request.user,
            phone=phone,
            is_first_login=False,  # Set to False since no password change needed
            first_name=first_name,
            last_name=last_name,
            branch_location=manager_branch, # Set employee's branch to match manager's branch

        )
        
        try:
            send_mail(
                subject='Your Eduquity Hardware Management Account Credentials',
                message=f'''
Dear {name},

Welcome to Eduquity Hardware Management System!

Your account has been created successfully. Here are your login credentials:

Username: {username}
Password: {default_password}
Login URL: http://eduquityinventory.co.in/login/

Important Instructions:
1. Use the password above to login
2. Keep your credentials secure
3. Do not share your password with anyone

Best regards,
Eduquity Hardware Management Team
                ''',
                html_message=f'''
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; padding: 20px; text-align: center; border-radius: 5px 5px 0 0; }}
        .content {{ background: #f9f9f9; padding: 30px; border: 1px solid #ddd; border-top: none; border-radius: 0 0 5px 5px; }}
        .credentials {{ background: #e8f4fc; border: 2px solid #3498db; padding: 15px; margin: 20px 0; border-radius: 5px; }}
        .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 12px; }}
        .btn {{ display: inline-block; padding: 10px 20px; background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; text-decoration: none; border-radius: 5px; margin: 10px 0; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h2>Welcome to Eduquity Hardware Management</h2>
        </div>
        <div class="content">
            <p>Dear <strong>{name}</strong>,</p>
            
            <p>Your account has been created successfully in the Eduquity Hardware Management System.</p>
            
            <div class="credentials">
                <h3>Your Login Credentials:</h3>
                <p><strong>Full Name:</strong> {name}</p>
                <p><strong>Username:</strong> {username}</p>
                <p><strong>Password:</strong> <code style="background: #fff; padding: 5px 10px; border-radius: 3px; font-size: 14px;">{default_password}</code></p>
                <p><strong>Login URL:</strong> <a href="http://eduquityinventory.co.in/login/">http://eduquityinventory.co.in/login/</a></p>
                <a href="http://eduquityinventory.co.in/login/" class="btn">Login Now</a>
            </div>
            
            <p><strong>About the System:</strong><br>
            The Eduquity Hardware Management System allows you to:
            <ul>
                <li>View your hardware assignments</li>
                <li>Enter serial numbers of assigned hardware</li>
                <li>Track hardware status</li>
                <li>Communicate with your manager</li>
            </ul>
            </p>
            
            <div class="footer">
                <p><strong>Eduquity Hardware Management Team</strong><br>
                Established in 2000 - Thought-leader in the Indian assessment industry</p>
                <p><em>This is an automated email. Please do not reply to this message.</em></p>
            </div>
        </div>
    </div>
</body>
</html>
                ''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
            
            messages.success(request, f'Employee account created successfully for {name}! Login credentials have been sent to {email}.')
            
        except Exception as e:
            messages.warning(request, f'Employee account created for {name} but email could not be sent. Error: {str(e)}')
        
        return redirect('create_employee')
    
    all_employees = CustomUser.objects.filter(
        user_type='employee', 
        manager=request.user
    ).order_by('-date_joined')
    
    recent_employees = all_employees[:5]
    
    total_employees = all_employees.count()
    active_employees = all_employees.filter(is_active=True).count()
    pending_employees = all_employees.filter(is_active=False).count()
    
    context = {
        'recent_employees': recent_employees,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'pending_employees': pending_employees,
        'default_password': DEFAULT_PASSWORD,
        'manager_branch': manager_branch,  # Add manager branch to context

    }
    return render(request, 'manager/create_employee.html', context)

import io
import csv
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.hashers import make_password

@login_required
def bulk_create_employees(request):
    """Bulk create employees via CSV upload with default password"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    manager_branch = getattr(request.user, 'branch_location', None) or 'Head Office'

    
    if request.method == 'POST':
        csv_file = request.FILES.get('csv_file')
        
        if not csv_file:
            messages.error(request, 'Please select a CSV file to upload!')
            return redirect('bulk_create_employees')
        
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a valid CSV file!')
            return redirect('bulk_create_employees')
        
        try:
            decoded_file = csv_file.read().decode('utf-8-sig')
            io_string = io.StringIO(decoded_file)
            
            dialect = csv.Sniffer().sniff(decoded_file[:1024])
            reader = csv.DictReader(io_string, dialect=dialect)
            reader.fieldnames = [name.strip() for name in reader.fieldnames]
            
            created_count = 0
            error_count = 0
            errors = []
            created_employees = []
            
            default_password = DEFAULT_PASSWORD
            
            for row_num, row in enumerate(reader, start=2):
                name = row.get('name', '').strip()
                email = row.get('email', '').strip()
                phone = row.get('phone', '').strip()
                username = row.get('username', '').strip()
                
                if not username and email:
                    username = email.split('@')[0].lower()
                    username = username.replace('.', '_').replace('-', '_').replace('@', '_')
                
                if not name or len(name.strip()) < 2:
                    errors.append(f"Row {row_num}: Name is required and must be at least 2 characters")
                    error_count += 1
                    continue
                
                if not email:
                    errors.append(f"Row {row_num}: Email is required")
                    error_count += 1
                    continue
                
                if CustomUser.objects.filter(username=username).exists():
                    errors.append(f"Row {row_num}: Username '{username}' already exists")
                    error_count += 1
                    continue
                
                if CustomUser.objects.filter(email=email).exists():
                    errors.append(f"Row {row_num}: Email '{email}' already exists")
                    error_count += 1
                    continue
                
                name_parts = name.strip().split(' ', 1)
                first_name = name_parts[0]
                last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                try:
                    user = CustomUser.objects.create_user(
                        username=username,
                        email=email,
                        password=default_password,
                        user_type='employee',
                        manager=request.user,
                        phone=phone,
                        is_first_login=False,  # No password change needed
                        first_name=first_name,
                        last_name=last_name,
                        branch_location=manager_branch,  # Set employee's branch to match manager's branch

                    )
                    
                    created_employees.append({
                        'name': name,
                        'username': username,
                        'email': email,
                        'password': default_password
                    })
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1
            
            if created_employees:
                send_bulk_welcome_emails(created_employees, request.user)
            
            if created_count > 0:
                messages.success(request, f'Successfully created {created_count} employee(s)!')
                messages.info(request, f'Default password for all employees: {default_password}')
            
            if errors:
                messages.warning(request, f'Created {created_count} employee(s). {error_count} error(s):')
                for error in errors[:10]:
                    messages.warning(request, error)
            
            return redirect('bulk_create_employees')
            
        except Exception as e:
            messages.error(request, f'Error reading CSV file: {str(e)}')
            return redirect('bulk_create_employees')
    
    employee_count = CustomUser.objects.filter(
        user_type='employee',
        manager=request.user
    ).count()
    
    context = {
        'employee_count': employee_count,
        'sample_csv': sample_csv_template(),
        'default_password': DEFAULT_PASSWORD,
        'manager_branch': manager_branch,

    }
    return render(request, 'manager/bulk_create_employees.html', context)



def send_bulk_welcome_emails(employees, manager):
    """Send welcome emails to multiple employees with default password"""
    from django.core.mail import send_mail
    from django.conf import settings
    
    default_password = DEFAULT_PASSWORD
    
    for emp in employees:
        try:
            name = emp.get('name', 'Employee')
            username = emp.get('username', 'user')
            email = emp.get('email', '')
            
            if not email:
                print(f"Warning: No email provided for {name}")
                continue
                
            send_mail(
                subject='Your Eduquity Hardware Management Account Credentials',
                message=f'''
Dear {name},

Welcome to Eduquity Hardware Management System!

Your account has been created successfully. Here are your login credentials:

Username: {username}
Password: {default_password}
Login URL: http://eduquityinventory.co.in/login/

Important Instructions:
1. Use the password above to login
2. Keep your credentials secure
3. Do not share your password with anyone

Best regards,
Eduquity Hardware Management Team
                ''',
                html_message=f'''
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(90deg, #E04D00 0%, #FF6B1A 100%); color: white; padding: 20px; text-align: center; border-radius: 5px 5px 0 0; }}
        .content {{ background: #f9f9f9; padding: 30px; border: 1px solid #ddd; border-top: none; border-radius: 0 0 5px 5px; }}
        .credentials {{ background: #e8f4fc; border: 2px solid #E04D00; padding: 15px; margin: 20px 0; border-radius: 5px; }}
        .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 12px; }}
        .btn {{ display: inline-block; padding: 10px 20px; background: #E04D00; color: white; text-decoration: none; border-radius: 5px; margin: 10px 0; }}
        .btn:hover {{ background: #c44500; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h2>Welcome to Eduquity Hardware Management</h2>
        </div>
        <div class="content">
            <p>Dear <strong>{name}</strong>,</p>
            
            <p>Your account has been created successfully in the Eduquity Hardware Management System.</p>
            
            <div class="credentials">
                <h3>Your Login Credentials:</h3>
                <p><strong>Full Name:</strong> {name}</p>
                <p><strong>Username:</strong> {username}</p>
                <p><strong>Password:</strong> <code style="background: #fff; padding: 5px 10px; border-radius: 3px; font-size: 14px;">{default_password}</code></p>
                <p><strong>Login URL:</strong> <a href="http://eduquityinventory.co.in/login/">http://eduquityinventory.co.in/login/</a></p>
                <a href="http://eduquityinventory.co.in/login/" class="btn">Login Now</a>
            </div>
            
            <p><strong>About the System:</strong><br>
            The Eduquity Hardware Management System allows you to:
            <ul>
                <li>View your hardware assignments</li>
                <li>Enter serial numbers of assigned hardware</li>
                <li>Track hardware status</li>
                <li>Communicate with your manager</li>
            </ul>
            </p>
            
            <div class="footer">
                <p><strong>Eduquity Hardware Management Team</strong><br>
                Established in 2000 - Thought-leader in the Indian assessment industry</p>
                <p><em>This is an automated email. Please do not reply to this message.</em></p>
            </div>
        </div>
    </div>
</body>
</html>
                ''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )
        except Exception as e:
            print(f"Failed to send email to {emp.get('email', 'unknown')}: {str(e)}")




def sample_csv_template():
    """Generate sample CSV template content without spaces"""
    return """name,email,phone,username
John Doe,john.doe@example.com,+91 9876543210,john_doe
Jane Smith,jane.smith@example.com,+91 9876543211,jane_smith
Mike Johnson,mike.johnson@example.com,+91 9876543212,mike_j"""


@login_required
def download_sample_csv(request):
    """Download sample CSV template without spaces in headers"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employee_import_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['name', 'email', 'phone', 'username'])
    writer.writerow(['John Doe', 'john.doe@example.com', '+91 9876543210', 'john_doe'])
    writer.writerow(['Jane Smith', 'jane.smith@example.com', '+91 9876543211', 'jane_smith'])
    writer.writerow(['Mike Johnson', 'mike.johnson@example.com', '+91 9876543212', 'mike_j'])
    
    return response


@login_required
def delete_employee(request, employee_id):
    """Delete an employee account"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    employee = get_object_or_404(
        CustomUser, 
        id=employee_id, 
        user_type='employee', 
        manager=request.user
    )
    
    if request.method == 'POST':
        employee_name = employee.get_full_name() or employee.username
        
        active_assignments = HardwareAssignment.objects.filter(
            employee=employee,
            actual_return_date__isnull=True
        ).exists()
        
        if active_assignments:
            messages.error(
                request, 
                f'Cannot delete {employee_name} because they have active hardware assignments. Please return all hardware first.'
            )
            return redirect('create_employee')
        
        employee.delete()
        messages.success(request, f'Employee {employee_name} has been deleted successfully.')
        return redirect('create_employee')
    
    return redirect('create_employee')


@login_required
def employee_list(request):
    """View all employees with management options"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    employees = CustomUser.objects.filter(
        user_type='employee', 
        manager=request.user
    ).order_by('-date_joined')
    
    for emp in employees:
        emp.active_assignments = HardwareAssignment.objects.filter(
            employee=emp,
            actual_return_date__isnull=True
        ).count()
        emp.total_assignments = HardwareAssignment.objects.filter(
            employee=emp
        ).count()
    
    context = {
        'employees': employees,
        'total_employees': employees.count(),
        'active_employees': employees.filter(is_active=True).count(),
        'pending_employees': employees.filter(is_active=False).count(),
    }
    return render(request, 'manager/employee_list.html', context)

# Add this import at the top of your views.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

@login_required
def create_project(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        project_id = request.POST.get('project_id')
        project_name = request.POST.get('project_name')
        description = request.POST.get('description')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        location = request.POST.get('location')
        
        if Project.objects.filter(project_id=project_id).exists():
            messages.error(request, 'Project ID already exists!')
            return redirect('create_project')
        
        if Project.objects.filter(project_name=project_name).exists():
            messages.error(request, 'Project name already exists!')
            return redirect('create_project')
        
        Project.objects.create(
            project_id=project_id,
            project_name=project_name,
            description=description,
            start_date=start_date,
            end_date=end_date,
            location=location,
            created_by=request.user
        )
        
        messages.success(request, f'Project "{project_name}" created successfully!')
        return redirect('create_project')
    
    # Get all projects
    projects = Project.objects.filter(created_by=request.user).order_by('-created_at')
    
    total_projects = projects.count()
    active_projects = 0
    total_hardware_assigned = 0
    total_employees = set()
    
    for project in projects:
        # Get all active assignments for this project (not returned)
        assignments = HardwareAssignment.objects.filter(
            project=project,
            assigned_by=request.user,
            actual_return_date__isnull=True
        )
        
        # Count unique employees
        project.employee_count = assignments.values('employee').distinct().count()
        total_employees.add(project.employee_count)
        
        # Initialize counters
        project.total_hardware = 0
        project.active_hardware = 0
        project.assigned_hardware = 0
        project.available_hardware = 0
        hardware_by_type = {}
        employee_assignments = []
        
        for assignment in assignments:
            for item in assignment.hardwareassignmentitem_set.all():
                hardware = item.hardware
                project.total_hardware += 1
                
                # Count by hardware type
                hw_type = hardware.hardware_type.name if hardware.hardware_type else 'Unknown'
                hardware_by_type[hw_type] = hardware_by_type.get(hw_type, 0) + 1
                
                # Count by status
                if hardware.status == 'in_use':
                    project.active_hardware += 1
                elif hardware.status == 'assigned':
                    project.assigned_hardware += 1
                elif hardware.status == 'available':
                    project.available_hardware += 1
                
                # Get verification status
                verification_status = 'not_entered'
                verified = False
                try:
                    serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                    if serial_entry.verified:
                        verification_status = 'verified'
                        verified = True
                    else:
                        verification_status = 'pending'
                except HardwareSerialEntry.DoesNotExist:
                    pass
                
                # Add to employee assignments list
                employee_assignments.append({
                    'employee_name': assignment.employee.get_full_name() or assignment.employee.username,
                    'employee_email': assignment.employee.email,
                    'exam_city': assignment.exam_city,
                    'hardware_type': hw_type,
                    'serial_number': hardware.serial_number,
                    'model': hardware.model_name,
                    'status': hardware.status,
                    'verified': verified,
                    'verification_status': verification_status,
                    'assigned_date': assignment.assigned_date
                })
        
        project.hardware_by_type = hardware_by_type
        project.employee_assignments = employee_assignments
        project.assignments = assignments
        
        total_hardware_assigned += project.total_hardware
        
        if assignments.exists():
            active_projects += 1
    
    context = {
        'projects': projects,
        'total_projects': total_projects,
        'active_projects': active_projects,
        'total_hardware_assigned': total_hardware_assigned,
        'total_employees': len(total_employees),
    }
    return render(request, 'manager/create_project.html', context)


# @login_required
# def export_project_excel(request, project_id):
#     """Export project-wise hardware assignments to Excel"""
#     if request.user.user_type != 'manager':
#         return redirect('employee_dashboard')
    
#     # Get the project
#     project = get_object_or_404(Project, id=project_id, created_by=request.user)
    
#     # Get all active assignments for this project
#     assignments = HardwareAssignment.objects.filter(
#         project=project,
#         assigned_by=request.user,
#         actual_return_date__isnull=True
#     ).select_related('employee', 'project').prefetch_related('hardwareassignmentitem_set__hardware__hardware_type')
    
#     # Create workbook and worksheet
#     wb = openpyxl.Workbook()
    
#     # Create Summary Sheet
#     ws_summary = wb.active
#     ws_summary.title = "Project Summary"
    
#     # Define styles
#     header_font = Font(bold=True, color="FFFFFF", size=12)
#     header_fill = PatternFill(start_color="E04D00", end_color="E04D00", fill_type="solid")
#     subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
#     border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
#     center_alignment = Alignment(horizontal='center', vertical='center')
    
#     # Project Summary Sheet
#     ws_summary.merge_cells('A1:F1')
#     ws_summary['A1'] = f'PROJECT SUMMARY - {project.project_name}'
#     ws_summary['A1'].font = Font(bold=True, size=16)
#     ws_summary['A1'].alignment = center_alignment
    
#     ws_summary['A3'] = 'Project ID'
#     ws_summary['B3'] = project.project_id
#     ws_summary['A4'] = 'Project Name'
#     ws_summary['B4'] = project.project_name
#     ws_summary['A5'] = 'Location'
#     ws_summary['B5'] = project.location
#     ws_summary['A6'] = 'Duration'
#     ws_summary['B6'] = f'{project.start_date.strftime("%d-%m-%Y")} to {project.end_date.strftime("%d-%m-%Y")}'
#     ws_summary['A7'] = 'Description'
#     ws_summary['B7'] = project.description or 'N/A'
#     ws_summary['A8'] = 'Created On'
#     ws_summary['B8'] = project.created_at.strftime("%d-%m-%Y %H:%M:%S")
    
#     # Format summary headers
#     for row in range(3, 9):
#         ws_summary[f'A{row}'].font = Font(bold=True)
#         ws_summary[f'A{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
#     # Hardware Statistics
#     ws_summary['A10'] = 'HARDWARE STATISTICS'
#     ws_summary.merge_cells(f'A10:F10')
#     ws_summary['A10'].font = Font(bold=True, size=12)
#     ws_summary['A10'].fill = subheader_fill
#     ws_summary['A10'].font = Font(bold=True, color="FFFFFF")
    
#     ws_summary['A11'] = 'Total Hardware Items'
#     ws_summary['B11'] = 0
#     ws_summary['A12'] = 'Hardware In Use'
#     ws_summary['B12'] = 0
#     ws_summary['A13'] = 'Hardware Assigned'
#     ws_summary['B13'] = 0
#     ws_summary['A14'] = 'Available Hardware'
#     ws_summary['B14'] = 0
    
#     total_hardware = 0
#     active_hardware = 0
#     assigned_hardware = 0
#     available_hardware = 0
#     hardware_type_summary = {}
    
#     for assignment in assignments:
#         for item in assignment.hardwareassignmentitem_set.all():
#             hardware = item.hardware
#             total_hardware += 1
            
#             hw_type = hardware.hardware_type.name if hardware.hardware_type else 'Unknown'
#             hardware_type_summary[hw_type] = hardware_type_summary.get(hw_type, 0) + 1
            
#             if hardware.status == 'in_use':
#                 active_hardware += 1
#             elif hardware.status == 'assigned':
#                 assigned_hardware += 1
#             elif hardware.status == 'available':
#                 available_hardware += 1
    
#     ws_summary['B11'] = total_hardware
#     ws_summary['B12'] = active_hardware
#     ws_summary['B13'] = assigned_hardware
#     ws_summary['B14'] = available_hardware
    
#     # Hardware by Type
#     ws_summary['A16'] = 'HARDWARE BY TYPE'
#     ws_summary.merge_cells(f'A16:F16')
#     ws_summary['A16'].font = Font(bold=True, color="FFFFFF")
#     ws_summary['A16'].fill = subheader_fill
    
#     row = 17
#     for hw_type, count in hardware_type_summary.items():
#         ws_summary[f'A{row}'] = hw_type
#         ws_summary[f'B{row}'] = count
#         row += 1
    
#     # Employee Statistics
#     ws_summary['A20'] = 'EMPLOYEE STATISTICS'
#     ws_summary.merge_cells(f'A20:F20')
#     ws_summary['A20'].font = Font(bold=True, color="FFFFFF")
#     ws_summary['A20'].fill = subheader_fill
    
#     unique_employees = {}
#     for assignment in assignments:
#         emp_id = assignment.employee.id
#         if emp_id not in unique_employees:
#             unique_employees[emp_id] = {
#                 'name': assignment.employee.get_full_name() or assignment.employee.username,
#                 'email': assignment.employee.email,
#                 'phone': assignment.employee.phone or 'N/A'
#             }
    
#     ws_summary['A21'] = 'Total Unique Employees'
#     ws_summary['B21'] = len(unique_employees)
    
#     # Create Hardware Details Sheet
#     ws_hardware = wb.create_sheet("Hardware Details")
    
#     # Headers for Hardware Details
#     hardware_headers = ['S.No', 'Employee Name', 'Employee Email', 'Exam City', 'Hardware Type', 
#                         'Serial Number', 'Model', 'Status', 'Verification Status', 'Assigned Date']
    
#     for col, header in enumerate(hardware_headers, 1):
#         cell = ws_hardware.cell(row=1, column=col, value=header)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center_alignment
#         cell.border = border
    
#     # Add data to Hardware Details
#     row = 2
#     serial_no = 1
#     for assignment in assignments:
#         for item in assignment.hardwareassignmentitem_set.all():
#             hardware = item.hardware
            
#             # Get verification status
#             verification_status = 'Not Entered'
#             try:
#                 serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
#                 if serial_entry.verified:
#                     verification_status = 'Verified'
#                 else:
#                     verification_status = 'Pending'
#             except HardwareSerialEntry.DoesNotExist:
#                 pass
            
#             ws_hardware.cell(row=row, column=1, value=serial_no).border = border
#             ws_hardware.cell(row=row, column=2, value=assignment.employee.get_full_name() or assignment.employee.username).border = border
#             ws_hardware.cell(row=row, column=3, value=assignment.employee.email).border = border
#             ws_hardware.cell(row=row, column=4, value=assignment.exam_city).border = border
#             ws_hardware.cell(row=row, column=5, value=hardware.hardware_type.name if hardware.hardware_type else 'Unknown').border = border
#             ws_hardware.cell(row=row, column=6, value=hardware.serial_number).border = border
#             ws_hardware.cell(row=row, column=7, value=hardware.model_name or 'N/A').border = border
            
#             # Status with color
#             status_cell = ws_hardware.cell(row=row, column=8, value=hardware.status.upper())
#             if hardware.status == 'in_use':
#                 status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#             elif hardware.status == 'assigned':
#                 status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
#             # Verification status with color
#             verify_cell = ws_hardware.cell(row=row, column=9, value=verification_status)
#             if verification_status == 'Verified':
#                 verify_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
#             elif verification_status == 'Pending':
#                 verify_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
#             ws_hardware.cell(row=row, column=10, value=assignment.assigned_date.strftime("%d-%m-%Y")).border = border
            
#             row += 1
#             serial_no += 1
    
#     # Auto-adjust column widths for Hardware Details
#     for col in range(1, len(hardware_headers) + 1):
#         max_length = len(hardware_headers[col-1])
#         for row in range(2, row):
#             cell_value = ws_hardware.cell(row=row, column=col).value
#             if cell_value:
#                 max_length = max(max_length, len(str(cell_value)))
#         adjusted_width = min(max_length + 2, 30)
#         ws_hardware.column_dimensions[get_column_letter(col)].width = adjusted_width
    
#     # Create Employee Summary Sheet
#     ws_employee = wb.create_sheet("Employee Summary")
    
#     employee_headers = ['S.No', 'Employee Name', 'Email', 'Phone', 'Exam City', 'Hardware Count', 'Hardware Types']
    
#     for col, header in enumerate(employee_headers, 1):
#         cell = ws_employee.cell(row=1, column=col, value=header)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center_alignment
#         cell.border = border
    
#     # Group hardware by employee
#     employee_data = {}
#     for assignment in assignments:
#         emp_id = assignment.employee.id
#         emp_name = assignment.employee.get_full_name() or assignment.employee.username
        
#         if emp_id not in employee_data:
#             employee_data[emp_id] = {
#                 'name': emp_name,
#                 'email': assignment.employee.email,
#                 'phone': assignment.employee.phone or 'N/A',
#                 'exam_city': assignment.exam_city,
#                 'hardware_count': 0,
#                 'hardware_types': set()
#             }
        
#         for item in assignment.hardwareassignmentitem_set.all():
#             hardware = item.hardware
#             employee_data[emp_id]['hardware_count'] += 1
#             hw_type = hardware.hardware_type.name if hardware.hardware_type else 'Unknown'
#             employee_data[emp_id]['hardware_types'].add(hw_type)
    
#     row = 2
#     serial_no = 1
#     for emp_id, data in employee_data.items():
#         ws_employee.cell(row=row, column=1, value=serial_no).border = border
#         ws_employee.cell(row=row, column=2, value=data['name']).border = border
#         ws_employee.cell(row=row, column=3, value=data['email']).border = border
#         ws_employee.cell(row=row, column=4, value=data['phone']).border = border
#         ws_employee.cell(row=row, column=5, value=data['exam_city']).border = border
#         ws_employee.cell(row=row, column=6, value=data['hardware_count']).border = border
#         ws_employee.cell(row=row, column=7, value=', '.join(data['hardware_types'])).border = border
#         row += 1
#         serial_no += 1
    
#     # Auto-adjust column widths for Employee Summary
#     for col in range(1, len(employee_headers) + 1):
#         max_length = len(employee_headers[col-1])
#         for row in range(2, row):
#             cell_value = ws_employee.cell(row=row, column=col).value
#             if cell_value:
#                 max_length = max(max_length, len(str(cell_value)))
#         adjusted_width = min(max_length + 2, 35)
#         ws_employee.column_dimensions[get_column_letter(col)].width = adjusted_width
    
#     # Auto-adjust column widths for Summary Sheet
#     for col in ['A', 'B']:
#         ws_summary.column_dimensions[col].width = 25
    
#     # Prepare response
#     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#     filename = f"project_{project.project_id}_{timestamp}.xlsx"
    
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
#     wb.save(response)
#     return response


@login_required
def export_all_projects_excel(request):
    """Export all projects summary to Excel"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    projects = Project.objects.filter(created_by=request.user).order_by('-created_at')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Projects Summary"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="E04D00", end_color="E04D00", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['S.No', 'Project ID', 'Project Name', 'Location', 'Start Date', 'End Date', 
               'Total Hardware', 'Hardware In Use', 'Employees Count', 'Created Date']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    row = 2
    serial_no = 1
    
    for project in projects:
        # Get project statistics
        assignments = HardwareAssignment.objects.filter(
            project=project,
            assigned_by=request.user,
            actual_return_date__isnull=True
        )
        
        total_hardware = 0
        active_hardware = 0
        for assignment in assignments:
            for item in assignment.hardwareassignmentitem_set.all():
                total_hardware += 1
                if item.hardware.status == 'in_use':
                    active_hardware += 1
        
        employee_count = assignments.values('employee').distinct().count()
        
        ws.cell(row=row, column=1, value=serial_no).border = border
        ws.cell(row=row, column=2, value=project.project_id).border = border
        ws.cell(row=row, column=3, value=project.project_name).border = border
        ws.cell(row=row, column=4, value=project.location).border = border
        ws.cell(row=row, column=5, value=project.start_date.strftime("%d-%m-%Y")).border = border
        ws.cell(row=row, column=6, value=project.end_date.strftime("%d-%m-%Y")).border = border
        ws.cell(row=row, column=7, value=total_hardware).border = border
        ws.cell(row=row, column=8, value=active_hardware).border = border
        ws.cell(row=row, column=9, value=employee_count).border = border
        ws.cell(row=row, column=10, value=project.created_at.strftime("%d-%m-%Y %H:%M")).border = border
        
        row += 1
        serial_no += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col-1])
        for row in range(2, row):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"all_projects_summary_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def delete_project(request, project_id):
    """Delete a project if it has no active assignments"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    project = get_object_or_404(Project, id=project_id, created_by=request.user)
    
    # Check if project has any active hardware assignments
    has_active_assignments = HardwareAssignment.objects.filter(
        project=project,
        actual_return_date__isnull=True
    ).exists()
    
    if has_active_assignments:
        messages.error(request, f'Cannot delete "{project.project_name}" because it has active hardware assignments. Please return all hardware first.')
    else:
        project_name = project.project_name
        project.delete()
        messages.success(request, f'Project "{project_name}" has been deleted successfully.')
    
    return redirect('create_project')    
# Add this import at the top of your views.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
@login_required
def project_assignments(request, project_id):
    """View all employees assigned to a specific project"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    project = get_object_or_404(Project, id=project_id, created_by=request.user)
    
    # Get all active assignments for this project
    assignments = HardwareAssignment.objects.filter(
        project=project,
        assigned_by=request.user,
        actual_return_date__isnull=True
    ).select_related('employee').prefetch_related(
        'hardwareassignmentitem_set__hardware',
        'hardwareassignmentitem_set__hardware__hardware_type',
        'hardwareassignmentitem_set__asset_entry'
    ).order_by('-assigned_date')
    
    total_hardware = 0
    active_hardware = 0
    assigned_hardware = 0
    hardware_by_type = {}
    employees_dict = {}
    
    # Store detailed hardware list for export
    hardware_details_list = []
    
    for assignment in assignments:
        # Get hardware items count for this assignment
        hardware_items = assignment.hardwareassignmentitem_set.all()
        item_count = hardware_items.count()
        total_hardware += item_count
        
        # Get exam center name
        exam_center_name = getattr(assignment, 'exam_center_name', None)
        
        # Calculate hardware status counts and collect details
        for item in hardware_items:
            hardware = item.hardware
            hw_type = hardware.hardware_type.name if hardware.hardware_type else 'Unknown'
            hardware_by_type[hw_type] = hardware_by_type.get(hw_type, 0) + 1
            
            if hardware.status == 'in_use':
                active_hardware += 1
            elif hardware.status == 'assigned':
                assigned_hardware += 1
            
            # Get verification status using asset_entry
            verification_status = 'Not Entered'
            verified = False
            try:
                asset_entry = item.asset_entry
                if asset_entry.verified:
                    verification_status = 'Verified'
                    verified = True
                else:
                    # Check if asset number matches
                    expected_asset = hardware.asset_number if hardware.asset_number else 'N/A'
                    if asset_entry.entered_asset_number == expected_asset:
                        verification_status = 'Matched - Pending'
                    else:
                        verification_status = 'Mismatch'
            except HardwareAssetEntry.DoesNotExist:
                pass
            
            # Add to hardware details list for export
            hardware_details_list.append({
                'employee_name': assignment.employee.get_full_name() or assignment.employee.username,
                'employee_email': assignment.employee.email,
                'exam_city': assignment.exam_city,
                'exam_center_name': exam_center_name,
                'hardware_type': hw_type,
                'serial_number': hardware.serial_number,
                'model': hardware.model_name,
                'brand': hardware.brand or 'N/A',
                'status': 'In Use' if hardware.status == 'in_use' else 'Assigned',
                'asset_number': hardware.asset_number if hardware.asset_number else 'N/A',
                'entered_asset': asset_entry.entered_asset_number if hasattr(item, 'asset_entry') and item.asset_entry else 'N/A',
                'verification_status': verification_status,
                'verified': verified,
                'assigned_date': assignment.assigned_date.strftime('%d-%m-%Y'),
                'assignment_id': assignment.id
            })
        
        # Get overall verification status for this assignment (for employee view)
        verification_status_emp = 'not_entered'
        verified_emp = False
        
        # Check all hardware items for this assignment
        all_verified = True
        any_pending = False
        any_mismatch = False
        
        for item in hardware_items:
            try:
                asset_entry = item.asset_entry
                if asset_entry.verified:
                    pass
                else:
                    all_verified = False
                    any_pending = True
                    expected_asset = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
                    if asset_entry.entered_asset_number != expected_asset:
                        any_mismatch = True
            except HardwareAssetEntry.DoesNotExist:
                all_verified = False
        
        if all_verified and hardware_items.exists():
            verified_emp = True
            verification_status_emp = 'verified'
        elif any_mismatch:
            verification_status_emp = 'mismatch'
        elif any_pending:
            verification_status_emp = 'pending'
        else:
            verification_status_emp = 'not_entered'
        
        # Store employee data (one per employee)
        employee_id = assignment.employee.id
        if employee_id not in employees_dict:
            employees_dict[employee_id] = {
                'id': employee_id,
                'name': assignment.employee.get_full_name() or assignment.employee.username,
                'email': assignment.employee.email,
                'exam_city': assignment.exam_city,
                'exam_center_name': exam_center_name,
                'status': 'Active',
                'verified': verified_emp,
                'verification_status': verification_status_emp,
                'assignment_id': assignment.id,
                'hardware_count': item_count,
            }
        else:
            existing = employees_dict[employee_id]
            existing['hardware_count'] += item_count
            
            if verification_status_emp == 'verified' and existing['verification_status'] != 'verified':
                existing['verified'] = True
                existing['verification_status'] = 'verified'
            elif verification_status_emp == 'mismatch':
                existing['verification_status'] = 'mismatch'
                existing['verified'] = False
            elif verification_status_emp == 'pending' and existing['verification_status'] == 'not_entered':
                existing['verification_status'] = 'pending'
    
    # Check if export is requested
    if request.GET.get('export') == 'excel':
        return export_project_hardware_excel(project, hardware_details_list, total_hardware, active_hardware, assigned_hardware, hardware_by_type)
    
    employees_list = list(employees_dict.values())
    
    context = {
        'project': project,
        'employees': employees_list,
        'total_hardware': total_hardware,
        'active_hardware': active_hardware,
        'assigned_hardware': assigned_hardware,
        'hardware_by_type': hardware_by_type,
        'employee_count': len(employees_list),
    }
    return render(request, 'manager/project_assignments.html', context)



def export_project_hardware_excel(project, hardware_details_list, total_hardware, active_hardware, assigned_hardware, hardware_by_type):
    """Export project hardware details to Excel with employee grouping and filterable format"""
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E04D00", end_color="E04D00", fill_type="solid")
    employee_bg = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ========== SHEET 1: Hardware Details (Grouped by Employee with Gaps) ==========
    ws_data = wb.create_sheet("Hardware Details")
    
    # Group hardware by employee
    employee_hardware = {}
    for hardware in hardware_details_list:
        emp_name = hardware['employee_name']
        if emp_name not in employee_hardware:
            employee_hardware[emp_name] = {
                'email': hardware['employee_email'],
                'exam_city': hardware['exam_city'],
                'exam_center_name': hardware.get('exam_center_name', 'Not specified'),
                'hardware_list': []
            }
        employee_hardware[emp_name]['hardware_list'].append(hardware)
    
    # Headers - Updated to include Exam Center
    headers = ['S.No', 'Employee Name', 'Email', 'Exam City', 'Exam Center', 'Hardware Type', 'Asset Number', 'Entered Asset', 'Serial Number', 'Status', 'Verification']
    
    # Write main headers (Row 1)
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    current_row = 2
    global_sno = 1
    
    for emp_name, emp_data in sorted(employee_hardware.items()):
        # Employee Header Row (merged across all columns)
        ws_data.merge_cells(f'A{current_row}:K{current_row}')
        emp_header_cell = ws_data.cell(row=current_row, column=1, value=f"👤 EMPLOYEE: {emp_name}")
        emp_header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        emp_header_cell.fill = employee_bg
        emp_header_cell.alignment = left_alignment
        emp_header_cell.border = border
        current_row += 1
        
        # Write each hardware item as a complete row (with employee info repeated)
        for hardware in emp_data['hardware_list']:
            asset_number = hardware.get('asset_number', 'N/A')
            entered_asset = hardware.get('entered_asset', 'Not Entered')
            exam_center_name = hardware.get('exam_center_name', 'Not specified')
            
            # S.No
            ws_data.cell(row=current_row, column=1, value=global_sno).border = border
            # Employee info (repeated for filtering)
            ws_data.cell(row=current_row, column=2, value=emp_name).border = border
            ws_data.cell(row=current_row, column=3, value=emp_data['email']).border = border
            ws_data.cell(row=current_row, column=4, value=emp_data['exam_city']).border = border
            ws_data.cell(row=current_row, column=5, value=exam_center_name).border = border
            # Hardware info
            ws_data.cell(row=current_row, column=6, value=hardware['hardware_type']).border = border
            ws_data.cell(row=current_row, column=7, value=asset_number).border = border
            
            # Entered Asset with color coding
            entered_cell = ws_data.cell(row=current_row, column=8, value=entered_asset)
            if entered_asset != 'Not Entered' and entered_asset != 'N/A':
                if entered_asset == asset_number:
                    entered_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    entered_cell.font = Font(color="006100", bold=True)
                else:
                    entered_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    entered_cell.font = Font(color="9C0006", bold=True)
            entered_cell.border = border
            
            ws_data.cell(row=current_row, column=9, value=hardware['serial_number']).border = border
            
            # Status with color
            status_cell = ws_data.cell(row=current_row, column=10, value=hardware['status'])
            if hardware['status'] == 'In Use':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(color="9C5700", bold=True)
            status_cell.border = border
            
            # Verification with color
            verify_cell = ws_data.cell(row=current_row, column=11, value=hardware['verification_status'])
            if hardware['verification_status'] == 'Verified':
                verify_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                verify_cell.font = Font(color="006100", bold=True)
            elif hardware['verification_status'] == 'Matched - Pending':
                verify_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                verify_cell.font = Font(color="9C5700", bold=True)
            elif hardware['verification_status'] == 'Mismatch':
                verify_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                verify_cell.font = Font(color="9C0006", bold=True)
            else:
                verify_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                verify_cell.font = Font(color="9C0006", bold=True)
            verify_cell.border = border
            
            current_row += 1
            global_sno += 1
        
        # Employee Summary Row
        total_items = len(emp_data['hardware_list'])
        emp_verified = sum(1 for h in emp_data['hardware_list'] if h['verification_status'] == 'Verified')
        emp_matched = sum(1 for h in emp_data['hardware_list'] if h['verification_status'] == 'Matched - Pending')
        emp_mismatch = sum(1 for h in emp_data['hardware_list'] if h['verification_status'] == 'Mismatch')
        emp_not_entered = total_items - emp_verified - emp_matched - emp_mismatch
        
        ws_data.merge_cells(f'A{current_row}:E{current_row}')
        summary_label = ws_data.cell(row=current_row, column=1, value=f"📊 Summary for {emp_name}:")
        summary_label.font = Font(bold=True, size=10)
        summary_label.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        ws_data.merge_cells(f'F{current_row}:K{current_row}')
        summary_value = ws_data.cell(row=current_row, column=6, value=f"Total: {total_items} | Verified: {emp_verified} | Matched: {emp_matched} | Mismatch: {emp_mismatch} | Not Entered: {emp_not_entered}")
        summary_value.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        for col in range(1, 12):
            ws_data.cell(row=current_row, column=col).border = border
        current_row += 1
        
        # Add 2 empty rows as gap between employees
        current_row += 2
    
    # Enable AutoFilter on the entire data range
    ws_data.auto_filter.ref = f"A1:K{current_row-1}"
    
    # Freeze header row
    ws_data.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col-1])
        for row_idx in range(2, current_row):
            cell_value = ws_data.cell(row=row_idx, column=col).value
            if cell_value and cell_value != "📊 Summary":
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 30)
        ws_data.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # Set row heights
    ws_data.row_dimensions[1].height = 25
    
    # ========== SHEET 2: Employee Summary (Flat, Filterable) ==========
    ws_summary = wb.create_sheet("Employee Summary")
    
    # Group by employee for summary
    employee_summary = {}
    for hardware in hardware_details_list:
        emp_name = hardware['employee_name']
        if emp_name not in employee_summary:
            employee_summary[emp_name] = {
                'email': hardware['employee_email'],
                'exam_city': hardware['exam_city'],
                'exam_center_name': hardware.get('exam_center_name', 'Not specified'),
                'total': 0,
                'verified': 0,
                'matched': 0,
                'mismatch': 0,
                'not_entered': 0,
                'in_use': 0,
                'assigned': 0
            }
        employee_summary[emp_name]['total'] += 1
        if hardware['verification_status'] == 'Verified':
            employee_summary[emp_name]['verified'] += 1
        elif hardware['verification_status'] == 'Matched - Pending':
            employee_summary[emp_name]['matched'] += 1
        elif hardware['verification_status'] == 'Mismatch':
            employee_summary[emp_name]['mismatch'] += 1
        else:
            employee_summary[emp_name]['not_entered'] += 1
        
        if hardware['status'] == 'In Use':
            employee_summary[emp_name]['in_use'] += 1
        else:
            employee_summary[emp_name]['assigned'] += 1
    
    # Summary headers - Updated with Exam Center
    sum_headers = ['S.No', 'Employee Name', 'Email', 'Exam City', 'Exam Center', 'Total Items', 'In Use', 'Assigned', 
                   'Verified', 'Matched', 'Mismatch', 'Not Entered', 'Completion %']
    
    for col, header in enumerate(sum_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    row = 2
    for idx, (emp_name, data) in enumerate(sorted(employee_summary.items()), 1):
        completion = (data['verified'] / data['total'] * 100) if data['total'] > 0 else 0
        
        ws_summary.cell(row=row, column=1, value=idx).border = border
        ws_summary.cell(row=row, column=2, value=emp_name).border = border
        ws_summary.cell(row=row, column=3, value=data['email']).border = border
        ws_summary.cell(row=row, column=4, value=data['exam_city']).border = border
        ws_summary.cell(row=row, column=5, value=data['exam_center_name']).border = border
        ws_summary.cell(row=row, column=6, value=data['total']).border = border
        ws_summary.cell(row=row, column=7, value=data['in_use']).border = border
        ws_summary.cell(row=row, column=8, value=data['assigned']).border = border
        ws_summary.cell(row=row, column=9, value=data['verified']).border = border
        ws_summary.cell(row=row, column=10, value=data['matched']).border = border
        ws_summary.cell(row=row, column=11, value=data['mismatch']).border = border
        ws_summary.cell(row=row, column=12, value=data['not_entered']).border = border
        
        comp_cell = ws_summary.cell(row=row, column=13, value=f"{completion:.1f}%")
        if completion == 100:
            comp_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            comp_cell.font = Font(color="006100", bold=True)
        elif completion >= 50:
            comp_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            comp_cell.font = Font(color="9C5700", bold=True)
        else:
            comp_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            comp_cell.font = Font(color="9C0006", bold=True)
        comp_cell.border = border
        
        row += 1
    
    # Enable filter on summary sheet
    ws_summary.auto_filter.ref = f"A1:M{row-1}"
    ws_summary.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for col in range(1, len(sum_headers) + 1):
        max_length = len(sum_headers[col-1])
        for row_idx in range(2, row):
            cell_value = ws_summary.cell(row=row_idx, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 30)
        ws_summary.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # ========== SHEET 3: Hardware by Type ==========
    ws_type = wb.create_sheet("Hardware by Type")
    
    type_summary = {}
    for hardware in hardware_details_list:
        hw_type = hardware['hardware_type']
        if hw_type not in type_summary:
            type_summary[hw_type] = 0
        type_summary[hw_type] += 1
    
    type_headers = ['Hardware Type', 'Count', 'Percentage', 'Visual']
    
    for col, header in enumerate(type_headers, 1):
        cell = ws_type.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    row = 2
    for hw_type, count in sorted(type_summary.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / len(hardware_details_list) * 100) if hardware_details_list else 0
        
        ws_type.cell(row=row, column=1, value=hw_type).border = border
        ws_type.cell(row=row, column=2, value=count).border = border
        ws_type.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
        
        # Progress bar
        bar_length = int(percentage / 5)
        bar = "█" * bar_length + "░" * (20 - bar_length)
        ws_type.cell(row=row, column=4, value=bar).border = border
        
        row += 1
    
    ws_type.column_dimensions['A'].width = 25
    ws_type.column_dimensions['B'].width = 15
    ws_type.column_dimensions['C'].width = 15
    ws_type.column_dimensions['D'].width = 25
    
    # ========== SHEET 4: Project Information ==========
    ws_info = wb.create_sheet("Project Info")
    
    info_headers = ['Field', 'Value']
    for col, header in enumerate(info_headers, 1):
        cell = ws_info.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    verified_total = sum(1 for h in hardware_details_list if h['verification_status'] == 'Verified')
    matched_total = sum(1 for h in hardware_details_list if h['verification_status'] == 'Matched - Pending')
    mismatch_total = sum(1 for h in hardware_details_list if h['verification_status'] == 'Mismatch')
    not_entered_total = len(hardware_details_list) - verified_total - matched_total - mismatch_total
    
    project_details = [
        ['Project ID', project.project_id],
        ['Project Name', project.project_name],
        ['Location', project.location],
        ['Start Date', project.start_date.strftime("%d-%m-%Y")],
        ['End Date', project.end_date.strftime("%d-%m-%Y")],
        ['Duration', f'{(project.end_date - project.start_date).days} days'],
        ['Created By', project.created_by.get_full_name() or project.created_by.username],
        ['Created On', project.created_at.strftime("%d-%m-%Y %H:%M:%S")],
        ['', ''],
        ['📊 STATISTICS', ''],
        ['Total Hardware Items', len(hardware_details_list)],
        ['Hardware In Use', active_hardware],
        ['Hardware Assigned', assigned_hardware],
        ['', ''],
        ['✅ Verification Summary', ''],
        ['Verified', verified_total],
        ['Matched (Pending)', matched_total],
        ['Mismatch', mismatch_total],
        ['Not Entered', not_entered_total],
        ['', ''],
        ['Completion Rate', f"{(verified_total / len(hardware_details_list) * 100):.1f}%" if hardware_details_list else "0%"],
    ]
    
    row = 2
    for detail in project_details:
        if detail[0] in ['📊 STATISTICS', '✅ Verification Summary']:
            cell = ws_info.cell(row=row, column=1, value=detail[0])
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws_info.merge_cells(f'A{row}:B{row}')
            cell.alignment = center_alignment
            row += 1
        else:
            ws_info.cell(row=row, column=1, value=detail[0]).border = border
            ws_info.cell(row=row, column=2, value=detail[1]).border = border
            if detail[0]:
                ws_info.cell(row=row, column=1).font = Font(bold=True)
            row += 1
    
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 35
    
    # Prepare response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"PROJECT_{project.project_name.replace(' ', '_')}_HARDWARE_REPORT_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# ============== HARDWARE MANAGEMENT VIEWS ==============
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.paginator import Paginator
from django.db.models import Q, Count

@login_required
def manage_hardware(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware_types = HardwareType.objects.all()
    hardware_items = Hardware.objects.filter(created_by=request.user)
    
    hardware_items_count = hardware_items.values('hardware_type').annotate(
        count=Count('id')
    ).order_by()
    
    hardware_items_count_dict = {}
    for item in hardware_items_count:
        hardware_items_count_dict[item['hardware_type']] = item['count']
    
    search_query = request.GET.get('search', '')
    if search_query:
        hardware_items = hardware_items.filter(
            Q(asset_number__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(hardware_type__name__icontains=search_query)
        )
    
    type_filter = request.GET.get('type', '')
    if type_filter and type_filter != 'all':
        hardware_items = hardware_items.filter(hardware_type_id=type_filter)
    
    status_filter = request.GET.get('status', '')
    if status_filter and status_filter != 'all':
        hardware_items = hardware_items.filter(status=status_filter)
    
    available_count = hardware_items.filter(status='available').count()
    assigned_count = hardware_items.filter(status='assigned').count()
    in_use_count = hardware_items.filter(status='in_use').count()
    maintenance_count = hardware_items.filter(status='maintenance').count()
    
    paginator = Paginator(hardware_items, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    from django.template.defaulttags import register
    register.filter('get_item', lambda d, key: d.get(key, 0))
    
    context = {
        'hardware_types': hardware_types, 
        'hardware_items': page_obj,
        'hardware_items_count_dict': hardware_items_count_dict,
        'available_count': available_count,
        'assigned_count': assigned_count,
        'in_use_count': in_use_count,
        'maintenance_count': maintenance_count,
        'search_query': search_query,
        'type_filter': type_filter,
        'status_filter': status_filter,
        'paginator': paginator,
        'page_obj': page_obj,
    }
    return render(request, 'manager/manage_hardware.html', context)

# Add these imports at the top of your views.py
import openpyxl
import pandas as pd
from django.http import HttpResponse
from datetime import datetime
@login_required
def add_hardware(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware_types = HardwareType.objects.all()
    
    if request.method == 'POST':
        # Check if this is an Excel import
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Please upload an Excel file (.xlsx or .xls)')
                return redirect('add_hardware')
            
            try:
                # Read Excel file
                df = pd.read_excel(excel_file)
                
                # Expected columns - hardware_type first, then asset_number, then serial_number
                expected_columns = ['hardware_type', 'asset_number', 'serial_number']
                
                # Validate columns
                missing_columns = [col for col in expected_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f'Missing columns in Excel: {", ".join(missing_columns)}')
                    return redirect('add_hardware')
                
                success_count = 0
                error_count = 0
                errors = []
                
                for index, row in df.iterrows():
                    try:
                        hardware_type_name = str(row['hardware_type']).strip()
                        asset_number = str(row['asset_number']).strip()
                        serial_number = str(row['serial_number']).strip()
                        
                        # Validate required fields
                        if not hardware_type_name or not asset_number or not serial_number:
                            errors.append(f"Row {index + 2}: Missing required fields (hardware_type, asset_number, serial_number)")
                            error_count += 1
                            continue
                        
                        # Check if hardware type exists
                        try:
                            hardware_type = HardwareType.objects.get(name__iexact=hardware_type_name)
                        except HardwareType.DoesNotExist:
                            if hardware_type_name.isdigit():
                                try:
                                    hardware_type = HardwareType.objects.get(id=int(hardware_type_name))
                                except HardwareType.DoesNotExist:
                                    errors.append(f"Row {index + 2}: Hardware type ID '{hardware_type_name}' not found")
                                    error_count += 1
                                    continue
                            else:
                                errors.append(f"Row {index + 2}: Hardware type '{hardware_type_name}' not found")
                                error_count += 1
                                continue
                        
                        # Check if asset number already exists
                        if Hardware.objects.filter(asset_number=asset_number).exists():
                            errors.append(f"Row {index + 2}: Asset number '{asset_number}' already exists")
                            error_count += 1
                            continue
                        
                        # Check for duplicate serial number
                        if Hardware.objects.filter(serial_number=serial_number).exists():
                            errors.append(f"Row {index + 2}: Serial number '{serial_number}' already exists")
                            error_count += 1
                            continue
                        
                        # Create hardware
                        Hardware.objects.create(
                            hardware_type=hardware_type,
                            asset_number=asset_number,
                            serial_number=serial_number,
                            status='available',
                            created_by=request.user
                        )
                        success_count += 1
                        
                    except Exception as e:
                        errors.append(f"Row {index + 2}: {str(e)}")
                        error_count += 1
                
                if success_count > 0:
                    messages.success(request, f'Successfully imported {success_count} hardware items!')
                
                if error_count > 0:
                    error_message = f'Failed to import {error_count} items. '
                    if errors:
                        error_message += ' First few errors: ' + '; '.join(errors[:3])
                    messages.warning(request, error_message)
                
                return redirect('manage_hardware')
                
            except Exception as e:
                messages.error(request, f'Error reading Excel file: {str(e)}')
                return redirect('add_hardware')
        
        # Single hardware addition
        else:
            hardware_type_id = request.POST.get('hardware_type')
            asset_number = request.POST.get('asset_number')
            serial_number = request.POST.get('serial_number')
            
            # Validate required fields
            if not hardware_type_id:
                messages.error(request, 'Hardware type is required!')
                return redirect('add_hardware')
            
            if not asset_number:
                messages.error(request, 'Asset number is required!')
                return redirect('add_hardware')
            
            if not serial_number:
                messages.error(request, 'Serial number is required!')
                return redirect('add_hardware')
            
            if Hardware.objects.filter(asset_number=asset_number).exists():
                messages.error(request, 'Asset number already exists!')
                return redirect('add_hardware')
            
            if Hardware.objects.filter(serial_number=serial_number).exists():
                messages.error(request, 'Serial number already exists!')
                return redirect('add_hardware')
            
            try:
                hardware_type = HardwareType.objects.get(id=hardware_type_id)
            except HardwareType.DoesNotExist:
                messages.error(request, 'Invalid hardware type selected!')
                return redirect('add_hardware')
            
            hardware = Hardware.objects.create(
                hardware_type=hardware_type,
                asset_number=asset_number,
                serial_number=serial_number,
                status='available',
                created_by=request.user
            )
            
            messages.success(request, f'{hardware_type.name} added successfully! Asset Number: {asset_number}')
            return redirect('manage_hardware')
    
    context = {'hardware_types': hardware_types}
    return render(request, 'manager/add_hardware.html', context)



@login_required

@login_required
def download_hardware_template(request):
    """Download Excel template for bulk hardware import"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hardware Template"
    
    # Define headers - hardware_type first, then asset_number, then serial_number
    headers = ['hardware_type', 'asset_number', 'serial_number']
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E04D00", end_color="E04D00", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Get ALL hardware types from database
    hardware_types = HardwareType.objects.all().order_by('name')
    
    # Create a separate sheet for hardware types list
    ws_types_list = wb.create_sheet("HardwareTypesList")
    
    # Write all hardware types to the list sheet for reference
    ws_types_list.cell(row=1, column=1, value="Hardware Type")
    ws_types_list.cell(row=1, column=2, value="Description")
    
    # Style the header row
    for col in range(1, 3):
        cell = ws_types_list.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="E04D00", end_color="E04D00", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write all hardware types to the list sheet
    for idx, hw_type in enumerate(hardware_types, 2):
        ws_types_list.cell(row=idx, column=1, value=hw_type.name)
        ws_types_list.cell(row=idx, column=2, value=hw_type.description or '')
    
    # Create dropdown validation for hardware_type column (Column A)
    if hardware_types.exists():
        range_ref = f'HardwareTypesList!$A$2:$A${hardware_types.count() + 1}'
        
        from openpyxl.worksheet.datavalidation import DataValidation
        
        dv = DataValidation(type="list", formula1=range_ref, showDropDown=True)
        dv.error = 'Please select a valid hardware type from the list'
        dv.errorTitle = 'Invalid Hardware Type'
        dv.prompt = 'Select hardware type from dropdown'
        dv.promptTitle = 'Hardware Type'
        
        ws.add_data_validation(dv)
        dv.add('A2:A1000')  # Column A is hardware_type
    
    # Add example data
    example_data = []
    example_types = hardware_types[:10]
    
    for idx, hw_type in enumerate(example_types):
        example_data.append([
            hw_type.name,  # hardware_type
            f'AST-{str(idx+1).zfill(4)}',  # asset_number
            f'{hw_type.name[:3].upper()}-{str(idx+1).zfill(3)}'  # serial_number
        ])
    
    # Add a few blank rows for user to fill
    for i in range(5):
        example_data.append(['', '', ''])
    
    # Write example data
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            if value:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    # Add notes sheet
    ws_notes = wb.create_sheet("Instructions")
    
    notes = [
        ["📋 INSTRUCTIONS FOR BULK HARDWARE IMPORT"],
        [""],
        ["🔹 REQUIRED COLUMNS:"],
        ["   1. hardware_type - Select from dropdown or enter exact name from list below"],
        ["   2. asset_number - Unique asset tag number (must not exist in system)"],
        ["   3. serial_number - Unique serial number (must not exist in system)"],
        [""],
        ["⚠️ IMPORTANT NOTES:"],
        ["   • Hardware types must already exist (see complete list below)"],
        ["   • Asset numbers must be unique across entire system"],
        ["   • Serial numbers must be unique across entire system"],
        ["   • Hardware will be added with 'Available' status"],
        ["   • Do not modify the column headers"],
        ["   • Remove example rows before importing"],
        ["   • You can add multiple rows - the system will process all of them"],
        [""],
        ["📋 COMPLETE LIST OF AVAILABLE HARDWARE TYPES IN SYSTEM:"],
        [""],
    ]
    
    notes.append(["   " + "=" * 70])
    notes.append(["   {:.<30} {:.<20}".format("HARDWARE TYPE", "DESCRIPTION")])
    notes.append(["   " + "-" * 70])
    
    for hw_type in hardware_types:
        desc = hw_type.description if hw_type.description else "—"
        notes.append([f"   • {hw_type.name:<28} {desc:<30}"])
    
    notes.append(["   " + "=" * 70])
    notes.append([f"   Total Hardware Types: {hardware_types.count()}"])
    
    notes.append([""])
    notes.append(["📝 ASSET NUMBER FORMAT SUGGESTION:"])
    notes.append(["   • AST-0001, AST-0002, etc."])
    notes.append(["   • ORG-HW-001, ORG-HW-002, etc."])
    notes.append(["   • Use any format as long as it's unique"])
    notes.append([""])
    notes.append(["✅ EXAMPLE:"])
    notes.append(["   Row 2 shows: Laptop | AST-0001 | LAP-001"])
    notes.append(["   Replace with your actual data"])
    notes.append([""])
    notes.append(["❌ COMMON ERRORS TO AVOID:"])
    notes.append(["   • Invalid hardware type name - must match exactly from list"])
    notes.append(["   • Duplicate asset numbers - each must be unique"])
    notes.append(["   • Duplicate serial numbers - each must be unique"])
    notes.append(["   • Empty required fields - all required columns must be filled"])
    notes.append([""])
    notes.append(["📧 Need Help? Contact your system administrator"])
    
    # Style the notes sheet
    for row_idx, row_data in enumerate(notes, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_notes.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="E04D00")
            elif "COMPLETE LIST" in str(value):
                cell.font = Font(bold=True, size=12, color="0d6efd")
            elif "Total Hardware Types" in str(value):
                cell.font = Font(bold=True, size=12, color="28a745")
    
    # Auto-adjust column widths for main sheet
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col-1])
        for row in range(2, len(example_data) + 2):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 3, 35)
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Auto-adjust column widths for hardware types list sheet
    ws_types_list.column_dimensions['A'].width = 25
    ws_types_list.column_dimensions['B'].width = 40
    
    # Auto-adjust column width for notes sheet
    max_notes_length = 0
    for row in notes:
        for cell in row:
            if cell:
                max_notes_length = max(max_notes_length, len(str(cell)))
    ws_notes.column_dimensions['A'].width = min(max_notes_length + 5, 100)
    
    # Hide the HardwareTypesList sheet
    ws_types_list.sheet_state = 'hidden'
    
    # Prepare response
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"hardware_import_template_{timestamp}.xlsx"
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required

@login_required
def edit_hardware(request, hardware_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware = get_object_or_404(Hardware, id=hardware_id, created_by=request.user)
    hardware_types = HardwareType.objects.all()
    
    if request.method == 'POST':
        hardware_type_id = request.POST.get('hardware_type')
        asset_number = request.POST.get('asset_number')
        serial_number = request.POST.get('serial_number')
        status = request.POST.get('status')
        
        # Validate required fields
        if not hardware_type_id:
            messages.error(request, 'Hardware type is required!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        if not asset_number:
            messages.error(request, 'Asset number is required!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        if not serial_number:
            messages.error(request, 'Serial number is required!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        # Check if asset number already exists (excluding current hardware)
        if Hardware.objects.filter(asset_number=asset_number).exclude(id=hardware_id).exists():
            messages.error(request, 'Asset number already exists!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        # Check if serial number already exists (excluding current hardware)
        if Hardware.objects.filter(serial_number=serial_number).exclude(id=hardware_id).exists():
            messages.error(request, 'Serial number already exists!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        try:
            hardware.hardware_type = HardwareType.objects.get(id=hardware_type_id)
        except HardwareType.DoesNotExist:
            messages.error(request, 'Invalid hardware type selected!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        hardware.asset_number = asset_number
        hardware.serial_number = serial_number
        hardware.status = status
        hardware.save()
        
        messages.success(request, f'Hardware updated successfully! Asset Number: {asset_number}')
        return redirect('manage_hardware')
    
    context = {
        'hardware': hardware,
        'hardware_types': hardware_types,
    }
    return render(request, 'manager/edit_hardware.html', context)

@login_required
def delete_hardware(request, hardware_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware = get_object_or_404(Hardware, id=hardware_id, created_by=request.user)
    
    if request.method == 'POST':
        if hardware.status == 'assigned' or hardware.status == 'in_use':
            messages.error(request, 'Cannot delete hardware that is currently assigned!')
            return redirect('manage_hardware')
        
        hardware.delete()
        messages.success(request, 'Hardware deleted successfully!')
        return redirect('manage_hardware')
    
    context = {'hardware': hardware}
    return render(request, 'manager/delete_hardware.html', context)

@login_required
def manage_hardware_types(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware_types = HardwareType.objects.all()
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        
        if HardwareType.objects.filter(name=name).exists():
            messages.error(request, 'Hardware type already exists!')
        else:
            HardwareType.objects.create(
                name=name,
                description=description
            )
            messages.success(request, 'Hardware type added successfully!')
        
        return redirect('manage_hardware_types')
    
    context = {'hardware_types': hardware_types}
    return render(request, 'manager/manage_hardware_types.html', context)

# ============== ASSIGNMENT VIEWS ==============
@login_required


@login_required
def create_assignment(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    employees = CustomUser.objects.filter(user_type='employee', manager=request.user, is_active=True)
    projects = Project.objects.filter(created_by=request.user)
    hardware_types = HardwareType.objects.all()
    
    if request.method == 'POST':
        employee_id = request.POST.get('employee')
        project_id = request.POST.get('project')
        expected_return_date = request.POST.get('expected_return_date')
        exam_city = request.POST.get('exam_city')
        exam_center_name = request.POST.get('exam_center_name')  # New field
        notes = request.POST.get('notes')
        
        # Get hardware items as JSON or list
        hardware_items_data = request.POST.get('hardware_items_json', '[]')
        import json
        try:
            hardware_items = json.loads(hardware_items_data)
        except:
            hardware_items = []
        
        if not all([employee_id, project_id, expected_return_date, exam_city]):
            messages.error(request, 'Please fill in all required fields!')
            return redirect('create_assignment')
        
        if not hardware_items:
            messages.error(request, 'Please add at least one hardware item!')
            return redirect('create_assignment')
        
        try:
            employee = CustomUser.objects.get(id=employee_id, manager=request.user)
            project = Project.objects.get(id=project_id, created_by=request.user)
        except (CustomUser.DoesNotExist, Project.DoesNotExist):
            messages.error(request, 'Invalid employee or project selected!')
            return redirect('create_assignment')
        
        # Validate and collect hardware items - Asset Number is primary
        valid_hardware = []
        errors = []
        
        for item in hardware_items:
            hardware_type_id = item.get('hardware_type_id')
            asset_number = item.get('asset_number', '').strip()
            
            if not asset_number:
                errors.append(f"Asset number is required for hardware type {item.get('hardware_type_name', 'Unknown')}")
                continue
            
            try:
                # Find by asset number (primary identifier)
                hardware = Hardware.objects.get(
                    hardware_type_id=hardware_type_id,
                    asset_number=asset_number,
                    status='available',
                    created_by=request.user
                )
                valid_hardware.append(hardware)
            except Hardware.DoesNotExist:
                hardware_type = HardwareType.objects.get(id=hardware_type_id)
                errors.append(f"Asset '{asset_number}' (Type: {hardware_type.name}) not found or not available")
        
        if errors:
            for error in errors[:5]:
                messages.error(request, error)
            if len(errors) > 5:
                messages.error(request, f'...and {len(errors) - 5} more errors')
            return redirect('create_assignment')
        
        if not valid_hardware:
            messages.error(request, 'No valid hardware items to assign!')
            return redirect('create_assignment')
        
        # Create assignment
        assignment = HardwareAssignment.objects.create(
            employee=employee,
            project=project,
            assigned_by=request.user,
            expected_return_date=expected_return_date,
            exam_city=exam_city,
            exam_center_name=exam_center_name,  # New field
            notes=notes
        )
        
        # Create assignment items
        hardware_details = []
        for hardware in valid_hardware:
            HardwareAssignmentItem.objects.create(
                assignment=assignment,
                hardware=hardware,
                quantity=1,
                condition_at_assignment='Assigned for exam duty'
            )
            hardware.status = 'assigned'
            hardware.save()
            
            hardware_details.append({
                'type': hardware.hardware_type.name,
                'asset_number': hardware.asset_number,
                'serial_number': hardware.serial_number,
                'model': hardware.model_name,
                'brand': hardware.brand or 'N/A'
            })
        
        # Send email to employee
        try:
            send_assignment_email(assignment, employee, project, hardware_details)
            messages.success(request, f'Assignment created successfully with {len(valid_hardware)} hardware item(s)! Email sent to {employee.email}')
        except Exception as e:
            messages.success(request, f'Assignment created successfully with {len(valid_hardware)} hardware item(s)! But email could not be sent: {str(e)}')
        
        return redirect('view_assignments')
    
    # Get hardware with asset numbers for the template
    for hw_type in hardware_types:
        hw_type.hardware_list = Hardware.objects.filter(
            hardware_type=hw_type,
            status='available',
            created_by=request.user
        ).values('id', 'asset_number', 'serial_number')
    
    context = {
        'employees': employees,
        'projects': projects,
        'hardware_types': hardware_types,
        'today': timezone.now().date(),
    }
    return render(request, 'manager/create_assignment.html', context)


def send_assignment_email(assignment, employee, project, hardware_details):
    """Send assignment details email to employee"""
    from django.core.mail import send_mail
    from django.conf import settings
    
    try:
        employee_name = employee.get_full_name() or employee.username
        manager_name = assignment.assigned_by.get_full_name() or assignment.assigned_by.username
        
        # Build hardware list for email
        hardware_list_html = ''
        hardware_list_text = ''
        for idx, hw in enumerate(hardware_details, 1):
            hw_type = hw.get('type', 'Unknown')
            asset_number = hw.get('asset_number', 'N/A')
            
            hardware_list_html += f"""
                <tr>
                    <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;">{idx}</td>
                    <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;"><strong>{hw_type}</strong></td>
                    <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;"><code style="background: #f8f9fa; padding: 2px 6px; border-radius: 4px;">{asset_number}</code></td>
                </tr>
            """
            hardware_list_text += f"{idx}. {hw_type} - Asset: {asset_number}\n"
        
        subject = f'Hardware Assignment - {project.project_name} - {assignment.assignment_id}'
        
        # Get exam city safely
        exam_city = getattr(assignment, 'exam_city', 'Not specified')
        if not exam_city:
            exam_city = 'Not specified'
        
        # Get exam center name
        exam_center_name = getattr(assignment, 'exam_center_name', None)
        if not exam_center_name:
            exam_center_name = 'Not specified'
        
        # Format dates safely
        assigned_date = ''
        expected_return_date = ''
        
        if assignment.assigned_date:
            if hasattr(assignment.assigned_date, 'strftime'):
                assigned_date = assignment.assigned_date.strftime('%d %B %Y')
            else:
                assigned_date = str(assignment.assigned_date)
        else:
            assigned_date = 'N/A'
        
        if assignment.expected_return_date:
            if hasattr(assignment.expected_return_date, 'strftime'):
                expected_return_date = assignment.expected_return_date.strftime('%d %B %Y')
            else:
                expected_return_date = str(assignment.expected_return_date)
        else:
            expected_return_date = 'N/A'
        
        html_message = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 700px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; padding: 25px; text-align: center; border-radius: 8px 8px 0 0; }}
                .header h2 {{ margin: 0; font-weight: 300; }}
                .content {{ background: #ffffff; padding: 30px; border: 1px solid #e9ecef; border-top: none; border-radius: 0 0 8px 8px; }}
                .info-box {{ background: #f8f9fa; padding: 15px 20px; margin: 15px 0; border-radius: 6px; border-left: 4px solid #3498db; }}
                .info-box h6 {{ margin: 0 0 5px 0; color: #495057; }}
                .table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 14px; }}
                .table th {{ background: #2c3e50; color: white; padding: 10px 12px; text-align: left; }}
                .table td {{ padding: 10px 12px; border-bottom: 1px solid #e9ecef; }}
                .table tr:hover {{ background: #f8f9fa; }}
                .badge {{ display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; }}
                .badge-success {{ background: #28a745; color: white; }}
                .badge-primary {{ background: #3498db; color: white; }}
                .badge-warning {{ background: #ffc107; color: #212529; }}
                .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #e9ecef; color: #6c757d; font-size: 12px; text-align: center; }}
                .btn {{ display: inline-block; padding: 10px 24px; background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; text-decoration: none; border-radius: 6px; margin: 10px 0; }}
                .btn:hover {{ opacity: 0.9; }}
                .alert {{ padding: 12px 16px; border-radius: 6px; margin: 10px 0; }}
                .alert-info {{ background: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; }}
                .alert-warning {{ background: #fff3cd; border: 1px solid #ffeaa7; color: #856404; }}
                .highlight {{ background: #e8f4fc; padding: 2px 6px; border-radius: 4px; }}
                code {{ background: #f8f9fa; padding: 2px 6px; border-radius: 4px; font-family: 'Courier New', monospace; font-size: 13px; }}
                @media (max-width: 600px) {{
                    .table {{ font-size: 12px; }}
                    .table th, .table td {{ padding: 6px 8px; }}
                    .content {{ padding: 15px; }}
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>📋 Hardware Assignment Notification</h2>
                </div>
                <div class="content">
                    <p>Dear <strong>{employee_name}</strong>,</p>
                    
                    <p>You have been assigned hardware for the upcoming examination. Please review the details below.</p>
                    
                    <div class="info-box">
                        <h6>📌 Assignment Information</h6>
                        <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                            <tr>
                                <td style="padding: 4px 0; width: 35%;"><strong>Assignment ID:</strong></td>
                                <td style="padding: 4px 0;"><code>{assignment.assignment_id}</code></td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Project:</strong></td>
                                <td style="padding: 4px 0;">{project.project_name} ({project.project_id})</td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Exam City:</strong></td>
                                <td style="padding: 4px 0;"><span class="badge badge-success">{exam_city}</span></td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Exam Center:</strong></td>
                                <td style="padding: 4px 0;"><span class="badge badge-primary">{exam_center_name}</span></td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Assigned Date:</strong></td>
                                <td style="padding: 4px 0;">{assigned_date}</td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Expected Return:</strong></td>
                                <td style="padding: 4px 0;">{expected_return_date}</td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Assigned By:</strong></td>
                                <td style="padding: 4px 0;">{manager_name}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <h6 style="margin-top: 20px; margin-bottom: 10px;">🖥️ Assigned Hardware Items</h6>
                    <div style="overflow-x: auto;">
                        <table class="table">
                            <thead>
                                <tr>
                                    <th style="width: 40px;">#</th>
                                    <th>Hardware Type</th>
                                    <th>Asset Number</th>
                                </tr>
                            </thead>
                            <tbody>
                                {hardware_list_html}
                            </tbody>
                        </table>
                    </div>
                    <p style="margin: 10px 0;"><span class="badge badge-primary">Total Items: {len(hardware_details)}</span></p>
                    
                    <div class="alert alert-info">
                        <strong>📌 Next Steps:</strong>
                        <ol style="margin: 8px 0 0 20px;">
                            <li>Go to the <a href="http://eduquityinventory.co.in/" style="color: #3498db; text-decoration: none; font-weight: 600;">Eduquity Hardware Portal</a></li>
                            <li>Navigate to <strong>"My Assignments"</strong> section</li>
                            <li>Click <strong>"Enter Asset"</strong> to input the Asset Numbers from your physical devices</li>
                            <li>Your manager will verify the entries</li>
                        </ol>
                    </div>
                    
                    <div class="alert alert-warning">
                        <strong>⚠️ Important Instructions:</strong>
                        <ul style="margin: 8px 0 0 20px;">
                            <li>Asset Numbers must be entered accurately from the physical devices</li>
                            <li>Asset Number is the primary identifier for verification</li>
                            <li>Keep the hardware safe and in good condition</li>
                            <li>Return all hardware before the due date: <strong>{expected_return_date}</strong></li>
                            <li><strong>Bring hardware to {exam_center_name} ({exam_city}) for the exam</strong></li>
                        </ul>
                    </div>
                    
                    <p style="margin-top: 20px;">
                        <a href="http://eduquityinventory.co.in/" class="btn">🚀 Go to Hardware Portal</a>
                    </p>
                    
                    <div class="footer">
                        <p><strong>Eduquity Hardware Management Team</strong><br>
                        Established in 2000 - Thought-leader in the Indian assessment industry</p>
                        <p><em>This is an automated email. Please do not reply to this message.</em></p>
                        <p style="font-size: 11px;">If you have any issues, please contact your manager: {manager_name}</p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        plain_message = f"""
        HARDWARE ASSIGNMENT NOTIFICATION
        ================================
        
        Dear {employee_name},
        
        You have been assigned hardware for the upcoming examination.
        
        Assignment Information:
        -----------------------
        Assignment ID: {assignment.assignment_id}
        Project: {project.project_name} ({project.project_id})
        Exam City: {exam_city}
        Exam Center: {exam_center_name}
        Assigned Date: {assigned_date}
        Expected Return: {expected_return_date}
        Assigned By: {manager_name}
        
        Assigned Hardware Items:
        -----------------------
        {hardware_list_text}
        
        Total Items: {len(hardware_details)}
        
        Next Steps:
        ----------
        1. Go to the Eduquity Hardware Portal
        2. Navigate to "My Assignments" section
        3. Click "Enter Asset" to input the Asset Numbers from your physical devices
        4. Your manager will verify the entries
        
        Important Instructions:
        -----------------------
        - Asset Numbers must be entered accurately from the physical devices
        - Asset Number is the primary identifier for verification
        - Keep the hardware safe and in good condition
        - Return all hardware before the due date: {expected_return_date}
        - Bring hardware to {exam_center_name} ({exam_city}) for the exam
        
        If you have any issues, please contact your manager: {manager_name}
        
        ---
        Eduquity Hardware Management Team
        """
        
        # Send the email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[employee.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        return True
        
    except Exception as e:
        print(f"Email sending failed: {str(e)}")
        raise Exception(f"Email sending failed: {str(e)}")


def send_assignment_email(assignment, employee, project, hardware_details):
    """Send assignment details email to employee"""
    from django.core.mail import send_mail
    from django.conf import settings
    
    try:
        employee_name = employee.get_full_name() or employee.username
        manager_name = assignment.assigned_by.get_full_name() or assignment.assigned_by.username
        
        # Build hardware list for email
        hardware_list_html = ''
        hardware_list_text = ''
        for idx, hw in enumerate(hardware_details, 1):
            # Safely get values with defaults - only Hardware Type and Asset Number
            hw_type = hw.get('type', 'Unknown')
            asset_number = hw.get('asset_number', 'N/A')
            
            hardware_list_html += f"""
                <tr>
                    <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;">{idx}</td>
                    <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;"><strong>{hw_type}</strong></td>
                    <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;"><code style="background: #f8f9fa; padding: 2px 6px; border-radius: 4px;">{asset_number}</code></td>
                </tr>
            """
            hardware_list_text += f"{idx}. {hw_type} - Asset: {asset_number}\n"
        
        subject = f'Hardware Assignment - {project.project_name} - {assignment.assignment_id}'
        
        # Get exam city safely - it's a string field, not a date
        exam_city = getattr(assignment, 'exam_city', 'Not specified')
        if not exam_city:
            exam_city = 'Not specified'
        
        # Format dates safely - check if they are date objects
        assigned_date = ''
        expected_return_date = ''
        
        if assignment.assigned_date:
            if hasattr(assignment.assigned_date, 'strftime'):
                assigned_date = assignment.assigned_date.strftime('%d %B %Y')
            else:
                assigned_date = str(assignment.assigned_date)
        else:
            assigned_date = 'N/A'
        
        if assignment.expected_return_date:
            if hasattr(assignment.expected_return_date, 'strftime'):
                expected_return_date = assignment.expected_return_date.strftime('%d %B %Y')
            else:
                expected_return_date = str(assignment.expected_return_date)
        else:
            expected_return_date = 'N/A'
        
        html_message = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 700px; margin: 0 auto; padding: 20px; }}
                .header {{ background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; padding: 25px; text-align: center; border-radius: 8px 8px 0 0; }}
                .header h2 {{ margin: 0; font-weight: 300; }}
                .content {{ background: #ffffff; padding: 30px; border: 1px solid #e9ecef; border-top: none; border-radius: 0 0 8px 8px; }}
                .info-box {{ background: #f8f9fa; padding: 15px 20px; margin: 15px 0; border-radius: 6px; border-left: 4px solid #3498db; }}
                .info-box h6 {{ margin: 0 0 5px 0; color: #495057; }}
                .table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 14px; }}
                .table th {{ background: #2c3e50; color: white; padding: 10px 12px; text-align: left; }}
                .table td {{ padding: 10px 12px; border-bottom: 1px solid #e9ecef; }}
                .table tr:hover {{ background: #f8f9fa; }}
                .badge {{ display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; }}
                .badge-success {{ background: #28a745; color: white; }}
                .badge-primary {{ background: #3498db; color: white; }}
                .badge-warning {{ background: #ffc107; color: #212529; }}
                .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #e9ecef; color: #6c757d; font-size: 12px; text-align: center; }}
                .btn {{ display: inline-block; padding: 10px 24px; background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; text-decoration: none; border-radius: 6px; margin: 10px 0; }}
                .btn:hover {{ opacity: 0.9; }}
                .alert {{ padding: 12px 16px; border-radius: 6px; margin: 10px 0; }}
                .alert-info {{ background: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; }}
                .alert-warning {{ background: #fff3cd; border: 1px solid #ffeaa7; color: #856404; }}
                .highlight {{ background: #e8f4fc; padding: 2px 6px; border-radius: 4px; }}
                code {{ background: #f8f9fa; padding: 2px 6px; border-radius: 4px; font-family: 'Courier New', monospace; font-size: 13px; }}
                @media (max-width: 600px) {{
                    .table {{ font-size: 12px; }}
                    .table th, .table td {{ padding: 6px 8px; }}
                    .content {{ padding: 15px; }}
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>📋 Hardware Assignment Notification</h2>
                </div>
                <div class="content">
                    <p>Dear <strong>{employee_name}</strong>,</p>
                    
                    <p>You have been assigned hardware for the upcoming examination. Please review the details below.</p>
                    
                    <div class="info-box">
                        <h6>📌 Assignment Information</h6>
                        <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                            <tr>
                                <td style="padding: 4px 0; width: 35%;"><strong>Assignment ID:</strong></td>
                                <td style="padding: 4px 0;"><code>{assignment.assignment_id}</code></td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Project:</strong></td>
                                <td style="padding: 4px 0;">{project.project_name} ({project.project_id})</td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Exam City:</strong></td>
                                <td style="padding: 4px 0;"><span class="badge badge-success">{exam_city}</span></td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Assigned Date:</strong></td>
                                <td style="padding: 4px 0;">{assigned_date}</td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Expected Return:</strong></td>
                                <td style="padding: 4px 0;">{expected_return_date}</td>
                            </tr>
                            <tr>
                                <td style="padding: 4px 0;"><strong>Assigned By:</strong></td>
                                <td style="padding: 4px 0;">{manager_name}</td>
                            </tr>
                        </table>
                    </div>
                    
                    <h6 style="margin-top: 20px; margin-bottom: 10px;">🖥️ Assigned Hardware Items</h6>
                    <div style="overflow-x: auto;">
                        <table class="table">
                            <thead>
                                <tr>
                                    <th style="width: 40px;">#</th>
                                    <th>Hardware Type</th>
                                    <th>Asset Number</th>
                                </tr>
                            </thead>
                            <tbody>
                                {hardware_list_html}
                            </tbody>
                        </table>
                    </div>
                    <p style="margin: 10px 0;"><span class="badge badge-primary">Total Items: {len(hardware_details)}</span></p>
                    
                    <div class="alert alert-info">
                        <strong>📌 Next Steps:</strong>
                        <ol style="margin: 8px 0 0 20px;">
                            <li>Go to the <a href="http://eduquityinventory.co.in/" style="color: #3498db; text-decoration: none; font-weight: 600;">Eduquity Hardware Portal</a></li>
                            <li>Navigate to <strong>"My Assignments"</strong> section</li>
                            <li>Click <strong>"Enter Asset"</strong> to input the Asset Numbers from your physical devices</li>
                            <li>Your manager will verify the entries</li>
                        </ol>
                    </div>
                    
                    <div class="alert alert-warning">
                        <strong>⚠️ Important Instructions:</strong>
                        <ul style="margin: 8px 0 0 20px;">
                            <li>Asset Numbers must be entered accurately from the physical devices</li>
                            <li>Asset Number is the primary identifier for verification</li>
                            <li>Keep the hardware safe and in good condition</li>
                            <li>Return all hardware before the due date: <strong>{expected_return_date}</strong></li>
                            <li><strong>Bring hardware to {exam_city} for the exam</strong></li>
                        </ul>
                    </div>
                    
                    <p style="margin-top: 20px;">
                        <a href="http://eduquityinventory.co.in/" class="btn">🚀 Go to Hardware Portal</a>
                    </p>
                    
                    <div class="footer">
                        <p><strong>Eduquity Hardware Management Team</strong><br>
                        Established in 2000 - Thought-leader in the Indian assessment industry</p>
                        <p><em>This is an automated email. Please do not reply to this message.</em></p>
                        <p style="font-size: 11px;">If you have any issues, please contact your manager: {manager_name}</p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        plain_message = f"""
        HARDWARE ASSIGNMENT NOTIFICATION
        ================================
        
        Dear {employee_name},
        
        You have been assigned hardware for the upcoming examination.
        
        Assignment Information:
        -----------------------
        Assignment ID: {assignment.assignment_id}
        Project: {project.project_name} ({project.project_id})
        Exam City: {exam_city}
        Assigned Date: {assigned_date}
        Expected Return: {expected_return_date}
        Assigned By: {manager_name}
        
        Assigned Hardware Items:
        -----------------------
        {hardware_list_text}
        
        Total Items: {len(hardware_details)}
        
        Next Steps:
        ----------
        1. Go to the Eduquity Hardware Portal
        2. Navigate to "My Assignments" section
        3. Click "Enter Asset" to input the Asset Numbers from your physical devices
        4. Your manager will verify the entries
        
        Important Instructions:
        -----------------------
        - Asset Numbers must be entered accurately from the physical devices
        - Asset Number is the primary identifier for verification
        - Keep the hardware safe and in good condition
        - Return all hardware before the due date: {expected_return_date}
        - Bring hardware to {exam_city} for the exam
        
        If you have any issues, please contact your manager: {manager_name}
        
        ---
        Eduquity Hardware Management Team
        """
        
        # Send the email
        send_mail(
            subject=subject,
            message=plain_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[employee.email],
            html_message=html_message,
            fail_silently=False,
        )
        
        return True
        
    except Exception as e:
        # Log the error for debugging
        print(f"Email sending failed: {str(e)}")
        # Re-raise to be caught by the calling function
        raise Exception(f"Email sending failed: {str(e)}")


from django.db.models import Count, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone

from django.utils import timezone
from datetime import timedelta
from django.db.models import Q, Count

@login_required
def view_assignments(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    all_assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user
    ).order_by('-assigned_date')
    
    today = timezone.now().date()
    due_soon_date = today + timedelta(days=3)
    
    if request.GET.get('pending') == '1':
        assignments = all_assignments.filter(
            actual_return_date__isnull=True,
            expected_return_date__lte=due_soon_date
        )
    else:
        assignments = all_assignments  
    
    total_assignments = all_assignments.count()
    active_assignments = all_assignments.filter(actual_return_date__isnull=True).count()
    returned_assignments = all_assignments.filter(actual_return_date__isnull=False).count()
    
    overdue_count = all_assignments.filter(
        actual_return_date__isnull=True,
        expected_return_date__lt=today  
    ).count()
    
    due_soon_count = all_assignments.filter(
        actual_return_date__isnull=True,
        expected_return_date__gte=today,  
        expected_return_date__lte=due_soon_date
    ).count()
    
    pending_return_count = overdue_count + due_soon_count
    
    print(f"=== DEBUG ===")
    print(f"Today: {today}")
    print(f"Due soon date: {due_soon_date}")
    print(f"Active assignments: {active_assignments}")
    print(f"Overdue count: {overdue_count}")
    print(f"Due soon count: {due_soon_count}")
    print(f"Pending return count: {pending_return_count}")
    print(f"=============")
    
    active_assignments_list = all_assignments.filter(actual_return_date__isnull=True)
    for assignment in active_assignments_list:
        status = "OVERDUE" if assignment.expected_return_date < today else "DUE SOON" if assignment.expected_return_date <= due_soon_date else "FUTURE"
        print(f"Assignment {assignment.id}: Expected: {assignment.expected_return_date}, Status: {status}")
    
    unique_cities = all_assignments.exclude(
        exam_city__isnull=True
    ).exclude(
        exam_city__exact=''
    ).values('exam_city').distinct().count()
    
    context = {
        'assignments': assignments,
        'total_assignments': total_assignments,
        'active_assignments': active_assignments,
        'returned_assignments': returned_assignments,
        'pending_return_count': pending_return_count,
        'overdue_count': overdue_count,
        'due_soon_count': due_soon_count,
        'unique_cities': unique_cities,
        'today': today,
        'due_soon_date': due_soon_date,
    }
    return render(request, 'manager/view_assignments.html', context)



@login_required
def assignment_details(request, assignment_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(
        HardwareAssignment, 
        id=assignment_id, 
        assigned_by=request.user
    )
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    verified_count = 0
    pending_count = 0
    not_entered_count = 0
    
    for item in items:
        # FIX: Use asset_entry instead of serial_entry
        item.has_asset_entry = hasattr(item, 'asset_entry')
        if item.has_asset_entry:
            item.entered_asset = item.asset_entry.entered_asset_number
            item.is_verified = item.asset_entry.verified
            if item.asset_entry.verified:
                verified_count += 1
            else:
                pending_count += 1
        else:
            not_entered_count += 1
    
    # Get hardware details with asset numbers (primary)
    for item in items:
        item.asset_number = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
        item.serial_number = item.hardware.serial_number
    
    # Get exam center name
    exam_center_name = getattr(assignment, 'exam_center_name', None)
    
    context = {
        'assignment': assignment,
        'items': items,
        'verified_count': verified_count,
        'pending_count': pending_count,
        'not_entered_count': not_entered_count,
        'total_items': items.count(),
        'exam_center_name': exam_center_name,
    }
    return render(request, 'manager/assignment_details.html', context)
    

@login_required
def return_assignment(request, assignment_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, assigned_by=request.user)
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    # Add asset number to each item (primary identifier)
    for item in items:
        item.asset_number = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
        item.serial_number = item.hardware.serial_number
    
    if request.method == 'POST':
        verification_status = []
        all_verified = True
        mismatch_count = 0
        missing_count = 0
        
        for item in items:
            returned_asset = request.POST.get(f'returned_asset_{item.id}', '').strip()
            condition_notes = request.POST.get(f'condition_notes_{item.id}', '')
            
            if not returned_asset:
                missing_count += 1
                all_verified = False
                verification_status.append({
                    'item': item,
                    'status': 'missing',
                    'message': f'Return Asset Number not entered for {item.hardware.hardware_type.name} (Asset: {item.asset_number})'
                })
                continue
            
            # Check if returned asset matches the assigned asset number (primary verification)
            is_match = (returned_asset == item.asset_number)
            
            if not is_match:
                mismatch_count += 1
                all_verified = False
                verification_status.append({
                    'item': item,
                    'status': 'mismatch',
                    'message': f'Returned Asset "{returned_asset}" does not match expected Asset "{item.asset_number}"'
                })
            
            item.condition_at_return = condition_notes
            item.save()
        
        if not all_verified:
            error_messages = []
            if missing_count > 0:
                error_messages.append(f'{missing_count} item(s) missing return Asset Numbers')
            if mismatch_count > 0:
                error_messages.append(f'{mismatch_count} item(s) have mismatched Asset Numbers')
            
            messages.error(request, 'Return verification failed: ' + '; '.join(error_messages))
            
            request.session['verification_status'] = verification_status
            return redirect('return_assignment', assignment_id=assignment.id)
        
        # Update hardware status to available
        for item in items:
            item.hardware.status = 'available'
            item.hardware.save()
        
        assignment.actual_return_date = timezone.now().date()
        assignment.save()
        
        # Send email to employee about return
        try:
            send_return_confirmation_email(assignment, items)
            messages.success(request, f'Assignment returned successfully! All {items.count()} hardware items verified by Asset Number and marked as available. Email sent to {assignment.employee.email}')
        except Exception as e:
            messages.success(request, f'Assignment returned successfully! All {items.count()} hardware items verified by Asset Number and marked as available. But email could not be sent: {str(e)}')
        
        return redirect('view_assignments')
    
    verification_status = request.session.pop('verification_status', [])
    
    context = {
        'assignment': assignment,
        'items': items,
        'verification_status': verification_status,
        'today': timezone.now().date(),
    }
    return render(request, 'manager/return_assignment.html', context)



def send_return_confirmation_email(assignment, items):
    """Send return confirmation email to employee"""
    from django.core.mail import send_mail
    from django.conf import settings
    
    employee = assignment.employee
    employee_name = employee.get_full_name() or employee.username
    manager_name = assignment.assigned_by.get_full_name() or assignment.assigned_by.username
    
    # Build hardware list for email
    hardware_list_html = ''
    hardware_list_text = ''
    for idx, item in enumerate(items, 1):
        asset_number = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
        hw_type = item.hardware.hardware_type.name
        condition = item.condition_at_return or 'Good condition'
        
        hardware_list_html += f"""
            <tr>
                <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;">{idx}</td>
                <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;"><strong>{hw_type}</strong></td>
                <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;"><code style="background: #f8f9fa; padding: 2px 6px; border-radius: 4px;">{asset_number}</code></td>
                <td style="padding: 8px 12px; border-bottom: 1px solid #e9ecef;">{condition}</td>
            </tr>
        """
        hardware_list_text += f"{idx}. {hw_type} - Asset: {asset_number} | Condition: {condition}\n"
    
    subject = f'Hardware Return Confirmation - {assignment.project.project_name}'
    
    html_message = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 700px; margin: 0 auto; padding: 20px; }}
            .header {{ background: linear-gradient(90deg, #28a745 0%, #20c997 100%); color: white; padding: 25px; text-align: center; border-radius: 8px 8px 0 0; }}
            .header h2 {{ margin: 0; font-weight: 300; }}
            .content {{ background: #ffffff; padding: 30px; border: 1px solid #e9ecef; border-top: none; border-radius: 0 0 8px 8px; }}
            .info-box {{ background: #f8f9fa; padding: 15px 20px; margin: 15px 0; border-radius: 6px; border-left: 4px solid #28a745; }}
            .info-box h6 {{ margin: 0 0 5px 0; color: #495057; }}
            .table {{ width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 14px; }}
            .table th {{ background: #28a745; color: white; padding: 10px 12px; text-align: left; }}
            .table td {{ padding: 10px 12px; border-bottom: 1px solid #e9ecef; }}
            .table tr:hover {{ background: #f8f9fa; }}
            .badge {{ display: inline-block; padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 600; }}
            .badge-success {{ background: #28a745; color: white; }}
            .badge-info {{ background: #17a2b8; color: white; }}
            .badge-warning {{ background: #ffc107; color: #212529; }}
            .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #e9ecef; color: #6c757d; font-size: 12px; text-align: center; }}
            .btn {{ display: inline-block; padding: 10px 24px; background: linear-gradient(90deg, #28a745 0%, #20c997 100%); color: white; text-decoration: none; border-radius: 6px; margin: 10px 0; }}
            .btn:hover {{ opacity: 0.9; }}
            .alert {{ padding: 12px 16px; border-radius: 6px; margin: 10px 0; }}
            .alert-success {{ background: #d4edda; border: 1px solid #c3e6cb; color: #155724; }}
            .alert-info {{ background: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; }}
            code {{ background: #f8f9fa; padding: 2px 6px; border-radius: 4px; font-family: 'Courier New', monospace; font-size: 13px; }}
            @media (max-width: 600px) {{
                .table {{ font-size: 12px; }}
                .table th, .table td {{ padding: 6px 8px; }}
                .content {{ padding: 15px; }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h2>✅ Hardware Return Confirmation</h2>
            </div>
            <div class="content">
                <p>Dear <strong>{employee_name}</strong>,</p>
                
                <div class="alert alert-success">
                    <strong>✅ Assignment Returned Successfully!</strong>
                    <br>
                    All hardware items have been returned and verified by <strong>{manager_name}</strong>.
                </div>
                
                <div class="info-box">
                    <h6>📌 Assignment Information</h6>
                    <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                        <tr>
                            <td style="padding: 4px 0; width: 35%;"><strong>Assignment ID:</strong></td>
                            <td style="padding: 4px 0;"><code>{assignment.assignment_id}</code></td>
                        </tr>
                        <tr>
                            <td style="padding: 4px 0;"><strong>Project:</strong></td>
                            <td style="padding: 4px 0;">{assignment.project.project_name} ({assignment.project.project_id})</td>
                        </tr>
                        <tr>
                            <td style="padding: 4px 0;"><strong>Exam City:</strong></td>
                            <td style="padding: 4px 0;"><span class="badge badge-info">{assignment.exam_city}</span></td>
                        </tr>
                        <tr>
                            <td style="padding: 4px 0;"><strong>Return Date:</strong></td>
                            <td style="padding: 4px 0;"><span class="badge badge-success">{assignment.actual_return_date.strftime('%d %B %Y')}</span></td>
                        </tr>
                        <tr>
                            <td style="padding: 4px 0;"><strong>Verified By:</strong></td>
                            <td style="padding: 4px 0;">{manager_name}</td>
                        </tr>
                    </table>
                </div>
                
                <h6 style="margin-top: 20px; margin-bottom: 10px;">🖥️ Returned Hardware Items</h6>
                <div style="overflow-x: auto;">
                    <table class="table">
                        <thead>
                            <tr>
                                <th style="width: 40px;">#</th>
                                <th>Hardware Type</th>
                                <th>Asset Number</th>
                                <th>Condition at Return</th>
                            </tr>
                        </thead>
                        <tbody>
                            {hardware_list_html}
                        </tbody>
                    </table>
                </div>
                <p style="margin: 10px 0;"><span class="badge badge-success">Total Items Returned: {len(items)}</span></p>
                
                <div class="alert alert-info">
                    <strong>📌 What's Next:</strong>
                    <ul style="margin: 8px 0 0 20px;">
                        <li>You have successfully completed this assignment</li>
                        <li>All hardware has been returned and marked as <strong>Available</strong></li>
                        <li>You can view your completed assignments in the portal</li>
                        <li>Thank you for taking care of the hardware</li>
                    </ul>
                </div>
                
                <p style="margin-top: 20px;">
                    <a href="http://eduquityinventory.co.in/" class="btn">🚀 Go to Hardware Portal</a>
                </p>
                
                <div class="footer">
                    <p><strong>Eduquity Hardware Management Team</strong><br>
                    Established in 2000 - Thought-leader in the Indian assessment industry</p>
                    <p><em>This is an automated email. Please do not reply to this message.</em></p>
                    <p style="font-size: 11px;">For any questions, please contact your manager: {manager_name}</p>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    
    plain_message = f"""
    HARDWARE RETURN CONFIRMATION
    ============================
    
    Dear {employee_name},
    
    ✅ Assignment Returned Successfully!
    
    All hardware items have been returned and verified by {manager_name}.
    
    Assignment Information:
    -----------------------
    Assignment ID: {assignment.assignment_id}
    Project: {assignment.project.project_name} ({assignment.project.project_id})
    Exam City: {assignment.exam_city}
    Return Date: {assignment.actual_return_date.strftime('%d %B %Y')}
    Verified By: {manager_name}
    
    Returned Hardware Items:
    -----------------------
    {hardware_list_text}
    
    Total Items Returned: {len(items)}
    
    What's Next:
    -----------
    - You have successfully completed this assignment
    - All hardware has been returned and marked as Available
    - You can view your completed assignments in the portal
    - Thank you for taking care of the hardware
    
    For any questions, please contact your manager: {manager_name}
    
    ---
    Eduquity Hardware Management Team
    """
    
    send_mail(
        subject=subject,
        message=plain_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[employee.email],
        html_message=html_message,
        fail_silently=False,
    )

# ============== SERIAL NUMBER VIEWS ==============


@login_required
def view_serial_entries(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True
    ).prefetch_related(
        'hardwareassignmentitem_set__hardware__hardware_type',
        'hardwareassignmentitem_set__asset_entry__entered_by',  # This is correct - using related_name
        'hardwareassignmentitem_set__asset_entry__verified_by'
    ).order_by('-assigned_date')
    
    total_verified = 0
    total_matched = 0
    total_mismatch = 0
    total_pending = 0
    
    for assignment in assignments:
        items = assignment.hardwareassignmentitem_set.all()
        assignment.total_items = items.count()
        assignment.verified_count = 0
        assignment.matched_count = 0
        assignment.mismatch_count = 0
        assignment.pending_count = 0
        assignment.asset_entries = []
        
        for item in items:
            expected_asset = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
            
            try:
                # CRITICAL FIX: Use the correct field name
                # The model has hardware_item, but related_name is 'asset_entry'
                # So we access it as item.asset_entry (using related_name)
                asset_entry = item.asset_entry  # This works because related_name='asset_entry'
                
                # Check if entered asset number matches expected asset number
                is_match = (asset_entry.entered_asset_number == expected_asset)
                
                entry_data = {
                    'id': asset_entry.id,
                    'item_id': item.id,
                    'entered_asset_number': asset_entry.entered_asset_number,
                    'expected_asset': expected_asset,
                    'hardware_type': item.hardware.hardware_type.name,
                    'model': item.hardware.model_name,
                    'asset_number': expected_asset,
                    'entered_by': asset_entry.entered_by,
                    'entered_at': asset_entry.entered_at,
                    'verified': asset_entry.verified,
                    'verified_by': asset_entry.verified_by,
                    'verified_at': asset_entry.verified_at,
                    'is_match': is_match,
                    'match_status': 'verified' if asset_entry.verified else ('matched' if is_match else 'mismatch'),
                }
                assignment.asset_entries.append(entry_data)
                
                if asset_entry.verified:
                    assignment.verified_count += 1
                    total_verified += 1
                else:
                    if is_match:
                        assignment.matched_count += 1
                        total_matched += 1
                    else:
                        assignment.mismatch_count += 1
                        total_mismatch += 1
                        
            except HardwareAssetEntry.DoesNotExist:
                entry_data = {
                    'id': None,
                    'item_id': item.id,
                    'entered_asset_number': None,
                    'expected_asset': expected_asset,
                    'hardware_type': item.hardware.hardware_type.name,
                    'model': item.hardware.model_name,
                    'asset_number': expected_asset,
                    'entered_by': None,
                    'entered_at': None,
                    'verified': False,
                    'verified_by': None,
                    'verified_at': None,
                    'is_match': False,
                    'match_status': 'pending',
                }
                assignment.asset_entries.append(entry_data)
                assignment.pending_count += 1
                total_pending += 1
    
    total_assignments = assignments.count()
    
    context = {
        'assignments': assignments,
        'total_assignments': total_assignments,
        'total_verified': total_verified,
        'total_matched': total_matched,
        'total_mismatch': total_mismatch,
        'total_pending': total_pending,
    }
    return render(request, 'manager/view_asset_entries.html', context)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

@login_required
def export_all_employees_hardware(request):
    """Export current active hardware data for all employees to Excel"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    # Get ONLY ACTIVE assignments (not returned)
    assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True  # Only active assignments
    ).order_by('-assigned_date')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Current Active Hardware"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10)
    subheader_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    success_fill = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")
    warning_fill = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")
    danger_fill = PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid")
    info_fill = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    title_cell = ws.cell(row=1, column=1, value="EDUQUITY HARDWARE MANAGEMENT SYSTEM")
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = center_alignment
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    subtitle_cell = ws.cell(row=2, column=1, value=f"CURRENT ACTIVE HARDWARE REPORT - Generated on {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    subtitle_cell.font = Font(size=11, italic=True)
    subtitle_cell.alignment = center_alignment
    
    # Summary Row
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    summary_title = ws.cell(row=3, column=1, value="▶ CURRENT ACTIVE HARDWARE SUMMARY")
    summary_title.font = Font(bold=True, size=12)
    summary_title.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # Prepare data - ONLY ACTIVE HARDWARE
    employee_data = {}
    total_hardware = 0
    total_verified = 0
    total_matched = 0
    total_mismatch = 0
    total_pending = 0
    total_overdue = 0
    total_due_soon = 0
    
    today = timezone.now().date()
    due_soon_date = today + timezone.timedelta(days=3)
    
    for assignment in assignments:
        employee = assignment.employee
        emp_id = employee.id
        if emp_id not in employee_data:
            employee_data[emp_id] = {
                'employee_name': employee.get_full_name() or employee.username,
                'email': employee.email,
                'phone': employee.phone or '-',
                'created_at': employee.date_joined.strftime("%d/%m/%Y") if employee.date_joined else '-',
                'hardware_items': []
            }
        
        # Get hardware items for this assignment
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        
        for item in items:
            total_hardware += 1
            
            # Check if overdue or due soon
            is_overdue = assignment.expected_return_date < today
            is_due_soon = assignment.expected_return_date <= due_soon_date and assignment.expected_return_date >= today
            
            if is_overdue:
                total_overdue += 1
            elif is_due_soon:
                total_due_soon += 1
            
            hardware_info = {
                'assignment_id': str(assignment.assignment_id)[:8],
                'exam_city': assignment.exam_city or 'Not Specified',
                'assigned_date': assignment.assigned_date.strftime("%d/%m/%Y"),
                'expected_return': assignment.expected_return_date.strftime("%d/%m/%Y"),
                'is_overdue': is_overdue,
                'is_due_soon': is_due_soon,
                'hardware_type': item.hardware.hardware_type.name,
                'model': item.hardware.model_name,
                'brand': item.hardware.brand or '-',
                'assigned_serial': item.hardware.serial_number,
                'status': item.hardware.get_status_display(),
                'entered_serial': '-',
                'verification_status': 'Pending',
                'verified_by': '-',
                'verified_at': '-',
                'condition': item.condition_at_assignment or '-'
            }
            
            try:
                serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                hardware_info['entered_serial'] = serial_entry.serial_number
                
                if serial_entry.verified:
                    hardware_info['verification_status'] = 'Verified'
                    hardware_info['verified_by'] = serial_entry.verified_by.get_full_name() or serial_entry.verified_by.username if serial_entry.verified_by else '-'
                    hardware_info['verified_at'] = serial_entry.verified_at.strftime("%d/%m/%Y %H:%M") if serial_entry.verified_at else '-'
                    total_verified += 1
                    
                    if serial_entry.serial_number == item.hardware.serial_number:
                        hardware_info['match_status'] = 'Verified - Correct'
                    else:
                        hardware_info['match_status'] = 'Verified - Mismatch'
                else:
                    if serial_entry.serial_number == item.hardware.serial_number:
                        hardware_info['verification_status'] = 'Matched - Pending'
                        total_matched += 1
                    else:
                        hardware_info['verification_status'] = 'Mismatch'
                        total_mismatch += 1
                        
            except HardwareSerialEntry.DoesNotExist:
                total_pending += 1
                hardware_info['verification_status'] = 'Not Entered'
            
            employee_data[emp_id]['hardware_items'].append(hardware_info)
    
    # Write Main Report
    current_row = 5
    
    for emp_id, emp_data in employee_data.items():
        if not emp_data['hardware_items']:
            continue
            
        # Employee Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
        emp_header = ws.cell(row=current_row, column=1, value=f"EMPLOYEE: {emp_data['employee_name']} - {emp_data['email']}")
        emp_header.font = Font(bold=True, size=12, color="FFFFFF")
        emp_header.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        emp_header.alignment = center_alignment
        current_row += 1
        
        # Employee Details
        details = [
            ["📧 Email", emp_data['email']],
            ["📞 Phone", emp_data['phone']],
            ["📅 Employee Since", emp_data['created_at']],
            ["🖥️ Active Hardware", len(emp_data['hardware_items'])]
        ]
        
        for i, (label, value) in enumerate(details):
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        
        current_row += 1
        
        # Hardware Table Header
        headers = ['S.No', 'Assignment ID', 'Exam City', 'Assigned Date', 'Expected Return', 'Return Status',
                   'Hardware Type', 'Model', 'Brand', 'Assigned Serial', 'Entered Serial', 
                   'Verification Status', 'Condition']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_alignment
            cell.border = border
        
        current_row += 1
        
        # Write Hardware Items
        sno = 1
        for item in emp_data['hardware_items']:
            return_status = "Normal"
            row_fill = None
            
            if item['is_overdue']:
                return_status = "⚠️ OVERDUE"
                row_fill = danger_fill
            elif item['is_due_soon']:
                return_status = "⚡ DUE SOON"
                row_fill = warning_fill
            
            row_data = [
                sno, item['assignment_id'], item['exam_city'], item['assigned_date'],
                item['expected_return'], return_status, item['hardware_type'],
                item['model'], item['brand'], item['assigned_serial'], item['entered_serial'],
                item['verification_status'], item['condition']
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = border
                
                # Apply color coding
                if row_fill:
                    cell.fill = row_fill
                elif item['verification_status'] == 'Verified':
                    cell.fill = success_fill
                elif item['verification_status'] == 'Matched - Pending':
                    cell.fill = info_fill
                elif item['verification_status'] == 'Mismatch':
                    cell.fill = danger_fill
                elif item['verification_status'] == 'Not Entered':
                    cell.fill = warning_fill
                
                if col_num in [4, 5, 6, 12]:
                    cell.alignment = center_alignment
            
            sno += 1
            current_row += 1
        
        current_row += 2
    
    # Grand Summary
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
    summary_cell = ws.cell(row=current_row, column=1, value="▶ GRAND SUMMARY - CURRENT ACTIVE HARDWARE")
    summary_cell.font = Font(bold=True, size=14, color="FFFFFF")
    summary_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    summary_cell.alignment = center_alignment
    current_row += 1
    
    summary_data = [
        ["Total Active Employees", len([e for e in employee_data.values() if e['hardware_items']])],
        ["Total Active Hardware Items", total_hardware],
        ["✅ Verified Items", total_verified],
        ["🟦 Matched Items (Pending)", total_matched],
        ["❌ Mismatched Items", total_mismatch],
        ["⏳ Pending Entry Items", total_pending],
        ["⚠️ Overdue Items", total_overdue],
        ["⚡ Due Soon Items", total_due_soon],
        ["📈 Completion Rate", f"{round((total_verified / total_hardware * 100) if total_hardware > 0 else 0, 2)}%"],
        ["👤 Generated By", request.user.get_full_name() or request.user.username],
        ["📅 Generated On", datetime.now().strftime("%d/%m/%Y %H:%M:%S")]
    ]
    
    for label, value in summary_data:
        label_cell = ws.cell(row=current_row, column=1, value=label)
        label_cell.font = Font(bold=True)
        value_cell = ws.cell(row=current_row, column=2, value=value)
        current_row += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, current_row):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 4, 45)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    filename = f"active_hardware_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def verify_asset_entry(request, entry_id):
    """
    Verify a single asset entry
    """
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    asset_entry = get_object_or_404(HardwareAssetEntry, id=entry_id)
    
    # FIX: Use 'hardware_item' instead of 'assignment_item'
    if asset_entry.hardware_item.assignment.assigned_by != request.user:
        messages.error(request, 'Unauthorized access!')
        return redirect('view_serial_entries')
    
    # Check if entered asset number matches expected asset number
    expected_asset = asset_entry.hardware_item.hardware.asset_number
    is_match = asset_entry.entered_asset_number == expected_asset
    
    if is_match:
        # Mark as verified
        asset_entry.verified = True
        asset_entry.verified_by = request.user
        asset_entry.verified_at = timezone.now()
        asset_entry.save()
        
        # Update hardware status
        hardware = asset_entry.hardware_item.hardware
        hardware.status = 'in_use'
        hardware.save()
        
        messages.success(
            request, 
            f'Asset entry verified successfully! Asset {asset_entry.entered_asset_number} is now marked as in use.'
        )
    else:
        messages.error(
            request, 
            f'Cannot verify - Entered asset number "{asset_entry.entered_asset_number}" does not match expected asset number "{expected_asset}"!'
        )
    
    return redirect('view_serial_entries')


@login_required
def verify_serial_entry(request, entry_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    serial_entry = get_object_or_404(HardwareSerialEntry, id=entry_id)
    
    if serial_entry.assignment_item.assignment.assigned_by != request.user:
        messages.error(request, 'Unauthorized access!')
        return redirect('view_serial_entries')
    
    is_match = serial_entry.serial_number == serial_entry.assignment_item.hardware.serial_number
    
    if is_match:
        serial_entry.verified = True
        serial_entry.verified_by = request.user
        serial_entry.verified_at = timezone.now()
        serial_entry.save()
        
        hardware = serial_entry.assignment_item.hardware
        hardware.status = 'in_use'
        hardware.save()
        
        messages.success(request, f'Serial entry verified successfully! Hardware is now marked as in use.')
    else:
        messages.error(request, 'Cannot verify - Serial number does not match the assigned hardware!')
    
    return redirect('view_serial_entries')



@login_required
def verify_all_employee_entries(request, assignment_id):
    """
    Single click verify all pending asset entries for an employee
    """
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(
        HardwareAssignment,
        id=assignment_id,
        assigned_by=request.user,
        actual_return_date__isnull=True
    )
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment)
    verified_count = 0
    skipped_count = 0
    mismatch_count = 0
    no_entry_count = 0
    
    for item in items:
        try:
            # Get the asset entry
            asset_entry = item.asset_entry
            expected_asset = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
            is_match = asset_entry.entered_asset_number == expected_asset
            
            if not asset_entry.verified and is_match:
                # Verify this entry
                asset_entry.verified = True
                asset_entry.verified_by = request.user
                asset_entry.verified_at = timezone.now()
                asset_entry.save()
                
                # Update hardware status
                item.hardware.status = 'in_use'
                item.hardware.save()
                verified_count += 1
                
            elif asset_entry.verified:
                skipped_count += 1
                
            elif not is_match and asset_entry.entered_asset_number:
                mismatch_count += 1
                
        except HardwareAssetEntry.DoesNotExist:
            no_entry_count += 1
            continue
    
    employee_name = assignment.employee.get_full_name() or assignment.employee.username
    
    # Build response messages
    if verified_count > 0:
        messages.success(
            request, 
            f'✅ Successfully verified {verified_count} asset(s) for {employee_name}!'
        )
        
        if mismatch_count > 0:
            messages.warning(
                request,
                f'⚠️ Skipped {mismatch_count} item(s) with asset number mismatch for {employee_name}.'
            )
            
        if skipped_count > 0:
            messages.info(
                request,
                f'ℹ️ {skipped_count} item(s) were already verified.'
            )
            
        if no_entry_count > 0:
            messages.info(
                request,
                f'ℹ️ {no_entry_count} item(s) have no asset entry yet.'
            )
    else:
        if mismatch_count > 0 and no_entry_count == 0:
            messages.error(
                request, 
                f'❌ No items verified. Found {mismatch_count} item(s) with asset number mismatch for {employee_name}.'
            )
        elif skipped_count > 0 and mismatch_count == 0 and no_entry_count == 0:
            messages.warning(
                request, 
                f'ℹ️ All {skipped_count} item(s) are already verified for {employee_name}.'
            )
        elif no_entry_count > 0 and mismatch_count == 0 and skipped_count == 0:
            messages.warning(
                request, 
                f'ℹ️ No asset entries found for {employee_name}. Employee needs to enter asset numbers first.'
            )
        else:
            messages.warning(
                request, 
                f'⚠️ No eligible items found for verification for {employee_name}. Items must have matching asset numbers and not be already verified.'
            )
    
    # Redirect to the asset entries view
    return redirect('view_serial_entries')


@login_required
def manager_verification_status(request):
    """Manager dashboard to see verification status across all assignments"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    total_verified = 0
    total_matched = 0
    total_mismatch = 0
    total_pending = 0
    
    for assignment in assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        total_items = items.count()
        verified_items = 0
        matched_items = 0
        mismatch_items = 0
        pending_items = 0
        
        for item in items:
            try:
                # FIX: Use asset_entry instead of serial_entry
                asset_entry = item.asset_entry
                expected_asset = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
                is_match = (asset_entry.entered_asset_number == expected_asset)
                
                if asset_entry.verified:
                    verified_items += 1
                else:
                    if is_match:
                        matched_items += 1
                    else:
                        mismatch_items += 1
            except HardwareAssetEntry.DoesNotExist:
                pending_items += 1
        
        verified_percentage = (verified_items / total_items * 100) if total_items > 0 else 0
        matched_percentage = (matched_items / total_items * 100) if total_items > 0 else 0
        mismatch_percentage = (mismatch_items / total_items * 100) if total_items > 0 else 0
        pending_percentage = (pending_items / total_items * 100) if total_items > 0 else 0
        
        assignment.verification_stats = {
            'total': total_items,
            'verified': verified_items,
            'matched': matched_items,
            'mismatch': mismatch_items,
            'pending': pending_items,
            'verified_percentage': verified_percentage,
            'matched_percentage': matched_percentage,
            'mismatch_percentage': mismatch_percentage,
            'pending_percentage': pending_percentage,
            'progress': verified_percentage
        }
        
        total_verified += verified_items
        total_matched += matched_items
        total_mismatch += mismatch_items
        total_pending += pending_items
    
    context = {
        'assignments': assignments,
        'total_verified': total_verified,
        'total_matched': total_matched,
        'total_mismatch': total_mismatch,
        'total_pending': total_pending,
        'today': timezone.now().date(),
    }
    return render(request, 'manager/verification_status.html', context)
@login_required
def manager_verification_details(request, assignment_id):
    """Manager view to see detailed verification status for an assignment"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(
        HardwareAssignment,
        id=assignment_id,
        assigned_by=request.user,
        actual_return_date__isnull=True
    )
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    verification_details = []
    all_verified = True
    
    for item in items:
        # Get asset number
        expected_asset = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
        
        try:
            # FIX: Use asset_entry instead of serial_entry
            asset_entry = item.asset_entry
            # Check if entered value matches asset number
            is_match = (asset_entry.entered_asset_number == expected_asset)
            
            verification_details.append({
                'item': item,
                'asset_entry': asset_entry,
                'is_match': is_match,
                'entered_asset': asset_entry.entered_asset_number,
                'expected_asset': expected_asset,
                'hardware_type': item.hardware.hardware_type.name,
                'model': item.hardware.model_name,
                'entered_by': asset_entry.entered_by,
                'entered_at': asset_entry.entered_at,
                'verified_at': asset_entry.verified_at,
                'verified_by': asset_entry.verified_by
            })
            if not is_match:
                all_verified = False
        except HardwareAssetEntry.DoesNotExist:
            verification_details.append({
                'item': item,
                'asset_entry': None,
                'is_match': False,
                'entered_asset': None,
                'expected_asset': expected_asset,
                'hardware_type': item.hardware.hardware_type.name,
                'model': item.hardware.model_name,
                'entered_by': None,
                'entered_at': None,
                'verified_at': None,
                'verified_by': None
            })
            all_verified = False
    
    total_items = len(verification_details)
    verified_count = sum(1 for d in verification_details if d['verified_at'] is not None)
    mismatch_count = sum(1 for d in verification_details if d['entered_asset'] and not d['is_match'])
    pending_count = sum(1 for d in verification_details if not d['entered_asset'])
    
    context = {
        'assignment': assignment,
        'verification_details': verification_details,
        'all_verified': all_verified,
        'total_items': total_items,
        'verified_count': verified_count,
        'mismatch_count': mismatch_count,
        'pending_count': pending_count,
        'verified_percentage': (verified_count / total_items * 100) if total_items > 0 else 0,
    }
    return render(request, 'manager/verification_details.html', context)


# ============== EMPLOYEE VIEWS ==============

@login_required
def employee_dashboard(request):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    # Get branch location with multiple fallbacks
    branch_location = getattr(request.user, 'branch_location', None)
    
    # If branch_location is None or empty, try to get from manager
    if not branch_location and request.user.manager:
        branch_location = getattr(request.user.manager, 'branch_location', None)
    
    # Final fallback
    if not branch_location:
        branch_location = 'Head Office'
    
    current_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    current_assignments_count = current_assignments.count()
    completed_assignments_count = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=False
    ).count()
    
    current_exam_city = None
    current_exam_center_name = None
    if current_assignments.exists():
        first_assignment = current_assignments.first()
        current_exam_city = first_assignment.exam_city
        current_exam_center_name = getattr(first_assignment, 'exam_center_name', None)
    
    pending_asset_count = 0
    for assignment in current_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        assignment.total_items = items.count()
        assignment.entered_asset_count = 0
        assignment.verified_asset_count = 0
        assignment.pending_asset_count = 0
        
        for item in items:
            if hasattr(item, 'asset_entry'):
                assignment.entered_asset_count += 1
                if item.asset_entry.verified:
                    assignment.verified_asset_count += 1
            else:
                assignment.pending_asset_count += 1
        
        pending_asset_count += assignment.pending_asset_count
    
    employee_name = request.user.get_full_name() or request.user.username
    
    context = {
        'current_assignments': current_assignments,
        'current_assignments_count': current_assignments_count,
        'completed_assignments_count': completed_assignments_count,
        'pending_asset_count': pending_asset_count,
        'current_exam_city': current_exam_city,
        'current_exam_center_name': current_exam_center_name,
        'employee_name': employee_name,
        'today': timezone.now().date(),
        'current_time': timezone.now(),
        'branch_location': branch_location,
        'user': request.user,
    }
    return render(request, 'employee/dashboard.html', context)

@login_required
def view_my_assignments(request):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignments = HardwareAssignment.objects.filter(employee=request.user).order_by('-assigned_date')
    
    for assignment in assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        assignment.total_items = items.count()
        assignment.pending_asset_count = 0  
        assignment.entered_asset_count = 0  
        assignment.verified_asset_count = 0  
        
        for item in items:
            if hasattr(item, 'asset_entry'):
                assignment.entered_asset_count += 1
                if item.asset_entry.verified:
                    assignment.verified_asset_count += 1
            else:
                assignment.pending_asset_count += 1
    
    context = {
        'assignments': assignments,
    }
    return render(request, 'employee/view_my_assignments.html', context)


@login_required
def my_assignment_details(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    for item in items:
        # Get asset number (primary) and serial number (reference)
        item.asset_number = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
        item.serial_number_display = item.hardware.serial_number
        
        # FIX: Use asset_entry instead of serial_entry
        item.has_asset_entry = hasattr(item, 'asset_entry')
        if item.has_asset_entry:
            # The entered value is the Asset Number (stored in asset_entry)
            item.entered_asset = item.asset_entry.entered_asset_number
            item.is_verified = item.asset_entry.verified
        else:
            item.entered_asset = None
            item.is_verified = False
    
    total_items = items.count()
    pending_count = sum(1 for item in items if not hasattr(item, 'asset_entry'))
    entered_count = total_items - pending_count
    verified_count = sum(1 for item in items if hasattr(item, 'asset_entry') and item.asset_entry.verified)
    
    # Get exam center name
    exam_center_name = getattr(assignment, 'exam_center_name', None)
    
    context = {
        'assignment': assignment,
        'items': items,
        'total_items': total_items,
        'pending_count': pending_count,
        'entered_count': entered_count,
        'verified_count': verified_count,
        'exam_center_name': exam_center_name,
    }
    return render(request, 'employee/my_assignment_details.html', context)
    


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import HardwareAssignment, HardwareAssignmentItem, HardwareType, HardwareAssetEntry
@login_required
def export_assignment_excel(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    # Check if there are any items
    if not items.exists():
        messages.warning(request, 'No hardware items found in this assignment to export.')
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    # Get all hardware types from the items
    hardware_type_ids = items.values_list('hardware__hardware_type', flat=True).distinct()
    hardware_types = HardwareType.objects.filter(id__in=hardware_type_ids).order_by('name')
    
    if not hardware_types.exists():
        messages.warning(request, 'No hardware types found in this assignment.')
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    # Group items by hardware type
    items_by_type = {}
    max_items_per_type = 0
    
    for hw_type in hardware_types:
        type_items = [item for item in items if item.hardware.hardware_type and item.hardware.hardware_type.id == hw_type.id]
        items_by_type[hw_type.id] = type_items
        if len(type_items) > max_items_per_type:
            max_items_per_type = len(type_items)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # ==============================================
    # FILENAME: exam center name_exam city_project name_employee name.xlsx
    # ==============================================
    
    # Get exam center name
    exam_center_name = getattr(assignment, 'exam_center_name', 'Not Specified')
    if not exam_center_name or exam_center_name == '':
        exam_center_name = 'NoCenter'
    
    # Get exam city
    exam_city = assignment.exam_city or 'NoCity'
    
    # Get project name
    project_name = assignment.project.project_name if assignment.project else 'NoProject'
    
    # Get employee name
    employee_name = assignment.employee.get_full_name() or assignment.employee.username
    
    # Clean up strings for filename (remove special characters, replace spaces with underscores)
    import re
    def clean_filename(text):
        # Replace spaces with underscores
        text = str(text).replace(' ', '_')
        # Remove special characters except underscores and hyphens
        text = re.sub(r'[^a-zA-Z0-9_\-]', '', text)
        return text
    
    exam_center_name_clean = clean_filename(exam_center_name)
    exam_city_clean = clean_filename(exam_city)
    project_name_clean = clean_filename(project_name)
    employee_name_clean = clean_filename(employee_name)
    
    # Build filename
    filename = f"{exam_center_name_clean}_{exam_city_clean}_{project_name_clean}_{employee_name_clean}.xlsx"
    
    assignment_id_str = str(assignment.assignment_id).replace('-', '')[:8]
    ws.title = f"Assign_{assignment_id_str}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E04D00", end_color="E04D00", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Calculate total columns (4 per hardware type: Type Name, Asset, Entered, Status)
    total_hw_types = hardware_types.count()
    total_columns = total_hw_types * 4
    
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws['A1'] = f"ASSIGNMENT DETAILS - {assignment.assignment_id}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_alignment
    
    # Assignment info with Exam Center included
    info_start = 3
    info_data = [
        ['Project:', assignment.project.project_name if assignment.project else 'N/A'],
        ['Exam City:', assignment.exam_city or 'Not specified'],
        ['Exam Center:', getattr(assignment, 'exam_center_name', 'Not specified') or 'Not specified'],
        ['Employee:', assignment.employee.get_full_name() or assignment.employee.username],
        ['Assigned Date:', assignment.assigned_date.strftime('%d %b %Y')],
        ['Expected Return:', assignment.expected_return_date.strftime('%d %b %Y')],
    ]
    
    for i, (label, value) in enumerate(info_data, start=info_start):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    
    # Headers - Each hardware type gets 4 columns
    header_row = info_start + len(info_data) + 2
    
    # Main header row with Hardware Type as merged cells (4 columns each)
    current_col = 1
    for hw_type in hardware_types:
        end_col = current_col + 3
        ws.merge_cells(start_row=header_row, start_column=current_col, end_row=header_row, end_column=end_col)
        hw_cell = ws.cell(row=header_row, column=current_col, value=f"{hw_type.name}")
        hw_cell.font = header_font
        hw_cell.fill = header_fill
        hw_cell.alignment = center_alignment
        hw_cell.border = border
        current_col = end_col + 1
    
    # Sub-headers: Asset Number, Entered Asset, Status, Verified By
    sub_header_row = header_row + 1
    current_col = 1
    for hw_type in hardware_types:
        ws.merge_cells(start_row=sub_header_row, start_column=current_col, end_row=sub_header_row, end_column=current_col + 3)
        type_label = ws.cell(row=sub_header_row, column=current_col, value=f"{hw_type.name} Details")
        type_label.font = Font(bold=True, color="FFFFFF", size=9)
        type_label.fill = subheader_fill
        type_label.alignment = center_alignment
        type_label.border = border
        current_col += 4
    
    # Data headers row
    data_header_row = sub_header_row + 1
    current_col = 1
    for hw_type in hardware_types:
        # Asset Number
        asset_cell = ws.cell(row=data_header_row, column=current_col, value="Asset Number")
        asset_cell.font = Font(bold=True, color="FFFFFF", size=8)
        asset_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        asset_cell.alignment = center_alignment
        asset_cell.border = border
        
        # Entered Asset
        entered_cell = ws.cell(row=data_header_row, column=current_col + 1, value="Entered Asset")
        entered_cell.font = Font(bold=True, color="FFFFFF", size=8)
        entered_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        entered_cell.alignment = center_alignment
        entered_cell.border = border
        
        # Verification Status
        status_cell = ws.cell(row=data_header_row, column=current_col + 2, value="Status")
        status_cell.font = Font(bold=True, color="FFFFFF", size=8)
        status_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        status_cell.alignment = center_alignment
        status_cell.border = border
        
        # Verified By
        verified_by_cell = ws.cell(row=data_header_row, column=current_col + 3, value="Verified By")
        verified_by_cell.font = Font(bold=True, color="FFFFFF", size=8)
        verified_by_cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        verified_by_cell.alignment = center_alignment
        verified_by_cell.border = border
        
        current_col += 4
    
    # Data rows
    data_start_row = data_header_row + 1
    
    # Track stats
    total_items = items.count()
    verified_count = 0
    pending_count = 0
    not_entered_count = 0
    mismatch_count = 0
    
    for row_offset in range(max_items_per_type):
        current_row = data_start_row + row_offset
        current_col = 1
        
        for hw_type in hardware_types:
            type_items = items_by_type.get(hw_type.id, [])
            
            if row_offset < len(type_items):
                item = type_items[row_offset]
                
                # Get asset number (expected)
                asset_number = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
                
                # Get entered asset number from asset_entry
                entered_asset = 'Not Entered'
                verification_status = 'Not Entered'
                verified_by = '—'
                is_verified = False
                is_match = False
                
                try:
                    asset_entry = item.asset_entry
                    entered_asset = asset_entry.entered_asset_number
                    is_verified = asset_entry.verified
                    is_match = (entered_asset == asset_number)
                    
                    if is_verified:
                        verification_status = 'Verified ✓'
                        verified_count += 1
                        if asset_entry.verified_by:
                            verified_by = asset_entry.verified_by.get_full_name() or asset_entry.verified_by.username
                    elif is_match:
                        verification_status = 'Matched - Pending'
                        pending_count += 1
                    else:
                        verification_status = 'Mismatch ✗'
                        mismatch_count += 1
                except HardwareAssetEntry.DoesNotExist:
                    not_entered_count += 1
                
                # Asset Number cell
                asset_cell = ws.cell(row=current_row, column=current_col, value=asset_number)
                asset_cell.border = border
                asset_cell.alignment = center_alignment
                asset_cell.fill = PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid")
                asset_cell.font = Font(color="0d6efd", bold=True)
                
                # Entered Asset cell
                entered_cell = ws.cell(row=current_row, column=current_col + 1, value=entered_asset)
                entered_cell.border = border
                entered_cell.alignment = center_alignment
                
                if is_verified:
                    entered_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    entered_cell.font = Font(color="006100", bold=True)
                elif entered_asset != 'Not Entered' and is_match:
                    entered_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    entered_cell.font = Font(color="9C5700", bold=True)
                elif entered_asset != 'Not Entered' and not is_match:
                    entered_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    entered_cell.font = Font(color="9C0006", bold=True)
                else:
                    entered_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    entered_cell.font = Font(color="666666", italic=True)
                
                # Status cell
                status_cell = ws.cell(row=current_row, column=current_col + 2, value=verification_status)
                status_cell.border = border
                status_cell.alignment = center_alignment
                status_cell.font = Font(bold=True)
                
                if verification_status == 'Verified ✓':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100", bold=True)
                elif verification_status == 'Matched - Pending':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    status_cell.font = Font(color="9C5700", bold=True)
                elif verification_status == 'Mismatch ✗':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006", bold=True)
                else:
                    status_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    status_cell.font = Font(color="666666", italic=True)
                
                # Verified By cell
                verified_by_cell = ws.cell(row=current_row, column=current_col + 3, value=verified_by)
                verified_by_cell.border = border
                verified_by_cell.alignment = center_alignment
                
                if is_verified:
                    verified_by_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    verified_by_cell.font = Font(color="006100")
                
            else:
                # Empty cells for rows without items
                for i in range(4):
                    empty_cell = ws.cell(row=current_row, column=current_col + i, value="—")
                    empty_cell.border = border
                    empty_cell.alignment = center_alignment
                    empty_cell.font = Font(color="999999", italic=True)
            
            current_col += 4
    
    # Summary section
    summary_row = data_start_row + max_items_per_type + 3
    
    # Summary header
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=total_columns)
    summary_header = ws.cell(row=summary_row, column=1, value="📊 ASSIGNMENT SUMMARY")
    summary_header.font = Font(bold=True, size=12, color="FFFFFF")
    summary_header.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    summary_header.alignment = center_alignment
    
    # Summary stats
    summary_start = summary_row + 2
    
    stats_data = [
        ['Total Hardware Items:', str(total_items)],
        ['✅ Verified:', str(verified_count)],
        ['⏳ Matched - Pending:', str(pending_count)],
        ['❌ Mismatch:', str(mismatch_count)],
        ['📝 Not Entered:', str(not_entered_count)],
        ['', ''],
        ['📈 Completion Rate:', f"{round((verified_count / total_items * 100) if total_items > 0 else 0, 1)}%"],
    ]
    
    for idx, (label, value) in enumerate(stats_data):
        row = summary_start + idx
        
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=True)
        label_cell.border = border
        
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.border = border
        
        if 'Verified' in label and verified_count > 0:
            value_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            value_cell.font = Font(color="006100", bold=True)
        elif 'Matched' in label and pending_count > 0:
            value_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            value_cell.font = Font(color="9C5700", bold=True)
        elif 'Mismatch' in label and mismatch_count > 0:
            value_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            value_cell.font = Font(color="9C0006", bold=True)
        elif 'Not Entered' in label and not_entered_count > 0:
            value_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            value_cell.font = Font(color="666666", bold=True)
    
    # Asset Number Summary
    asset_summary_row = summary_start + len(stats_data) + 2
    ws.merge_cells(start_row=asset_summary_row, start_column=1, end_row=asset_summary_row, end_column=2)
    asset_summary_header = ws.cell(row=asset_summary_row, column=1, value="📋 ASSET NUMBER SUMMARY")
    asset_summary_header.font = Font(bold=True, size=11, color="FFFFFF")
    asset_summary_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    asset_summary_header.alignment = center_alignment
    
    # Collect all asset numbers with status
    asset_summary = []
    for item in items:
        asset_number = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
        status = "Not Entered"
        
        try:
            asset_entry = item.asset_entry
            if asset_entry.verified:
                status = "Verified ✓"
            elif asset_entry.entered_asset_number == asset_number:
                status = "Matched - Pending"
            else:
                status = "Mismatch ✗"
        except HardwareAssetEntry.DoesNotExist:
            pass
        
        asset_summary.append({
            'asset': asset_number,
            'hardware_type': item.hardware.hardware_type.name if item.hardware.hardware_type else 'Unknown',
            'status': status
        })
    
    row = asset_summary_row + 2
    for idx, asset_info in enumerate(asset_summary, 1):
        ws.cell(row=row, column=1, value=f"{idx}. {asset_info['asset']}").border = border
        ws.cell(row=row, column=2, value=f"{asset_info['hardware_type']} - {asset_info['status']}").border = border
        
        # Color code the status
        status_cell = ws.cell(row=row, column=2)
        if 'Verified' in asset_info['status']:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100", bold=True)
        elif 'Matched' in asset_info['status']:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C5700", bold=True)
        elif 'Mismatch' in asset_info['status']:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            status_cell.font = Font(color="666666", italic=True)
        
        row += 1
    
    # Set column widths
    for col in range(1, total_columns + 1):
        column_letter = get_column_letter(col)
        if col % 4 == 1:  # Asset Number column
            ws.column_dimensions[column_letter].width = 16
        elif col % 4 == 2:  # Entered Asset column
            ws.column_dimensions[column_letter].width = 16
        elif col % 4 == 3:  # Status column
            ws.column_dimensions[column_letter].width = 18
        else:  # Verified By column
            ws.column_dimensions[column_letter].width = 18
    
    # Info columns (A, B) for summary
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    
    # Set row heights
    ws.row_dimensions[header_row].height = 30
    ws.row_dimensions[sub_header_row].height = 25
    ws.row_dimensions[data_header_row].height = 25
    
    for row in range(data_start_row, data_start_row + max_items_per_type):
        ws.row_dimensions[row].height = 24
    
    # Freeze header row
    ws.freeze_panes = ws.cell(row=data_header_row + 1, column=1)
    
    # Prepare response with the new filename format
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
# ============== EMPLOYEE SERIAL NUMBER ENTRY ==============


@login_required
def enter_serial_numbers(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    
    if assignment.actual_return_date:
        messages.error(request, 'This assignment has already been returned!')
        return redirect('employee_dashboard')
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    verified_items = []
    pending_items = []
    
    for item in items:
        item.asset_number = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
        item.serial_number = item.hardware.serial_number
        
        # FIX: Use 'asset_entry' related_name
        if hasattr(item, 'asset_entry') and item.asset_entry.entered_asset_number:
            item.existing_asset = item.asset_entry.entered_asset_number
            verified_items.append(item)
        else:
            item.existing_asset = ''
            pending_items.append(item)
    
    if request.method == 'POST':
        success_count = 0
        error_count = 0
        
        for item in items:
            asset_number = request.POST.get(f'asset_{item.id}')
            if asset_number and asset_number.strip():
                # FIX: Use 'asset_entry' related_name
                if hasattr(item, 'asset_entry'):
                    asset_entry = item.asset_entry
                    asset_entry.entered_asset_number = asset_number.strip()
                    asset_entry.entered_by = request.user
                    asset_entry.save()
                    success_count += 1
                else:
                    HardwareAssetEntry.objects.create(
                        hardware_item=item,  # Using the field name 'hardware_item'
                        entered_asset_number=asset_number.strip(),
                        entered_by=request.user
                    )
                    success_count += 1
            else:
                error_count += 1
        
        if success_count > 0:
            messages.success(request, f'Successfully submitted {success_count} asset number(s)!')
        if error_count > 0:
            messages.warning(request, f'{error_count} item(s) were not submitted (empty asset number)')
        
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    context = {
        'assignment': assignment,
        'items': items,
        'verified_items': verified_items,
        'pending_items': pending_items,
        'total_items': items.count(),
        'editing': bool(verified_items)
    }
    return render(request, 'employee/enter_serial_numbers.html', context)

@login_required
def edit_serial_numbers(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    
    if assignment.actual_return_date:
        messages.error(request, 'This assignment has already been returned!')
        return redirect('employee_dashboard')
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    has_verified_serials = False
    for item in items:
        if hasattr(item, 'serial_entry') and item.serial_entry.verified:
            has_verified_serials = True
            break
    
    if has_verified_serials:
        messages.error(request, 'Cannot edit serial numbers that have been verified by manager!')
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    if request.method == 'POST':
        success_count = 0
        error_count = 0
        
        for item in items:
            serial_number = request.POST.get(f'serial_{item.id}')
            if serial_number and serial_number.strip():
                if hasattr(item, 'serial_entry'):
                    serial_entry = item.serial_entry
                    serial_entry.serial_number = serial_number.strip()
                    serial_entry.save()
                    success_count += 1
                else:
                    HardwareSerialEntry.objects.create(
                        assignment_item=item,
                        serial_number=serial_number.strip(),
                        entered_by=request.user
                    )
                    success_count += 1
            else:
                error_count += 1
        
        if success_count > 0:
            messages.success(request, f'Successfully updated {success_count} serial number(s)!')
        if error_count > 0:
            messages.warning(request, f'{error_count} item(s) were not updated (empty serial number)')
        
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    for item in items:
        if hasattr(item, 'serial_entry'):
            item.existing_serial = item.serial_entry.serial_number
        else:
            item.existing_serial = ''
    
    context = {
        'assignment': assignment,
        'items': items,
        'editing': True,
    }
    return render(request, 'employee/enter_serial_numbers.html', context)

# ============== API VIEWS ==============

@csrf_exempt
@login_required
def api_get_hardware_by_type(request):
    if request.method == 'GET':
        hardware_type_id = request.GET.get('type_id')
        
        if request.user.user_type == 'manager':
            hardware_items = Hardware.objects.filter(
                hardware_type_id=hardware_type_id,
                status='available',
                created_by=request.user
            ).values('id', 'serial_number', 'model_name', 'brand')
        else:
            return JsonResponse({'error': 'Unauthorized'}, status=403)
        
        return JsonResponse(list(hardware_items), safe=False)

@csrf_exempt
@login_required
def api_get_assignment_details(request, assignment_id):
    if request.method == 'GET':
        assignment = get_object_or_404(HardwareAssignment, id=assignment_id)
        
        if request.user.user_type == 'employee' and assignment.employee != request.user:
            return JsonResponse({'error': 'Unauthorized'}, status=403)
        
        items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
        
        data = {
            'assignment_id': str(assignment.assignment_id),
            'project': {
                'id': assignment.project.id,
                'name': assignment.project.project_name,
                'location': assignment.project.location,
            },
            'employee': assignment.employee.username,
            'assigned_by': assignment.assigned_by.username,
            'assigned_date': assignment.assigned_date.strftime('%Y-%m-%d'),
            'expected_return_date': assignment.expected_return_date.strftime('%Y-%m-%d'),
            'actual_return_date': assignment.actual_return_date.strftime('%Y-%m-%d') if assignment.actual_return_date else None,
            'notes': assignment.notes,
            'hardware_items': [
                {
                    'id': item.id,
                    'hardware_id': item.hardware.id,
                    'type': item.hardware.hardware_type.name,
                    'model': item.hardware.model_name,
                    'serial_number': item.hardware.serial_number,
                    'brand': item.hardware.brand,
                    'has_serial_entry': hasattr(item, 'serial_entry'),
                    'entered_serial': item.serial_entry.serial_number if hasattr(item, 'serial_entry') else None,
                    'verified': item.serial_entry.verified if hasattr(item, 'serial_entry') else False,
                }
                for item in items
            ]
        }
        
        return JsonResponse(data)

@csrf_exempt
@login_required
def api_check_serial_exists(request):
    if request.method == 'GET':
        serial_number = request.GET.get('serial_number')
        
        if not serial_number:
            return JsonResponse({'exists': False})
        
        exists = HardwareSerialEntry.objects.filter(serial_number=serial_number).exists()
        return JsonResponse({'exists': exists})

# ============== UTILITY VIEWS ==============

@login_required
def profile(request):
    context = {'user': request.user}
    return render(request, 'profile.html', context)

@login_required
def update_profile(request):
    if request.method == 'POST':
        user = request.user
        user.email = request.POST.get('email')
        user.phone = request.POST.get('phone')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.save()
        
        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')
    
    context = {'user': request.user}
    return render(request, 'update_profile.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.contrib.auth.hashers import make_password
from .models import PasswordResetOTP
import uuid

User = get_user_model()

def forgot_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        
        try:
            user = User.objects.get(username=username)
            
            otp_obj = PasswordResetOTP.generate_otp(user)
            
            subject = 'Password Reset OTP - Eduquity Hardware Management'
            html_message = render_to_string('auth/password_reset_email.html', {
                'user': user,
                'otp': otp_obj.otp,
                'expiry_minutes': 5,
                'token': otp_obj.token,
            })
            
            send_mail(
                subject=subject,
                message=f'Your OTP for password reset is: {otp_obj.otp}. This OTP is valid for 5 minutes.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            
            messages.success(request, f'OTP has been sent to your registered email address ({user.email}). Please check your inbox.')
            return redirect('verify_otp', token=otp_obj.token)
            
        except User.DoesNotExist:
            messages.error(request, 'No account found with this Employee ID.')
        except Exception as e:
            messages.error(request, f'Failed to send OTP. Error: {str(e)}')
    
    return render(request, 'auth/forgot_password.html')

def verify_otp(request, token):
    try:
        otp_obj = PasswordResetOTP.objects.get(token=token, is_used=False)
        
        if otp_obj.is_expired():
            messages.error(request, 'OTP has expired. Please request a new one.')
            return redirect('forgot_password')
        
        if request.method == 'POST':
            entered_otp = request.POST.get('otp')
            
            if entered_otp == otp_obj.otp:
                otp_obj.is_used = True
                otp_obj.save()
                return redirect('reset_password', token=token)
            else:
                messages.error(request, 'Invalid OTP. Please try again.')
        
        return render(request, 'auth/verify_otp.html', {
            'token': token,
            'email': otp_obj.user.email[:3] + '*****' + otp_obj.user.email[otp_obj.user.email.find('@'):]
        })
    
    except PasswordResetOTP.DoesNotExist:
        messages.error(request, 'Invalid or expired OTP link.')
        return redirect('forgot_password')
def reset_password(request, token):
    try:
        otp_obj = PasswordResetOTP.objects.get(token=token)
        
        if otp_obj.is_used == False:
            messages.error(request, 'Please verify OTP first.')
            return redirect('verify_otp', token=token)
        
        if request.method == 'POST':
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')
            
            if password != confirm_password:
                messages.error(request, 'Passwords do not match!')
                return render(request, 'auth/reset_password.html', {'token': token})
            
            if len(password) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return render(request, 'auth/reset_password.html', {'token': token})
            
            user = otp_obj.user
            user.set_password(password)
            
            if user.user_type == 'employee' and user.is_first_login:
                user.is_first_login = False
            
            user.save()
            
            try:
                subject = 'Password Reset Successful - Eduquity Hardware Management'
                
                try:
                    html_message = render_to_string('auth/password_reset_success_email.html', {
                        'user': user,
                        'now': timezone.now(),
                    })
                except TemplateDoesNotExist:
                    html_message = f'''
                    <!DOCTYPE html>
                    <html>
                    <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                        <h2>Password Reset Successful</h2>
                        <p>Hello {user.username},</p>
                        <p>Your password has been successfully reset for the Eduquity Hardware Management System.</p>
                        <p>If you did not request this password reset, please contact your system administrator immediately.</p>
                        <p><strong>Eduquity Hardware Management Team</strong></p>
                    </body>
                    </html>
                    '''
                
                send_mail(
                    subject=subject,
                    message=f'Your password has been reset successfully on {timezone.now().strftime("%B %d, %Y at %I:%M %p")}.',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    html_message=html_message,
                    fail_silently=False,
                )
                
            except Exception as e:
                print(f"Failed to send password reset email: {str(e)}")
            
            messages.success(request, 'Password reset successfully! You can now login with your new password.')
            return redirect('login')
        
        return render(request, 'auth/reset_password.html', {'token': token})
    
    except PasswordResetOTP.DoesNotExist:
        messages.error(request, 'Invalid reset link.')
        return redirect('forgot_password')
    

@login_required
def my_hardware(request):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    active_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    completed_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=False
    ).order_by('-assigned_date')[:5]  
    
    total_items = 0
    verified_count = 0
    pending_count = 0
    matched_count = 0
    mismatch_count = 0
    
    for assignment in active_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        assignment.hardware_count = items.count()
        total_items += assignment.hardware_count
        
        assignment.verified_count = 0
        assignment.matched_count = 0
        assignment.mismatch_count = 0
        assignment.pending_count = 0
        assignment.items_list = []  
        
        for item in items:
            # Get asset number (expected) and serial number (reference)
            expected_asset = item.hardware.asset_number if item.hardware.asset_number else 'N/A'
            
            hardware_data = {
                'id': item.id,
                'hardware_type': item.hardware.hardware_type.name,
                'model': item.hardware.model_name,
                'brand': item.hardware.brand,
                'expected_asset': expected_asset,
                'serial_number': item.hardware.serial_number,  # Display this instead
                'status': item.hardware.status,
            }
            
            try:
                # FIX: Use asset_entry instead of serial_entry
                asset_entry = item.asset_entry
                hardware_data['entered_asset'] = asset_entry.entered_asset_number
                hardware_data['verified'] = asset_entry.verified
                hardware_data['verified_by'] = asset_entry.verified_by
                hardware_data['verified_at'] = asset_entry.verified_at
                hardware_data['entered_at'] = asset_entry.entered_at
                
                if asset_entry.verified:
                    assignment.verified_count += 1
                    verified_count += 1
                else:
                    if asset_entry.entered_asset_number == expected_asset:
                        assignment.matched_count += 1
                        matched_count += 1
                    else:
                        assignment.mismatch_count += 1
                        mismatch_count += 1
                        
            except HardwareAssetEntry.DoesNotExist:
                hardware_data['entered_asset'] = None
                hardware_data['verified'] = False
                assignment.pending_count += 1
                pending_count += 1
            
            assignment.items_list.append(hardware_data)
    
    for assignment in completed_assignments:
        assignment.hardware_count = HardwareAssignmentItem.objects.filter(assignment=assignment).count()
    
    context = {
        'assignments': active_assignments,
        'completed_assignments': completed_assignments,
        'total_items': total_items,
        'verified_count': verified_count,
        'pending_count': pending_count,
        'matched_count': matched_count,
        'mismatch_count': mismatch_count,
        'active_assignments': active_assignments.count(),
    }
    return render(request, 'employee/my_hardware.html', context)
    
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

@login_required
def export_my_hardware_excel(request):
    """Export employee's hardware data to Excel with filename: examcity_employeename_date.xlsx"""
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    active_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    exam_city = "NoCity"
    if active_assignments.exists():
        exam_city = active_assignments.first().exam_city.replace(" ", "_")
    
    employee_name = request.user.get_full_name() or request.user.username
    employee_name = employee_name.replace(" ", "_")
    
    current_date = datetime.now().strftime("%Y%m%d")
    
    filename = f"{exam_city}_{employee_name}_{current_date}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Hardware Report"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    success_fill = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")  
    info_fill = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")      
    warning_fill = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")    
    danger_fill = PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid")     
    
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    
    headers = [
        'Assignment ID', 'Project', 'Exam City', 'Assigned Date', 'Expected Return',
        'Hardware Type', 'Model', 'Brand', 'Assigned Serial', 'Entered Serial',
        'Entry Status', 'Verification Status', 'Verified By', 'Verified On'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row_num = 2
    total_items = 0
    
    for assignment in active_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        
        for item in items:
            total_items += 1
            
            serial_entry = None
            entered_serial = "Not entered"
            entry_status = "Pending"
            verification_status = "Pending"
            verified_by = "-"
            verified_on = "-"
            
            try:
                serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                entered_serial = serial_entry.serial_number
                
                if serial_entry.verified:
                    verification_status = "Verified"
                    verified_by = serial_entry.verified_by.get_full_name() or serial_entry.verified_by.username if serial_entry.verified_by else "-"
                    verified_on = serial_entry.verified_at.strftime("%d/%m/%Y %H:%M") if serial_entry.verified_at else "-"
                    
                    if serial_entry.serial_number == item.hardware.serial_number:
                        entry_status = "Verified - Correct"
                    else:
                        entry_status = "Verified - Mismatch"
                else:
                    if serial_entry.serial_number == item.hardware.serial_number:
                        entry_status = "Matched - Pending"
                        verification_status = "Pending Verification"
                    else:
                        entry_status = "Mismatch"
                        verification_status = "Not Verified"
                        
            except HardwareSerialEntry.DoesNotExist:
                entry_status = "Not Entered"
                verification_status = "Pending"
            
            if serial_entry and serial_entry.verified:
                row_fill = success_fill
            elif serial_entry and serial_entry.serial_number == item.hardware.serial_number:
                row_fill = info_fill
            elif serial_entry and serial_entry.serial_number != item.hardware.serial_number:
                row_fill = danger_fill
            else:
                row_fill = warning_fill
            
            row_data = [
                str(assignment.assignment_id)[:8],
                assignment.project.project_name,
                assignment.exam_city,
                assignment.assigned_date.strftime("%d/%m/%Y"),
                assignment.expected_return_date.strftime("%d/%m/%Y"),
                item.hardware.hardware_type.name,
                item.hardware.model_name,
                item.hardware.brand or "-",
                item.hardware.serial_number,
                entered_serial,
                entry_status,
                verification_status,
                verified_by,
                verified_on
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="center")
                
                if col_num in [4, 5, 14]:  
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1
    
    row_num += 2
    summary_row = row_num
    
    verified_count = 0
    matched_count = 0
    mismatch_count = 0
    pending_count = 0
    
    for assignment in active_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        for item in items:
            try:
                serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                if serial_entry.verified:
                    verified_count += 1
                elif serial_entry.serial_number == item.hardware.serial_number:
                    matched_count += 1
                else:
                    mismatch_count += 1
            except HardwareSerialEntry.DoesNotExist:
                pending_count += 1
    
    summary_headers = ['Summary Statistics', 'Value']
    summary_data = [
        ['Total Items', total_items],
        ['Verified Items', verified_count],
        ['Matched Items (Pending)', matched_count],
        ['Mismatched Items', mismatch_count],
        ['Pending Entry', pending_count],
        ['Generated On', datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
        ['Employee Name', request.user.get_full_name() or request.user.username],
        ['Employee Email', request.user.email],
        ['Exam City', assignment.exam_city if active_assignments.exists() else 'Not Assigned'],
    ]
    
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=summary_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for i, (label, value) in enumerate(summary_data, summary_row + 1):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.border = border
        label_cell.font = Font(bold=True)
        
        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.border = border
        value_cell.alignment = Alignment(horizontal="center")
    
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, row_num):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def request_hardware_transfer(request):
    """Employee requests hardware transfer from another employee (multiple items)"""
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    employees = CustomUser.objects.filter(
        user_type='employee',
        manager=request.user.manager,
        is_active=True
    ).exclude(id=request.user.id)
    
    for emp in employees:
        active_assignment = HardwareAssignment.objects.filter(
            employee=emp,
            actual_return_date__isnull=True
        ).first()
        emp.current_exam_city = active_assignment.exam_city if active_assignment else 'Not Assigned'
    
    active_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=True
    ).prefetch_related('hardwareassignmentitem_set__hardware__hardware_type')
    
    my_hardware = []
    for assignment in active_assignments:
        for item in assignment.hardwareassignmentitem_set.all():
            hardware = item.hardware
            if hardware.status == 'in_use':
                my_hardware.append({
                    'id': hardware.id,
                    'hardware_type': hardware.hardware_type.name,
                    'asset_number': hardware.asset_number if hardware.asset_number else 'N/A',  # Added asset number
                    'serial_number': hardware.serial_number,
                    'model_name': hardware.model_name,
                    'brand': hardware.brand,
                    'current_exam_city': assignment.exam_city or 'Unknown',
                    'assignment_id': assignment.id
                })
    
    if request.method == 'POST':
        hardware_ids = request.POST.getlist('hardware_ids')  
        to_employee_id = request.POST.get('to_employee_id')
        transfer_type = request.POST.get('transfer_type')
        reason = request.POST.get('reason')
        
        # Handle date formats - DD-MM-YYYY to YYYY-MM-DD
        expected_arrival_date = None
        expected_return_date = None
        
        if request.POST.get('expected_arrival_date'):
            try:
                arrival_date_str = request.POST.get('expected_arrival_date')
                # Check if date is in DD-MM-YYYY format
                if '-' in arrival_date_str:
                    parts = arrival_date_str.split('-')
                    if len(parts) == 3 and len(parts[0]) == 2 and len(parts[1]) == 2 and len(parts[2]) == 4:
                        # DD-MM-YYYY format
                        expected_arrival_date = datetime.strptime(arrival_date_str, '%d-%m-%Y').date()
                    else:
                        expected_arrival_date = datetime.strptime(arrival_date_str, '%Y-%m-%d').date()
                else:
                    expected_arrival_date = datetime.strptime(arrival_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                expected_arrival_date = None
        
        if transfer_type == 'temporary' and request.POST.get('expected_return_date'):
            try:
                return_date_str = request.POST.get('expected_return_date')
                # Check if date is in DD-MM-YYYY format
                if '-' in return_date_str:
                    parts = return_date_str.split('-')
                    if len(parts) == 3 and len(parts[0]) == 2 and len(parts[1]) == 2 and len(parts[2]) == 4:
                        # DD-MM-YYYY format
                        expected_return_date = datetime.strptime(return_date_str, '%d-%m-%Y').date()
                    else:
                        expected_return_date = datetime.strptime(return_date_str, '%Y-%m-%d').date()
                else:
                    expected_return_date = datetime.strptime(return_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                expected_return_date = None
        
        requester_notes = request.POST.get('requester_notes')
        delivery_method = request.POST.get('delivery_method')
        
        if not hardware_ids:
            messages.error(request, 'Please select at least one hardware item to transfer!')
            return redirect('request_hardware_transfer')
        
        try:
            to_employee = CustomUser.objects.get(id=to_employee_id)
            
            valid_hardware = []
            errors = []
            
            for hw_id in hardware_ids:
                try:
                    hardware = Hardware.objects.get(id=hw_id)
                    is_assigned = HardwareAssignmentItem.objects.filter(
                        hardware=hardware,
                        assignment__employee=request.user,
                        assignment__actual_return_date__isnull=True
                    ).exists()
                    
                    if not is_assigned:
                        errors.append(f"Hardware '{hardware.serial_number}' is not assigned to you!")
                    else:
                        valid_hardware.append(hardware)
                except Hardware.DoesNotExist:
                    errors.append(f"Hardware with ID {hw_id} not found!")
            
            if errors:
                messages.error(request, '\n'.join(errors))
                return redirect('request_hardware_transfer')
            
            if not valid_hardware:
                messages.error(request, 'No valid hardware items selected!')
                return redirect('request_hardware_transfer')
            
            first_hardware = valid_hardware[0]
            assignment_item = HardwareAssignmentItem.objects.filter(
                hardware=first_hardware,
                assignment__employee=request.user,
                assignment__actual_return_date__isnull=True
            ).first()
            from_exam_city = assignment_item.assignment.exam_city or 'Unknown'
            from_project = assignment_item.assignment.project
            
            to_employee_assignment = HardwareAssignment.objects.filter(
                employee=to_employee,
                actual_return_date__isnull=True
            ).first()
            to_exam_city = to_employee_assignment.exam_city if to_employee_assignment else 'Unknown'
            to_project = to_employee_assignment.project if to_employee_assignment else None
            
            transfer = EmployeeHardwareTransfer.objects.create(
                from_employee=request.user,
                to_employee=to_employee,
                from_exam_city=from_exam_city,
                to_exam_city=to_exam_city,
                from_project=from_project,
                to_project=to_project,
                transfer_type=transfer_type,
                reason=reason,
                expected_arrival_date=expected_arrival_date,
                expected_return_date=expected_return_date,
                requester_notes=requester_notes,
                delivery_method=delivery_method,
                status='requested',
                created_by=request.user
            )
            
            for hardware in valid_hardware:
                TransferItem.objects.create(
                    transfer=transfer,
                    hardware=hardware,
                    status='pending'
                )
            
            TransferHistory.objects.create(
                transfer=transfer,
                action=f"Transfer request created with {len(valid_hardware)} item(s)",
                status='requested',
                notes=f"Reason: {reason}",
                updated_by=request.user
            )
            
            manager = request.user.manager
            if manager:
                TransferNotification.objects.create(
                    transfer=transfer,
                    recipient=manager,
                    message=f"Transfer request #{transfer.transfer_id} from {request.user.get_full_name() or request.user.username} to {to_employee.get_full_name() or to_employee.username} with {len(valid_hardware)} item(s)"
                )
            
            messages.success(request, f'Transfer request #{transfer.transfer_id} created successfully with {len(valid_hardware)} hardware item(s)!')
            return redirect('my_transfers')
            
        except Exception as e:
            messages.error(request, f'Error creating transfer: {str(e)}')
            return redirect('request_hardware_transfer')
    
    context = {
        'employees': employees,
        'my_hardware': my_hardware,
        'today': timezone.now().date(),
    }
    return render(request, 'employee/request_transfer.html', context)
    
@login_required
def update_transfer_status_employee(request, transfer_id):
    """Employee updates transfer status with proper hardware transfer"""
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    transfer = get_object_or_404(
        EmployeeHardwareTransfer,
        id=transfer_id
    )
    
    if request.user not in [transfer.from_employee, transfer.to_employee]:
        messages.error(request, 'You are not authorized for this action.')
        return redirect('my_transfers')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        if action == 'initiate' and request.user == transfer.from_employee and transfer.status == 'approved_by_manager':
            for item in transfer.transfer_items.all():
                item.status = 'in_transit'
                item.transfer_date = timezone.now().date()
                item.save()
                
                item.hardware.status = 'maintenance'
                item.hardware.save()
            
            transfer.status = 'in_transit'
            transfer.transfer_date = timezone.now().date()
            transfer.save()
            
            messages.success(request, f'Transfer initiated! {transfer.transfer_items.count()} hardware items marked as in transit.')
            
        elif action == 'receive' and request.user == transfer.to_employee and transfer.status == 'in_transit':
            transferred_count = 0
            
            for item in transfer.transfer_items.all():
                item.status = 'received'
                item.received_date = timezone.now().date()
                item.condition_after = notes
                item.save()
                
                sender_item = HardwareAssignmentItem.objects.filter(
                    hardware=item.hardware,
                    assignment__employee=transfer.from_employee,
                    assignment__actual_return_date__isnull=True
                ).first()
                
                if sender_item:
                    sender_assignment = sender_item.assignment
                    
                    sender_item.delete()
                    
                    remaining_items = HardwareAssignmentItem.objects.filter(
                        assignment=sender_assignment
                    ).count()
                    
                    if remaining_items == 0:
                        sender_assignment.actual_return_date = timezone.now().date()
                        sender_assignment.save()
                
                receiver_assignment = HardwareAssignment.objects.filter(
                    employee=transfer.to_employee,
                    project=transfer.to_project,
                    actual_return_date__isnull=True
                ).first()
                
                if not receiver_assignment:
                    sender_assignment_item = HardwareAssignmentItem.objects.filter(
                        hardware=item.hardware,
                        assignment__employee=transfer.from_employee,
                        assignment__actual_return_date__isnull=True
                    ).first()
                    
                    project = transfer.to_project
                    if not project and sender_assignment_item:
                        project = sender_assignment_item.assignment.project
                    
                    receiver_assignment = HardwareAssignment.objects.create(
                        employee=transfer.to_employee,
                        project=project,
                        exam_city=transfer.to_exam_city,
                        assigned_by=transfer.approved_by or request.user.manager,
                        expected_return_date=transfer.expected_return_date if transfer.transfer_type == 'temporary' else (timezone.now().date() + timezone.timedelta(days=30)),
                        notes=f"Hardware transferred via Transfer ID: {transfer.transfer_id}"
                    )
                
                existing_item = HardwareAssignmentItem.objects.filter(
                    assignment=receiver_assignment,
                    hardware=item.hardware
                ).first()
                
                if existing_item:
                    existing_item.quantity += 1
                    existing_item.condition_at_assignment = notes
                    existing_item.save()
                else:
                    HardwareAssignmentItem.objects.create(
                        assignment=receiver_assignment,
                        hardware=item.hardware,
                        quantity=1,
                        condition_at_assignment=notes
                    )
                
                sender_serial_entry = HardwareSerialEntry.objects.filter(
                    assignment_item=sender_item
                ).first() if sender_item else None
                
                if sender_serial_entry:
                    receiver_item = HardwareAssignmentItem.objects.filter(
                        assignment=receiver_assignment,
                        hardware=item.hardware
                    ).first()
                    
                    if receiver_item:
                        existing_serial = HardwareSerialEntry.objects.filter(
                            assignment_item=receiver_item
                        ).first()
                        
                        if not existing_serial:
                            HardwareSerialEntry.objects.create(
                                assignment_item=receiver_item,
                                serial_number=sender_serial_entry.serial_number,
                                entered_by=sender_serial_entry.entered_by,
                                entered_at=sender_serial_entry.entered_at,
                                verified=sender_serial_entry.verified,
                                verified_by=sender_serial_entry.verified_by,
                                verified_at=sender_serial_entry.verified_at
                            )
                
                transferred_count += 1
                
                item.hardware.status = 'in_use'
                item.hardware.save()
            
            transfer.status = 'received_by_receiver'
            transfer.actual_arrival_date = timezone.now().date()
            transfer.save()
            
            messages.success(request, f'{transferred_count} hardware item(s) transferred successfully to {transfer.to_employee.get_full_name() or transfer.to_employee.username}!')
            
        elif action == 'complete' and request.user == transfer.from_employee and transfer.status == 'received_by_receiver':
            if transfer.transfer_type == 'permanent':
                transfer.status = 'completed'
                transfer.save()
                messages.success(request, 'Permanent transfer completed successfully!')
            
            elif transfer.transfer_type == 'temporary':
                returned_count = 0
                
                for item in transfer.transfer_items.all():
                    item.status = 'returned'
                    item.return_date = timezone.now().date()
                    item.save()
                    
                    receiver_item = HardwareAssignmentItem.objects.filter(
                        hardware=item.hardware,
                        assignment__employee=transfer.to_employee,
                        assignment__actual_return_date__isnull=True
                    ).first()
                    
                    if receiver_item:
                        receiver_assignment = receiver_item.assignment
                        
                        receiver_item.delete()
                        
                        remaining_items = HardwareAssignmentItem.objects.filter(
                            assignment=receiver_assignment
                        ).count()
                        
                        if remaining_items == 0:
                            receiver_assignment.actual_return_date = timezone.now().date()
                            receiver_assignment.save()
                    
                    sender_assignment = HardwareAssignment.objects.filter(
                        employee=transfer.from_employee,
                        project=transfer.from_project,
                        actual_return_date__isnull=True
                    ).first()
                    
                    if not sender_assignment:
                        sender_assignment = HardwareAssignment.objects.create(
                            employee=transfer.from_employee,
                            project=transfer.from_project,
                            exam_city=transfer.from_exam_city,
                            assigned_by=transfer.approved_by,
                            expected_return_date=timezone.now().date() + timezone.timedelta(days=30),
                            notes=f"Hardware returned from {transfer.to_employee.get_full_name() or transfer.to_employee.username} - Transfer ID: {transfer.transfer_id}"
                        )
                    
                    existing_item = HardwareAssignmentItem.objects.filter(
                        assignment=sender_assignment,
                        hardware=item.hardware
                    ).first()
                    
                    if existing_item:
                        existing_item.quantity += 1
                        existing_item.save()
                    else:
                        HardwareAssignmentItem.objects.create(
                            assignment=sender_assignment,
                            hardware=item.hardware,
                            quantity=1,
                            condition_at_assignment=notes
                        )
                    
                    receiver_serial_entry = HardwareSerialEntry.objects.filter(
                        assignment_item=receiver_item
                    ).first() if receiver_item else None
                    
                    if receiver_serial_entry:
                        sender_item = HardwareAssignmentItem.objects.filter(
                            assignment=sender_assignment,
                            hardware=item.hardware
                        ).first()
                        
                        if sender_item:
                            existing_serial = HardwareSerialEntry.objects.filter(
                                assignment_item=sender_item
                            ).first()
                            
                            if not existing_serial:
                                HardwareSerialEntry.objects.create(
                                    assignment_item=sender_item,
                                    serial_number=receiver_serial_entry.serial_number,
                                    entered_by=receiver_serial_entry.entered_by,
                                    entered_at=receiver_serial_entry.entered_at,
                                    verified=receiver_serial_entry.verified,
                                    verified_by=receiver_serial_entry.verified_by,
                                    verified_at=receiver_serial_entry.verified_at
                                )
                    
                    returned_count += 1
                    item.hardware.status = 'in_use'
                    item.hardware.save()
                
                transfer.status = 'completed'
                transfer.return_date = timezone.now().date()
                transfer.save()
                
                messages.success(request, f'Temporary transfer completed! {returned_count} hardware item(s) returned to original owner.')
        
        return redirect('transfer_tracking', transfer_id=transfer.id)
    
    return redirect('my_transfers')
@login_required
def manager_transfer_requests(request):
    """Manager views all transfer requests"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    transfers = EmployeeHardwareTransfer.objects.filter(
        from_employee__manager=request.user
    ).order_by('-requested_date')
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        transfers = transfers.filter(status=status_filter)
    
    total = transfers.count()
    pending = transfers.filter(status='requested').count()
    approved = transfers.filter(status='approved_by_manager').count()
    in_transit = transfers.filter(status='in_transit').count()
    received = transfers.filter(status='received_by_receiver').count()  # ADD THIS
    completed = transfers.filter(status='completed').count()
    rejected = transfers.filter(status='rejected').count()
    
    context = {
        'transfers': transfers,
        'total': total,
        'pending': pending,
        'approved': approved,
        'in_transit': in_transit,
        'received': received,  
        'completed': completed,
        'rejected': rejected,
        'status_filter': status_filter,
    }
    return render(request, 'manager/transfer_requests.html', context)

@login_required
def approve_transfer_request(request, transfer_id):
    """Manager approves transfer request"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    transfer = get_object_or_404(
        EmployeeHardwareTransfer, 
        id=transfer_id, 
        from_employee__manager=request.user,
        status='requested'
    )
    
    if request.method == 'POST':
        manager_notes = request.POST.get('manager_notes', '')
        
        transfer.status = 'approved_by_manager'
        transfer.approved_by = request.user
        transfer.approved_date = timezone.now()
        transfer.manager_notes = manager_notes
        transfer.save()
        
        TransferHistory.objects.create(
            transfer=transfer,
            status='approved_by_manager',
            notes=f"Approved by {request.user.get_full_name()}. Notes: {manager_notes}",
            updated_by=request.user
        )
        
        TransferNotification.objects.create(
            transfer=transfer,
            recipient=transfer.from_employee,
            message=f"Your transfer request #{transfer.transfer_id} has been approved by manager. Please prepare the hardware for transfer."
        )
        TransferNotification.objects.create(
            transfer=transfer,
            recipient=transfer.to_employee,
            message=f"Transfer request #{transfer.transfer_id} has been approved. Please prepare to receive the hardware."
        )
        
        messages.success(request, f'Transfer #{transfer.transfer_id} approved successfully!')
        return redirect('manager_transfer_requests')
    
    context = {'transfer': transfer}
    return render(request, 'manager/approve_transfer.html', context)


@login_required
def reject_transfer_request(request, transfer_id):
    """Manager rejects transfer request"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    transfer = get_object_or_404(
        EmployeeHardwareTransfer, 
        id=transfer_id, 
        from_employee__manager=request.user,
        status='requested'
    )
    
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        
        transfer.status = 'rejected'
        transfer.manager_notes = rejection_reason
        transfer.save()
        
        TransferHistory.objects.create(
            transfer=transfer,
            status='rejected',
            notes=f"Rejected by {request.user.get_full_name()}. Reason: {rejection_reason}",
            updated_by=request.user
        )
        
        TransferNotification.objects.create(
            transfer=transfer,
            recipient=transfer.from_employee,
            message=f"Your transfer request #{transfer.transfer_id} has been rejected. Reason: {rejection_reason}"
        )
        
        messages.warning(request, f'Transfer #{transfer.transfer_id} rejected.')
        return redirect('manager_transfer_requests')
    
    context = {'transfer': transfer}
    return render(request, 'manager/reject_transfer.html', context)


@login_required
def transfer_details(request, transfer_id):
    """View transfer details"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    transfer = get_object_or_404(EmployeeHardwareTransfer, id=transfer_id)
    
    transfer_items = transfer.transfer_items.all().select_related('hardware__hardware_type')
    
    hardware_details = []
    for item in transfer_items:
        asset_entry = None
        assignment_item = HardwareAssignmentItem.objects.filter(
            hardware=item.hardware,
            assignment__employee=transfer.to_employee if transfer.status == 'received' else transfer.from_employee,
            assignment__actual_return_date__isnull=True
        ).first()
        
        if assignment_item:
            try:
                # FIX: Use 'hardware_item' instead of 'assignment_item'
                asset_entry = HardwareAssetEntry.objects.get(hardware_item=assignment_item)
            except HardwareAssetEntry.DoesNotExist:
                pass
        
        hardware_details.append({
            'item': item,
            'has_asset_entry': asset_entry is not None,
            'asset_entry': asset_entry
        })
    
    total_items = transfer_items.count()
    pending_items = transfer_items.filter(status='pending').count()
    in_transit_items = transfer_items.filter(status='in_transit').count()
    received_items = transfer_items.filter(status='received').count()
    returned_items = transfer_items.filter(status='returned').count()
    
    history = []
    if hasattr(transfer, 'history'):
        history = transfer.history.all()
    
    context = {
        'transfer': transfer,
        'transfer_items': transfer_items,
        'hardware_details': hardware_details,
        'total_items': total_items,
        'pending_items': pending_items,
        'in_transit_items': in_transit_items,
        'received_items': received_items,
        'returned_items': returned_items,
        'history': history,
    }
    return render(request, 'manager/transfer_details.html', context)

@login_required
def my_transfers(request):
    """Employee views their transfers (sent and received)"""
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    sent_transfers = EmployeeHardwareTransfer.objects.filter(from_employee=request.user).order_by('-requested_date')
    received_transfers = EmployeeHardwareTransfer.objects.filter(to_employee=request.user).order_by('-requested_date')
    
    pending_sent = sent_transfers.filter(status='requested').count()
    pending_received = received_transfers.filter(status='in_transit').count()
    
    context = {
        'sent_transfers': sent_transfers,
        'received_transfers': received_transfers,
        'pending_sent': pending_sent,
        'pending_received': pending_received,
    }
    return render(request, 'employee/my_transfers.html', context)



@login_required
def transfer_tracking(request, transfer_id):
    """Track transfer details and status"""
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    transfer = get_object_or_404(
        EmployeeHardwareTransfer,
        id=transfer_id
    )
    
    if request.user not in [transfer.from_employee, transfer.to_employee, request.user.manager]:
        messages.error(request, 'You are not authorized to view this transfer.')
        return redirect('my_transfers')
    
    context = {
        'transfer': transfer,
        'can_act': (request.user == transfer.from_employee or request.user == transfer.to_employee),
    }
    return render(request, 'employee/transfer_tracking.html', context)

import json
import time
from datetime import datetime, timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Q, Count, Sum
from .models import (
    Hardware, HardwareType, HardwareAssignment, HardwareAssignmentItem, 
    HardwareSerialEntry, CustomUser, Project, EmployeeHardwareTransfer, TransferItem
)


@login_required
def chatbot_view(request):
    """Chatbot interface for hardware tracking"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    from .models import ChatbotConversation
    
    conversations = ChatbotConversation.objects.filter(
        user=request.user
    )[:20]
    
    total_hardware = Hardware.objects.filter(created_by=request.user).count()
    active_assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True
    ).count()
    
    verified_items = HardwareSerialEntry.objects.filter(
        assignment_item__assignment__assigned_by=request.user,
        verified=True
    ).count()
    
    overdue_items = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True,
        expected_return_date__lt=timezone.now().date()
    ).count()
    
    employee_count = CustomUser.objects.filter(
        user_type='employee',
        manager=request.user
    ).count()
    
    pending_transfers = EmployeeHardwareTransfer.objects.filter(
        from_employee__manager=request.user,
        status='requested'
    ).count()
    
    in_transit_transfers = EmployeeHardwareTransfer.objects.filter(
        from_employee__manager=request.user,
        status='in_transit'
    ).count()
    
    context = {
        'conversations': conversations,
        'total_hardware': total_hardware,
        'active_assignments': active_assignments,
        'verified_items': verified_items,
        'overdue_items': overdue_items,
        'employee_count': employee_count,
        'pending_transfers': pending_transfers,
        'in_transit_transfers': in_transit_transfers,
    }
    return render(request, 'manager/chatbot.html', context)


@login_required
@csrf_exempt
def chatbot_api(request):
    """Chatbot API endpoint"""
    if request.user.user_type != 'manager':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        user_message = data.get('message', '').strip().lower()
        
        if not user_message:
            return JsonResponse({'error': 'Message is empty'}, status=400)
        
        response_data = process_chatbot_message(user_message, request.user)
        
        from .models import ChatbotConversation
        ChatbotConversation.objects.create(
            user=request.user,
            message=user_message,
            response=response_data['message'],
            intent=response_data.get('intent', 'chat')
        )
        
        return JsonResponse(response_data)
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def chatbot_stats_api(request):
    """Get real-time stats for chatbot dashboard"""
    if request.user.user_type != 'manager':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    total_hardware = Hardware.objects.filter(created_by=request.user).count()
    active_assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True
    ).count()
    
    verified_items = HardwareSerialEntry.objects.filter(
        assignment_item__assignment__assigned_by=request.user,
        verified=True
    ).count()
    
    overdue_items = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True,
        expected_return_date__lt=timezone.now().date()
    ).count()
    
    employee_count = CustomUser.objects.filter(
        user_type='employee',
        manager=request.user
    ).count()
    
    pending_transfers = EmployeeHardwareTransfer.objects.filter(
        from_employee__manager=request.user,
        status='requested'
    ).count()
    
    in_transit_transfers = EmployeeHardwareTransfer.objects.filter(
        from_employee__manager=request.user,
        status='in_transit'
    ).count()
    
    return JsonResponse({
        'total_hardware': total_hardware,
        'active_assignments': active_assignments,
        'verified_items': verified_items,
        'overdue_items': overdue_items,
        'employee_count': employee_count,
        'pending_transfers': pending_transfers,
        'in_transit_transfers': in_transit_transfers,
    })


def process_chatbot_message(message, user):
    """Process user message and return appropriate response"""
    
    if any(word in message for word in ['hi', 'hello', 'hey', 'greetings', 'good morning', 'good evening', 'gm', 'gd']):
        return {
            'message': f"""👋 Hello {user.get_full_name() or user.username}! I'm your Hardware Management Assistant.

I can help you track hardware, employees, assignments, and transfers.

🔍 **Try these commands:**
• "Total hardware" - See inventory summary
• "All employees" - List all employees
• "Active assignments" - View current assignments  
• "Verification status" - Check verification progress
• "Transfer status" - Check pending transfers
• "Help" - See all commands

What would you like to know today?""",
            'intent': 'greeting'
        }
    
    elif 'help' in message or 'what can you do' in message or 'commands' in message:
        return {
            'message': """🤖 **Hardware Bot Help - Available Commands**

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 **Hardware Management:**
• "Total hardware" - View inventory summary
• "Laptop status" - Check laptop availability
• "Hardware by type" - See distribution by type
• "Available hardware" - List available items

👥 **Employee Management:**
• "All employees" - List all employees
• "Active employees" - Employees with active assignments
• "Employee [name]" - Get employee details
• "Employee hardware [name]" - See employee's hardware

📋 **Assignment Tracking:**
• "Active assignments" - View current assignments
• "Overdue assignments" - List overdue returns
• "Due soon" - View assignments due in 3 days
• "Assignment by employee" - Group by employee

🔐 **Verification:**
• "Verification status" - Check verification progress
• "Pending verification" - View pending items
• "Verified items" - Count of verified hardware

🔄 **Transfer Management:**
• "Transfer status" - Check pending transfers
• "In transit transfers" - Hardware being transferred
• "Completed transfers" - Completed transfers

💡 **Tip:** Be specific with your questions for best results!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━""",
            'intent': 'help'
        }
    
    # ============ HARDWARE STATISTICS ============
    elif 'total hardware' in message or 'hardware count' in message or 'inventory summary' in message:
        total = Hardware.objects.filter(created_by=user).count()
        available = Hardware.objects.filter(created_by=user, status='available').count()
        assigned = Hardware.objects.filter(created_by=user, status='assigned').count()
        in_use = Hardware.objects.filter(created_by=user, status='in_use').count()
        maintenance = Hardware.objects.filter(created_by=user, status='maintenance').count()
        retired = Hardware.objects.filter(created_by=user, status='retired').count()
        
        return {
            'message': f"""📊 **Hardware Inventory Summary**

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📦 **Total Hardware:** {total}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ Available: **{available}**
📋 Assigned: **{assigned}**
💻 In Use: **{in_use}**
🔧 Maintenance: **{maintenance}**
📦 Retired: **{retired}**
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📈 **Utilization Rate:** {round(((assigned + in_use) / total * 100) if total > 0 else 0, 1)}%

Need more details? Ask for "hardware by type" or specific hardware status!""",
            'intent': 'hardware_summary'
        }
    
    elif 'hardware by type' in message or 'distribution' in message or 'type wise' in message:
        hardware_by_type = Hardware.objects.filter(
            created_by=user
        ).values('hardware_type__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        if hardware_by_type:
            response = "📊 **Hardware Distribution by Type**\n\n"
            response += "━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            for item in hardware_by_type:
                bar_length = min(20, item['count'] * 2)
                bar = "█" * bar_length + "░" * (20 - bar_length)
                response += f"• **{item['hardware_type__name']}**: {item['count']}\n"
                response += f"  {bar} {round((item['count']/hardware_by_type.aggregate(Sum('count'))['count__sum'])*100, 1)}%\n\n"
            
            total = hardware_by_type.aggregate(Sum('count'))['count__sum']
            response += f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            response += f"📦 **Total:** {total} items"
        else:
            response = "No hardware data available. Please add hardware inventory first."
        
        return {'message': response, 'intent': 'hardware_by_type'}
    
    elif 'laptop' in message or 'laptops' in message:
        laptops = Hardware.objects.filter(
            created_by=user,
            hardware_type__name__icontains='laptop'
        )
        total = laptops.count()
        available = laptops.filter(status='available').count()
        assigned = laptops.filter(status='assigned').count()
        in_use = laptops.filter(status='in_use').count()
        maintenance = laptops.filter(status='maintenance').count()
        
        return {
            'message': f"""💻 **Laptops Status Report**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📦 **Total Laptops:** {total}
━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ Available: **{available}** (Ready to assign)
📋 Assigned: **{assigned}** (Pending pickup)
💻 In Use: **{in_use}** (Currently active)
🔧 Maintenance: **{maintenance}** (Under repair)
━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 **Availability Rate:** {round((available / total * 100) if total > 0 else 0, 1)}%

Want laptop details by model? Ask for "laptop models"!""",
            'intent': 'hardware_type'
        }
    
    elif 'firewall' in message:
        firewalls = Hardware.objects.filter(
            created_by=user,
            hardware_type__name__icontains='firewall'
        )
        total = firewalls.count()
        available = firewalls.filter(status='available').count()
        assigned = firewalls.filter(status='assigned').count()
        
        return {
            'message': f"""🛡️ **Firewall Status Report**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📦 **Total Firewalls:** {total}
━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ Available: **{available}**
📋 Assigned: **{assigned}**
🎯 Utilization: **{round(((total - available) / total * 100) if total > 0 else 0, 1)}%**
━━━━━━━━━━━━━━━━━━━━━━━━━━━

Need specific firewall details?""",
            'intent': 'hardware_type'
        }
    
    elif 'available hardware' in message or 'available items' in message:
        available_items = Hardware.objects.filter(
            created_by=user,
            status='available'
        ).select_related('hardware_type')[:20]
        
        count = available_items.count()
        
        if count > 0:
            response = f"✅ **Available Hardware ({count} items)**\n\n"
            for item in available_items:
                response += f"• **{item.hardware_type.name}** - {item.serial_number}\n"
                response += f"  {item.model_name} | {item.brand or 'No brand'}\n\n"
            if count >= 20:
                response += "Showing first 20 items. Use filters for more specific results."
        else:
            response = "✅ **No available hardware at the moment.** All hardware is either assigned or in use."
        
        return {'message': response, 'intent': 'available_hardware'}
    
    elif 'all employees' in message or 'list employees' in message or 'show employees' in message:
        employees = CustomUser.objects.filter(
            user_type='employee',
            manager=user
        ).order_by('first_name')[:15]
        
        total_employees = employees.count()
        
        if total_employees > 0:
            response = f"👥 **Employee List ({total_employees} employees)**\n\n"
            for emp in employees:
                active_count = HardwareAssignment.objects.filter(
                    employee=emp,
                    actual_return_date__isnull=True
                ).count()
                response += f"• **{emp.get_full_name() or emp.username}**\n"
                response += f"  📧 {emp.email} | 📞 {emp.phone or 'No phone'}\n"
                response += f"  💻 Active Hardware: {active_count} items\n\n"
        else:
            response = "👥 No employees found. Please create employee accounts first."
        
        return {'message': response, 'intent': 'all_employees'}
    
    elif 'active employees' in message:
        active_employees = CustomUser.objects.filter(
            user_type='employee',
            manager=user
        ).count()
        
        employees_with_assignments = CustomUser.objects.filter(
            user_type='employee',
            manager=user,
            hardware_assignments__actual_return_date__isnull=True
        ).distinct().count()
        
        response = f"""👥 **Active Employees Report**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📊 **Total Employees:** {active_employees}
━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ **With Active Assignments:** {employees_with_assignments}
⏳ **Idle Employees:** {active_employees - employees_with_assignments}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

📈 **Activity Rate:** {round((employees_with_assignments / active_employees * 100) if active_employees > 0 else 0, 1)}%

Want to see specific employee details? Ask "Employee [name]!" """
        
        return {'message': response, 'intent': 'active_employees'}
    
    elif 'employee' in message and ('details' in message or 'info' in message or 'show' in message):
        words = message.split()
        emp_name = None
        for i, word in enumerate(words):
            if word not in ['employee', 'details', 'info', 'show', 'for', 'about']:
                emp_name = word
                break
        
        if emp_name:
            employees = CustomUser.objects.filter(
                user_type='employee',
                manager=user,
                username__icontains=emp_name
            ) | CustomUser.objects.filter(
                user_type='employee',
                manager=user,
                first_name__icontains=emp_name
            ) | CustomUser.objects.filter(
                user_type='employee',
                manager=user,
                last_name__icontains=emp_name
            )
            
            if employees.exists():
                emp = employees.first()
                
                assignments = HardwareAssignment.objects.filter(
                    employee=emp,
                    actual_return_date__isnull=True
                )
                
                hardware_count = 0
                hardware_list = []
                for assignment in assignments:
                    for item in assignment.hardwareassignmentitem_set.all():
                        hardware_count += 1
                        hardware_list.append(f"  • {item.hardware.hardware_type.name} - {item.hardware.serial_number}")
                
                completed_count = HardwareAssignment.objects.filter(
                    employee=emp,
                    actual_return_date__isnull=False
                ).count()
                
                response = f"""👤 **Employee Details: {emp.get_full_name() or emp.username}**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📧 **Email:** {emp.email}
📞 **Phone:** {emp.phone or 'Not provided'}
📅 **Joined:** {emp.date_joined.strftime('%d %b %Y')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

💻 **Current Hardware:** {hardware_count} item(s)
✅ **Completed Assignments:** {completed_count}
📊 **Status:** {'Active' if hardware_count > 0 else 'Idle'}

"""
                if hardware_list:
                    response += "\n".join(hardware_list[:5])
                    if len(hardware_list) > 5:
                        response += f"\n  ... and {len(hardware_list) - 5} more"

                response += "\n\nNeed to assign hardware? Use the assignment form!"
            else:
                response = f"❌ No employee found with name '{emp_name}'. Please check the name and try again."
        else:
            response = "Please provide an employee name. Example: 'Employee details John'"
        
        return {'message': response, 'intent': 'employee_details'}
    
    elif 'employee hardware' in message or 'hardware assigned to' in message:
        words = message.split()
        emp_name = None
        for word in words:
            if word not in ['employee', 'hardware', 'assigned', 'to', 'show', 'for', 'of']:
                emp_name = word
                break
        
        if emp_name:
            employees = CustomUser.objects.filter(
                user_type='employee',
                manager=user,
                username__icontains=emp_name
            ) | CustomUser.objects.filter(
                user_type='employee',
                manager=user,
                first_name__icontains=emp_name
            )
            
            if employees.exists():
                emp = employees.first()
                assignments = HardwareAssignment.objects.filter(
                    employee=emp,
                    actual_return_date__isnull=True
                )
                
                hardware_list = []
                for assignment in assignments:
                    for item in assignment.hardwareassignmentitem_set.all():
                        verification_status = "Not Verified"
                        try:
                            serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                            if serial_entry.verified:
                                verification_status = "✅ Verified"
                            else:
                                verification_status = "⏳ Pending"
                        except HardwareSerialEntry.DoesNotExist:
                            verification_status = "📝 Not Entered"
                        
                        hardware_list.append({
                            'type': item.hardware.hardware_type.name,
                            'serial': item.hardware.serial_number,
                            'model': item.hardware.model_name,
                            'exam_city': assignment.exam_city,
                            'status': verification_status
                        })
                
                if hardware_list:
                    response = f"💻 **Hardware Assigned to {emp.get_full_name() or emp.username}**\n\n"
                    for hw in hardware_list:
                        response += f"• **{hw['type']}** - {hw['serial']}\n"
                        response += f"  Model: {hw['model']} | City: {hw['exam_city']}\n"
                        response += f"  Verification: {hw['status']}\n\n"
                else:
                    response = f"✅ No active hardware assigned to {emp.get_full_name() or emp.username}."
            else:
                response = f"❌ No employee found with name '{emp_name}'."
        else:
            response = "Please provide an employee name. Example: 'Employee hardware John'"
        
        return {'message': response, 'intent': 'employee_hardware'}
    
    elif 'active assignments' in message or 'current assignments' in message:
        today = timezone.now().date()
        active = HardwareAssignment.objects.filter(
            assigned_by=user,
            actual_return_date__isnull=True
        )
        
        total_active = active.count()
        overdue = active.filter(expected_return_date__lt=today).count()
        due_soon = active.filter(
            expected_return_date__gte=today,
            expected_return_date__lte=today + timedelta(days=3)
        ).count()
        
        city_stats = active.values('exam_city').annotate(count=Count('id')).order_by('-count')[:5]
        
        response = f"""📋 **Active Assignments Report**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📊 **Total Active:** {total_active}
━━━━━━━━━━━━━━━━━━━━━━━━━━━
⚠️ **Overdue:** {overdue}
⏰ **Due Soon (3 days):** {due_soon}
✅ **On Track:** {total_active - overdue - due_soon}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

📍 **By Exam City:**
"""
        for city in city_stats:
            response += f"• {city['exam_city']}: {city['count']} assignments\n"
        
        response += f"\n📈 **Completion Rate:** {round(((total_active - overdue) / total_active * 100) if total_active > 0 else 100, 1)}% on track"
        
        if overdue > 0:
            response += f"\n\n⚠️ **Action Required:** {overdue} assignment(s) are overdue. Please check the assignments page."
        
        return {'message': response, 'intent': 'assignments'}
    
    elif 'overdue' in message or 'show overdue' in message:
        today = timezone.now().date()
        overdue_assignments = HardwareAssignment.objects.filter(
            assigned_by=user,
            actual_return_date__isnull=True,
            expected_return_date__lt=today
        ).select_related('employee', 'project')[:10]
        
        if overdue_assignments.exists():
            response = "⚠️ **Overdue Assignments:**\n\n"
            for assign in overdue_assignments:
                days_overdue = (today - assign.expected_return_date).days
                response += f"• **{assign.employee.get_full_name() or assign.employee.username}**\n"
                response += f"  Project: {assign.project.project_name}\n"
                response += f"  Exam City: {assign.exam_city}\n"
                response += f"  Expected: {assign.expected_return_date} ({days_overdue} days overdue)\n\n"
            if overdue_assignments.count() >= 10:
                response += "Showing first 10 results. Please check the assignments page for complete list."
            response += "\n📋 Use 'return assignment' to process returns."
        else:
            response = "✅ **No overdue assignments!** All hardware is on track."
        
        return {'message': response, 'intent': 'overdue'}
    
    elif 'due soon' in message:
        today = timezone.now().date()
        due_soon_assignments = HardwareAssignment.objects.filter(
            assigned_by=user,
            actual_return_date__isnull=True,
            expected_return_date__gte=today,
            expected_return_date__lte=today + timedelta(days=3)
        ).select_related('employee', 'project')[:10]
        
        if due_soon_assignments.exists():
            response = "⏰ **Assignments Due Soon (within 3 days):**\n\n"
            for assign in due_soon_assignments:
                days_left = (assign.expected_return_date - today).days
                response += f"• **{assign.employee.get_full_name() or assign.employee.username}**\n"
                response += f"  Project: {assign.project.project_name}\n"
                response += f"  Due: {assign.expected_return_date} ({days_left} days left)\n\n"
        else:
            response = "✅ **No assignments due soon!** All due dates are more than 3 days away."
        
        return {'message': response, 'intent': 'due_soon'}
    
    elif 'assignment by employee' in message or 'group by employee' in message:
        assignments = HardwareAssignment.objects.filter(
            assigned_by=user,
            actual_return_date__isnull=True
        ).values('employee__first_name', 'employee__last_name', 'employee__username').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        if assignments.exists():
            response = "📋 **Assignments by Employee**\n\n"
            for assign in assignments:
                name = assign.get('employee__first_name') or assign.get('employee__username')
                if assign.get('employee__last_name'):
                    name += f" {assign.get('employee__last_name')}"
                bar_length = min(20, assign['count'] * 2)
                bar = "█" * bar_length + "░" * (20 - bar_length)
                response += f"• **{name}**: {assign['count']} assignments\n"
                response += f"  {bar}\n\n"
        else:
            response = "No active assignments found."
        
        return {'message': response, 'intent': 'assignment_by_employee'}
    
    elif 'verification' in message or 'verified' in message or 'pending verification' in message:
        total_entries = HardwareSerialEntry.objects.filter(
            assignment_item__assignment__assigned_by=user
        ).count()
        
        verified_count = HardwareSerialEntry.objects.filter(
            assignment_item__assignment__assigned_by=user,
            verified=True
        ).count()
        
        pending_count = total_entries - verified_count
        
        matched = 0
        mismatch = 0
        for entry in HardwareSerialEntry.objects.filter(assignment_item__assignment__assigned_by=user, verified=False):
            if entry.serial_number == entry.assignment_item.hardware.serial_number:
                matched += 1
            else:
                mismatch += 1
        
        not_entered = HardwareAssignmentItem.objects.filter(
            assignment__assigned_by=user,
            assignment__actual_return_date__isnull=True
        ).exclude(
            id__in=HardwareSerialEntry.objects.values('assignment_item_id')
        ).count()
        
        completion_rate = round((verified_count / total_entries * 100) if total_entries > 0 else 0, 1)
        
        response = f"""🔐 **Verification Status Report**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📊 **Total Entries:** {total_entries}
━━━━━━━━━━━━━━━━━━━━━━━━━━━
✅ **Verified:** {verified_count}
🟦 **Matched (Pending):** {matched}
❌ **Mismatch:** {mismatch}
📝 **Not Entered:** {not_entered}
━━━━━━━━━━━━━━━━━━━━━━━━━━━
📈 **Completion Rate:** {completion_rate}%

"""
        
        if matched > 0:
            response += f"💡 {matched} item(s) are matched and ready for verification. Use 'verify all' to approve them."
        elif mismatch > 0:
            response += f"⚠️ {mismatch} item(s) have mismatched serials. Please contact employees to correct them."
        elif not_entered > 0:
            response += f"📝 {not_entered} item(s) need serial numbers to be entered."
        
        return {'message': response, 'intent': 'verification'}
    
    # ============ TRANSFER MANAGEMENT ============
    elif 'transfer status' in message or 'pending transfers' in message:
        pending = EmployeeHardwareTransfer.objects.filter(
            from_employee__manager=user,
            status='requested'
        ).count()
        
        approved = EmployeeHardwareTransfer.objects.filter(
            from_employee__manager=user,
            status='approved_by_manager'
        ).count()
        
        in_transit = EmployeeHardwareTransfer.objects.filter(
            from_employee__manager=user,
            status='in_transit'
        ).count()
        
        completed = EmployeeHardwareTransfer.objects.filter(
            from_employee__manager=user,
            status='completed'
        ).count()
        
        rejected = EmployeeHardwareTransfer.objects.filter(
            from_employee__manager=user,
            status='rejected'
        ).count()
        
        response = f"""🔄 **Hardware Transfer Status**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📊 **Total Transfers:** {pending + approved + in_transit + completed + rejected}
━━━━━━━━━━━━━━━━━━━━━━━━━━━
⏳ **Pending Approval:** {pending}
✅ **Approved:** {approved}
🚚 **In Transit:** {in_transit}
🎯 **Completed:** {completed}
❌ **Rejected:** {rejected}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

"""
        if pending > 0:
            response += f"⚠️ {pending} transfer(s) pending your approval. Use 'transfer-requests' to review them."
        elif in_transit > 0:
            response += f"🚚 {in_transit} transfer(s) are in transit. Track them in the transfers section."
        
        return {'message': response, 'intent': 'transfer_status'}
    
    elif 'in transit transfers' in message:
        transfers = EmployeeHardwareTransfer.objects.filter(
            from_employee__manager=user,
            status='in_transit'
        ).select_related('from_employee', 'to_employee', 'hardware_item')[:10]
        
        if transfers.exists():
            response = "🚚 **In Transit Transfers:**\n\n"
            for t in transfers:
                response += f"• **{t.hardware_item.hardware_type.name}** - {t.hardware_item.serial_number}\n"
                response += f"  From: {t.from_employee.get_full_name() or t.from_employee.username} ({t.from_exam_city})\n"
                response += f"  To: {t.to_employee.get_full_name() or t.to_employee.username} ({t.to_exam_city})\n"
                response += f"  Type: {'Temporary' if t.transfer_type == 'temporary' else 'Permanent'}\n\n"
        else:
            response = "🚚 No hardware transfers are currently in transit."
        
        return {'message': response, 'intent': 'in_transit_transfers'}
    
    # ============ SEARCH HARDWARE ============
    elif 'search' in message or 'find' in message:
        import re
        # Extract search term
        search_match = re.search(r'(?:search|find)(?:\s+for)?\s+([a-zA-Z0-9-]+)', message)
        
        if search_match:
            search_term = search_match.group(1)
        else:
            words = message.split()
            for word in words:
                if len(word) > 3 and word not in ['search', 'find', 'hardware', 'for', 'show', 'me']:
                    search_term = word
                    break
        
        if search_term:
            hardware = Hardware.objects.filter(
                Q(serial_number__icontains=search_term) |
                Q(model_name__icontains=search_term) |
                Q(hardware_type__name__icontains=search_term),
                created_by=user
            )[:10]
            
            if hardware.exists():
                response = "🔍 **Search Results:**\n\n"
                for hw in hardware:
                    # Find current status/location
                    location = "Available (Not assigned)"
                    assignment_item = HardwareAssignmentItem.objects.filter(
                        hardware=hw,
                        assignment__actual_return_date__isnull=True
                    ).first()
                    
                    if assignment_item:
                        location = f"Assigned to: {assignment_item.assignment.employee.get_full_name() or assignment_item.assignment.employee.username} at {assignment_item.assignment.exam_city}"
                    
                    response += f"• **{hw.hardware_type.name}** - {hw.serial_number}\n"
                    response += f"  Model: {hw.model_name} | Brand: {hw.brand or 'N/A'}\n"
                    response += f"  Status: {hw.get_status_display()} | 📍 {location}\n\n"
            else:
                response = f"❌ No hardware found matching '{search_term}'"
        else:
            response = "Please provide a search term. Example: 'Search for laptop' or 'Find LAP-1234'"
        
        return {'message': response, 'intent': 'search'}
    
    # ============ SPECIFIC HARDWARE BY SERIAL ============
    elif re.search(r'[A-Z0-9]{4,}', message.upper()):
        import re
        serial_match = re.search(r'([A-Z0-9]{4,})', message.upper())
        
        if serial_match:
            serial_number = serial_match.group(1)
            try:
                hardware = Hardware.objects.get(serial_number__iexact=serial_number, created_by=user)
                
                # Find current assignment
                assignment_item = HardwareAssignmentItem.objects.filter(
                    hardware=hardware,
                    assignment__actual_return_date__isnull=True
                ).first()
                
                response = f"""🔍 **Hardware Details Found**

━━━━━━━━━━━━━━━━━━━━━━━━━━━
📦 **Type:** {hardware.hardware_type.name}
🔢 **Serial:** {hardware.serial_number}
📱 **Model:** {hardware.model_name}
🏷️ **Brand:** {hardware.brand or 'Not specified'}
📊 **Status:** {hardware.get_status_display()}
━━━━━━━━━━━━━━━━━━━━━━━━━━━

"""
                
                if assignment_item:
                    response += f"""👤 **Currently with:** {assignment_item.assignment.employee.get_full_name() or assignment_item.assignment.employee.username}
📍 **Location:** {assignment_item.assignment.exam_city}
📅 **Assigned Since:** {assignment_item.assignment.assigned_date.strftime('%d %b %Y')}
"""
                else:
                    response += "✅ This hardware is available for assignment."
                
                return {'message': response, 'intent': 'hardware_search'}
                
            except Hardware.DoesNotExist:
                pass
    
    # ============ DEFAULT RESPONSE ============
    else:
        return {
            'message': """🤔 I'm not sure I understand. Here's what I can help with:

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📊 **Hardware Tracking**
• "Total hardware" - Inventory summary
• "Laptop status" - Laptop availability
• "Hardware by type" - Distribution report

👥 **Employee Tracking**
• "All employees" - List employees
• "Employee [name]" - Employee details
• "Employee hardware [name]" - Assigned hardware

📋 **Assignment Tracking**
• "Active assignments" - Current assignments
• "Overdue assignments" - Overdue items
• "Due soon" - Approaching deadlines

🔐 **Verification**
• "Verification status" - Check progress

🔄 **Transfers**
• "Transfer status" - Pending transfers

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Try one of these commands or ask "Help" for more options!""",
            'intent': 'unknown'
        }
